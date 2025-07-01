from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page


# Create your views here.

from .serializers import (
    TalhaoSerializer,
    PlantioSerializer,
    DefensivoSerializer,
    AplicacaoSerializer,
    ColheitaSerializer,
    VisitasSerializer,
    RegistroVisitasSerializer,
    StProtheusIntegrationSerializer,
    ColheitaPlantioExtratoAreaSerializer,
    ColheitaResumoSerializer
)

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.authentication import TokenAuthentication
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import AuthenticationFailed


from django.contrib.auth.models import User
from rest_framework.renderers import JSONRenderer
from rest_framework import generics


import json
from django.http import JsonResponse, HttpResponse
from django.core.serializers import serialize
from django.db.models import Q, Sum, DecimalField, Count, FilteredRelation
import datetime
from dateutil.relativedelta import relativedelta
from decimal import *
from django.db.models import Q
from .utils import (
    get_dap,
    get_prev_app_date,
    get_quantidade_aplicar,
    get_base_date,
    get_index_dict_estagio,
    get_cargas_model,
    dictFarm,
    get_date,
    get_miliseconds,
    Spinner,
    is_older_than_7_days,
)

from qualidade_project.mongo_api import generate_file_run

# from qualidade_project.settings import db_name
from .models import (
    Talhao,
    Projeto,
    Variedade,
    Plantio,
    Safra,
    Ciclo,
    Defensivo,
    Aplicacao,
    Operacao,
    Colheita,
    Programa,
    PlannerPlantio,
    Deposito,
    Visitas,
    RegistroVisitas,
    PlantioDetail,
    CicloAtual,
    Fazenda,
    AppFarmboxIntegration,
    StProtheusIntegration,
    HeaderPlanejamentoAgricola,
    ColheitaPlantioExtratoArea,
    PlantioExtratoArea,
    SentSeeds,
    SeedStock,
    SeedConfig
)

from django.db.models import OuterRef, Subquery


from functools import reduce


import openpyxl
import json
import csv


from colorama import init as colorama_init
from colorama import Fore
from colorama import Style

import os
from django.conf import settings
from django.contrib.postgres.aggregates import ArrayAgg
import math

from django.db.models.functions import Round


from rest_framework.decorators import api_view, permission_classes
from django.db.models.functions.datetime import ExtractMonth, ExtractYear

import requests

from django.db.models import Case, When, DecimalField, Value, F
from django.db.models.functions import Coalesce, Round


import time
from .map_generate_pol import draw_cartoon_map
import base64

from rest_framework.authtoken.models import Token

from qualidade_project.settings import DEBUG
from qualidade_project.settings import FARMBOX_ID, PROTHEUS_TOKEN

from collections import defaultdict

from diamante.read_farm_data import get_applications, get_applications_pluvi

from django.core.mail import send_mail
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings

from django.db.models.functions import TruncDate
from itertools import groupby
from operator import itemgetter

from requests.auth import HTTPBasicAuth

from datetime import time as dateTime
from django.core.cache import cache

from django.db import transaction
from openpyxl import load_workbook


from threading import Thread
import logging

# Get a named logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# Avoid adding handlers multiple times if this config runs repeatedly
if not logger.handlers:
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    ch.setFormatter(formatter)
    logger.addHandler(ch)

# Now you can use logger.info or logger.error
logger.info('Logger configured successfully.')


main_path_upload_ids = (
    "http://localhost:5050"
    if DEBUG == True
    else "https://ubs-nodeserver.up.railway.app"
)

# --------------------- --------------------- START DEFENSIVOS MONGO API --------------------- --------------------- #


# --------------------- --------------------- END DEFENSIVOS MONGO API --------------------- --------------------- #


# --------------------- --------------------- START TALHAO API --------------------- --------------------- #


s_dict = {
    "2022/2023": 1,
    "2023/2024": 2,
    "2024/2025": 3,
    "2025/2026": 4,
}

c_dict = {"1": 3, "2": 4, "3": 5}


from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver

from concurrent.futures import ThreadPoolExecutor
from django.db.models import Sum, OuterRef, Subquery

import pandas as pd
from io import BytesIO
from openpyxl.styles import numbers


class CachedTokenAuthentication(TokenAuthentication):
    def authenticate(self, request):
        # Ensure that the Authorization header exists
        auth_header = request.META.get("HTTP_AUTHORIZATION")
        
        if not auth_header or not auth_header.startswith("Token "):
            # If the token is missing or improperly formatted, raise AuthenticationFailed
            raise AuthenticationFailed("No valid token provided")
        
        # Extract the token from the header
        token_key = auth_header.split(" ")[1]
        
        # Check if the token is cached
        cached_user = cache.get(f"token_{token_key}")
        
        if cached_user:
            return (cached_user, None)
        
        # If not cached, authenticate normally
        result = super().authenticate(request)
        
        if result:
            # Cache the authenticated user for future requests
            cache.set(f"token_{token_key}", result[0], timeout=3600)  # Cache user for 1 hour
        
        return result


class TalaoViewSet(viewsets.ModelViewSet):
    queryset = Talhao.objects.all().order_by("id_talhao")
    serializer_class = TalhaoSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=True, methods=["POST"])
    def save_talhao(self, request, pk=None):
        if request.user.is_authenticated:
            if "talhao" in request.data:
                file = request.FILES["talhao"]
                entradas = openpyxl.load_workbook(file, data_only=True)
                worksheet = entradas["talhao"]
                produtos = []
                projetos = Projeto.objects.all()
                for col in worksheet.iter_rows(min_row=1, max_col=26, max_row=30000):
                    if col[1].value != None and col[0].value != "FAZENDA":
                        fazenda = col[0].value
                        id_id = col[1].value
                        descricao = col[2].value
                        parcela = col[3].value
                        area = col[4].value
                        modulo = col[5].value
                        projeto = [x for x in projetos if x.id_d == fazenda][0]
                        try:
                            novo_talhao = Talhao(
                                id_talhao=parcela,
                                id_unico=id_id,
                                fazenda=projeto,
                                area_total=area,
                                modulo=modulo,
                            )
                            novo_talhao.save()
                            # produto = {
                            #     "fazenda": fazenda,
                            #     "id_unico": id,
                            #     "descricao": descricao,
                            #     "id_talhao": parcela,
                            #     "area_total": area,
                            #     "modulo": modulo,
                            # }
                            # produtos.append(produto)
                        except Exception as e:
                            print(f"Erro ao salvar o talhao {parcela} - {e}")
                # for i in produtos:
                #     print(i)
            qs = Talhao.objects.all()
            # Talhao.objects.aggregate(Sum('area_total'))
            serializer = TalhaoSerializer(qs, many=True)
            response = {
                "msg": f"Dados alterados com sucesso",
                "Quantidade Cadastrada": "Cadastros",
                "dados": serializer.data,
            }
            return Response(response, status=status.HTTP_200_OK)

    @action(detail=False, methods=["GET"])
    def get_talhao(self, request):
        if request.user.is_authenticated:
            try:
                qs = Talhao.objects.all()
                serializer = TalhaoSerializer(qs, many=True)
                area_total = Talhao.objects.aggregate(Sum("area_total"))
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs),
                    "Area Total dos Talhoes": area_total,
                    "dados": serializer.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- --------------------- END TALHAO API --------------------- --------------------- #

    # --------------------- --------------------- START PROJETOS API --------------------- --------------------- #

    @action(detail=False, methods=["GET"])
    def get_projetos(self, request):
        if request.user.is_authenticated:
            try:
                qs = Projeto.objects.values(
                    "nome",
                    "id_d",
                    "fazenda__nome",
                    "fazenda__id_d",
                )

                resumo = {}
                for i in qs:
                    resumo[i["fazenda__nome"]] = {}
                for i in qs:
                    resumo[i["fazenda__nome"]].update({i["id_d"]: i["nome"]})

                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "dados": resumo,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


# --------------------- --------------------- END PROJETOS API --------------------- --------------------- #


class PlantioViewSet(viewsets.ModelViewSet):
    queryset = Plantio.objects.all().order_by("data_plantio")
    serializer_class = PlantioSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    # --------------------- --------------------- PLANTIO API --------------------- --------------------- #

    @action(detail=True, methods=["POST"])
    def save_load_data(self, request, pk=None):
        if request.user.is_authenticated:
            print(request.data)
            if "plantio_arroz" in request.data:
                try:
                    file = request.FILES["plantio_arroz"]
                    entradas = openpyxl.load_workbook(file, data_only=True)
                    worksheet = entradas["Plantio"]
                    plantios = []

                    # DB CONSULT
                    talhao_list = Talhao.objects.all()
                    variedade_list = Variedade.objects.all()
                    safra = Safra.objects.all()[0]
                    ciclo = Ciclo.objects.all()[2]

                    for col in worksheet.iter_rows(min_row=1, max_col=14, max_row=3000):
                        if col[1].value != None and col[0].value != "ID":
                            id_talhao = col[0].value
                            finalizado = True if col[2].value == "Sim" else False
                            (
                                area_colher,
                                data_plantio,
                                dap,
                                id_variedade,
                            ) = (
                                col[8].value,
                                col[10].value,
                                col[11].value,
                                col[12].value,
                            )

                            if id_talhao:
                                try:
                                    talhao_id = [
                                        x
                                        for x in talhao_list
                                        if x.id_unico == id_talhao
                                    ][0]
                                except Exception as e:
                                    print(f"id sem cadastro: {id_talhao}")
                            else:
                                talhao_id = 0
                            try:
                                variedade_id = [
                                    x for x in variedade_list if x.id == id_variedade
                                ][0]
                            except Exception as e:
                                print(f"variedade sem cadastro: {id_variedade}")

                            try:
                                novo_plantio = Plantio(
                                    safra=safra,
                                    ciclo=ciclo,
                                    talhao=talhao_id,
                                    variedade=variedade_id,
                                    area_colheita=area_colher,
                                    data_plantio=data_plantio,
                                )

                                novo_plantio.save()
                            except Exception as e:
                                print(f"Problema em salvar o plantio: {id_variedade}")
                    qs_plantio = Plantio.objects.all()
                    total_plantado = Plantio.objects.aggregate(Sum("area_colheita"))
                    serializer_plantio = PlantioSerializer(qs_plantio, many=True)
                    response = {
                        "msg": f"Consulta realizada com sucesso!!",
                        "total_return": len(qs_plantio),
                        "Area Total dos Talhoes Plantados": total_plantado,
                        "dados": serializer_plantio.data,
                    }
                    return Response(response, status=status.HTTP_200_OK)
                except Exception as e:
                    response = {"message": f"Ocorreu um Erro: {e}"}
                    return Response(response, status=status.HTTP_400_BAD_REQUEST)
            else:
                response = {"message": "Arquivo desconhecido"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API --------------------- ----------------------#
    @action(detail=True, methods=["GET"])
    def get_plantio_from_farmBox(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                # file = request.FILES["plantio_arroz"]
                # file_ = open(os.path.join(settings.BASE_DIR, 'filename'))
                date_file = "2023-05-01"
                with open(f"static/files/dataset-{date_file}.json") as user_file:
                    file_contents = user_file.read()
                    parsed_json = json.loads(file_contents)
                    new_list = parsed_json
                # DB CONSULT
                talhao_list = Talhao.objects.all()
                variedade_list = Variedade.objects.all()
                safra = Safra.objects.all()[1]
                ciclo_list = Ciclo.objects.all()
                projetos = Projeto.objects.all()

                area_total_1 = 0
                area_total_2 = 0
                area_total_3 = 0
                cultura_list = []
                ciclo_and_cultura = []
                ciclo_1 = {}
                ciclo_2 = {}
                ciclo_3 = {}

                for i in new_list:
                    state = i["state"]
                    date_plantio = i["date"]
                    activation_date = ["activation_date"]
                    parcela = i["name"].replace(" ", "")
                    farm_name = i["farm_name"]
                    variedade_name = i["variety_name"]
                    area = i["area"]

                    variedade_planejada = i["planned_variety_name"]
                    cultura_planejada = i["planned_culture_name"]

                    culture_id = i["culture_id"]
                    variety_id = i["variety_id"]
                    fazenda_id = i["farm"]["id"]
                    variedade_planejada_id = i["planned_variety_id"]
                    cultura_planejada_id = i["planned_culture_id"]

                    safra_farm = i["harvest_name"]
                    ciclo_json = i["cycle"]

                    ciclo = ciclo_list[ciclo_json - 1]

                    id_talhao = [
                        x.id_d for x in projetos if x.id_farmbox == fazenda_id
                    ][0]
                    id_variedade = [
                        x
                        for x in variedade_list
                        if x.id_farmbox == variedade_planejada_id
                    ][0]

                    if id_talhao:
                        try:
                            talhao_id = f"{id_talhao}{parcela}"
                            talhao_id = [
                                x for x in talhao_list if x.id_unico == talhao_id
                            ][0]
                        except Exception as e:
                            print(
                                f"{Fore.RED}id sem cadastro: {id_talhao}{parcela} - {farm_name} - Ciclo: {ciclo}{Style.RESET_ALL}"
                            )
                    else:
                        talhao_id = 0
                    try:
                        id_variedade = [
                            x
                            for x in variedade_list
                            if x.id_farmbox == variedade_planejada_id
                        ][0]
                    except Exception as e:
                        print(
                            f"{Fore.RED}variedade sem cadastro: {id_variedade}{Style.RESET_ALL}"
                        )
                    if cultura_planejada:
                        if ciclo_json == 1:
                            if ciclo_1.get(cultura_planejada):
                                area_total = ciclo_1.get(cultura_planejada) + area
                                ciclo_1.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_1.update({cultura_planejada: round(area)})
                        if ciclo_json == 2:
                            if ciclo_2.get(cultura_planejada):
                                area_total = ciclo_2.get(cultura_planejada) + area
                                ciclo_2.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_2.update({cultura_planejada: round(area)})
                        if ciclo_json == 3:
                            if ciclo_3.get(cultura_planejada):
                                area_total = ciclo_3.get(cultura_planejada) + area
                                ciclo_3.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_3.update({cultura_planejada: round(area)})
                print(f"Area Total da Soja Ciclo 1: {area_total_1}")
                print(f"Area Total da Soja Ciclo 2: {area_total_2}")
                print(f"Area Total da Soja Ciclo 3: {area_total_3}")

                print(set(cultura_list))
                print("Ciclo 1")
                for k, v in ciclo_1.items():
                    print(f"Cultura: {k} - Area: {v}")
                print("Ciclo 2")
                for k, v in ciclo_2.items():
                    print(f"Cultura: {k} - Area: {v}")
                print("Ciclo 3")
                for k, v in ciclo_3.items():
                    print(f"Cultura: {k} - Area: {v}")
                qs_plantio = Plantio.objects.filter(safra__safra="2023/2024")

                total_plantado = Plantio.objects.filter(
                    safra__safra="2023/2024", ciclo__ciclo="1"
                ).aggregate(Sum("area_colheita"))

                serializer_plantio = PlantioSerializer(qs_plantio, many=True)
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs_plantio),
                    "Area Total dos Talhoes Plantados": total_plantado,
                    "dados": serializer_plantio.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API --------------------- ----------------------#

    # SAVE PLANTIO ON NEW SAFRA CICLO
    @action(detail=True, methods=["POST"])
    def save_plantio_from_farmBox_json(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                # file = request.FILES["plantio_arroz"]
                # file_ = open(os.path.join(settings.BASE_DIR, 'filename'))
                # date_file = "2024-06-13 08:07"
                date_file = request.data["filename"]
                with open(f"static/files/dataset-{date_file}.json") as user_file:
                    file_contents = user_file.read()
                    parsed_json = json.loads(file_contents)
                    new_list = parsed_json
                # DB CONSULT
                talhao_list = Talhao.objects.all()
                variedade_list = Variedade.objects.all()
                safra_list = Safra.objects.all()
                # safra = Safra.objects.all()[0]
                ciclo_list = Ciclo.objects.all()
                projetos = Projeto.objects.all()

                area_total_1 = 0
                area_total_2 = 0
                area_total_3 = 0
                cultura_list = []
                ciclo_and_cultura = []
                ciclo_1 = {}
                ciclo_2 = {}
                ciclo_3 = {}

                for i in new_list:
                    state = i["state"]
                    date_plantio = i["date"]
                    activation_date = ["activation_date"]
                    parcela = i["name"].replace(" ", "")
                    farm_name = i["farm_name"]
                    variedade_name = i["variety_name"]
                    area = i["area"]
                    area_planejamento = i["area"]
                    id_plantio_farmbox = i["id"]

                    planned_date = None
                    if i["planned_date"]:
                        planned_date = i["planned_date"].split('T')[0]

                    variedade_planejada = i["planned_variety_name"]
                    cultura_planejada = i["planned_culture_name"]

                    culture_id = i["culture_id"]
                    variety_id = i["variety_id"]
                    fazenda_id = i["farm"]["id"]
                    variedade_planejada_id = i["planned_variety_id"]
                    cultura_planejada_id = i["planned_culture_id"]

                    map_centro_id_farm = i["centroid"]
                    map_geo_points_farm = i["geo_points"]

                    safra_farm = i["harvest_name"]
                    safra = [x for x in safra_list if x.safra == safra_farm][0]
                    ciclo_json = i["cycle"]

                    ciclo = ciclo_list[ciclo_json - 1]

                    id_talhao = [
                        x.id_d for x in projetos if x.id_farmbox == fazenda_id
                    ][0]
                    id_variedade = [
                        x
                        for x in variedade_list
                        if x.id_farmbox == variedade_planejada_id
                    ][0]

                    if id_talhao:
                        try:
                            talhao_id = f"{id_talhao}{parcela}"
                            talhao_id = [
                                x for x in talhao_list if x.id_unico == talhao_id
                            ][0]
                        except Exception as e:
                            print(f"id sem cadastro: {id_talhao}{parcela}")
                    else:
                        talhao_id = 0
                    try:
                        # variedade_id = [
                        #     x for x in variedade_list if x.id == id_variedade
                        # ][0]
                        id_variedade = [
                            x
                            for x in variedade_list
                            if x.id_farmbox == variedade_planejada_id
                        ][0]
                    except Exception as e:
                        print(
                            f"{Fore.RED}variedade sem cadastro: {id_variedade}{Style.RESET_ALL}"
                        )
                    if cultura_planejada:
                        try:
                            with transaction.atomic():  # Ensu
                                novo_plantio = Plantio(
                                    safra=safra,
                                    ciclo=ciclo,
                                    talhao=talhao_id,
                                    variedade=id_variedade,
                                    area_colheita=area,
                                    map_centro_id=map_centro_id_farm,
                                    map_geo_points=map_geo_points_farm,
                                    id_farmbox=id_plantio_farmbox,
                                    # data_prevista_plantio=planned_date,
                                    area_planejamento_plantio=area_planejamento,
                                    # data_plantio=data_plantio,
                                )

                                novo_plantio.save()
                                print(
                                    f"{Fore.GREEN}Novo Plantio salvo com sucesso: {novo_plantio}{Style.RESET_ALL}"
                                )
                                # print(
                                #     f"{Fore.GREEN}Safra/Ciclo: {safra}-{ciclo} - Estado: {state} - Data Plantio: {date_plantio} - Parcela: {parcela} {Style.RESET_ALL}- Fazenda: {farm_name} - Cultura Planejada: {cultura_planejada} Variedade Planejada: {variedade_planejada}|{id_variedade} - Area: {area} - Cultura_ID: {culture_id} - Variedade_planejada_id: {variedade_planejada_id} - Fazenda_ID: {fazenda_id}"
                                # )
                                # print(
                                #     f"{Fore.YELLOW}Safra:{safra}-Ciclo:{ciclo} - Parcela: {talhao_id} - Variedade: {id_variedade} - area: {area}{Style.RESET_ALL}"
                                # )
                                print("\n")
                        except Exception as e:
                            print(
                                f"{Fore.RED}Problema em salvar o plantio: {id_variedade}{Style.RESET_ALL}{e}"
                            )
                        if ciclo_json == 1:
                            if ciclo_1.get(cultura_planejada):
                                area_total = ciclo_1.get(cultura_planejada) + area
                                ciclo_1.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_1.update({cultura_planejada: round(area)})
                        if ciclo_json == 2:
                            if ciclo_2.get(cultura_planejada):
                                area_total = ciclo_2.get(cultura_planejada) + area
                                ciclo_2.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_2.update({cultura_planejada: round(area)})
                        if ciclo_json == 3:
                            if ciclo_3.get(cultura_planejada):
                                area_total = ciclo_3.get(cultura_planejada) + area
                                ciclo_3.update({cultura_planejada: round(area_total)})
                            else:
                                ciclo_3.update({cultura_planejada: round(area)})
                    else:
                        try:
                            novo_plantio = Plantio(
                                safra=safra,
                                ciclo=ciclo,
                                talhao=talhao_id,
                                variedade=None,
                                area_colheita=area,
                                programa=None,
                                map_centro_id=map_centro_id_farm,
                                map_geo_points=map_geo_points_farm,
                                id_farmbox=id_plantio_farmbox,
                                # data_plantio=data_plantio,
                            )
                            novo_plantio.save()
                            print(
                                f"{Fore.GREEN}Novo Plantio salvo com sucesso: {novo_plantio}{Style.RESET_ALL}"
                            )
                        except Exception as e:
                            print(
                                f"{Fore.RED}Problema em salvar o plantio Não Planejado: {talhao_id} - {safra} - {ciclo}{Style.RESET_ALL}{e}"
                            )

                print(f"Area Total da Soja Ciclo 1: {area_total_1}")
                print(f"Area Total da Soja Ciclo 2: {area_total_2}")
                print(f"Area Total da Soja Ciclo 3: {area_total_3}")

                print(set(cultura_list))
                print("Ciclo 1")
                for k, v in ciclo_1.items():
                    print(f"Cultura: {k} - Area: {v}")
                print("Ciclo 2")
                for k, v in ciclo_2.items():
                    print(f"Cultura: {k} - Area: {v}")
                print("Ciclo 3")
                for k, v in ciclo_3.items():
                    print(f"Cultura: {k} - Area: {v}")
                qs_plantio = Plantio.objects.filter(safra__safra="2023/2024")
                total_plantado = Plantio.objects.filter(
                    safra__safra="2023/2024"
                ).aggregate(Sum("area_colheita"))
                serializer_plantio = PlantioSerializer(qs_plantio, many=True)
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs_plantio),
                    "Area Total dos Talhoes Plantados": total_plantado,
                    "dados": serializer_plantio.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API FROM FARMBOX START --------------------- ----------------------#
    @action(detail=True, methods=["GET", "POST"])
    def update_plantio_from_farmBox(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                new_list = request.data
                print('request data: ', new_list)
                # file = request.FILES["plantio_arroz"]
                # file_ = open(os.path.join(settings.BASE_DIR, 'filename'))
                # date_file = "2023-09-02 07:41"
                # with open(f"static/files/dataset-{date_file}.json") as user_file:
                #     file_contents = user_file.read()
                #     parsed_json = json.loads(file_contents)
                #     new_list = parsed_json
                # DB CONSULT
                talhao_list = Talhao.objects.all()
                variedade_list = Variedade.objects.all()
                safra_list = Safra.objects.all()
                safra_list = [s for s in safra_list]
                ciclo_list = Ciclo.objects.all()
                projetos = Projeto.objects.all()

                count_total = 0
                ids_farmbox_list = [x['id'] for x in new_list]
                print('newList IDs: ', ids_farmbox_list)
                plantio_extrato_area = (
                    PlantioExtratoArea.objects
                    .select_related('plantio')
                    .filter(plantio__id_farmbox__in=ids_farmbox_list)
                    .filter(ativo=True)
                    .values('plantio__id_farmbox')
                    .annotate(total_area_plantada=Sum('area_plantada'))
                )
                print('newList IDs plantio area: ', plantio_extrato_area)
                
                dev_mode = True
                if dev_mode:
                    for i in new_list:
                        state = i["state"]
                        date_plantio = i["date"]
                        emergence_date = i["emergence_date"]
                        activation_date = ["activation_date"]
                        parcela = i["name"].replace(" ", "")
                        farm_name = i["farm_name"]
                        variedade_name = i["variety_name"]
                        area = i["area"]
                        area_planejamento = i["area"]
                        id_plantio_farmbox = i["id"]

                        planned_date = None
                        if i["planned_date"]:
                            planned_date = i["planned_date"].split('T')[0]

                        variedade_planejada = i["planned_variety_name"]
                        cultura_planejada = i["planned_culture_name"]

                        culture_id = i["culture_id"]
                        variety_id = i["variety_id"]
                        fazenda_id = i["farm"]["id"]
                        variedade_planejada_id = i["planned_variety_id"]
                        cultura_planejada_id = i["planned_culture_id"]

                        map_centro_id_farm = i["centroid"]
                        map_geo_points_farm = i["geo_points"]

                        safra_farm = i["harvest_name"]
                        ciclo_json = i["cycle"]

                        ciclo = ciclo_list[ciclo_json - 1]
                        safra = [s for s in safra_list if s.safra == safra_farm][0]

                        id_talhao = [
                            x.id_d for x in projetos if x.id_farmbox == fazenda_id
                        ][0]

                        id_variedade = [
                            x
                            for x in variedade_list
                            if x.id_farmbox == variedade_planejada_id
                        ][0]

                        if id_talhao:
                            try:
                                talhao_id = f"{id_talhao}{parcela}"
                                talhao_id = [
                                    x for x in talhao_list if x.id_unico == talhao_id
                                ][0]
                            except Exception as e:
                                print(
                                    f"{Fore.RED}id sem cadastro: {id_talhao}{parcela} - {farm_name} - Ciclo: {ciclo}{Style.RESET_ALL}"
                                )
                        else:
                            talhao_id = 0
                        try:
                            id_variedade = [
                                x
                                for x in variedade_list
                                if x.id_farmbox == variedade_planejada_id
                            ][0]
                        except Exception as e:
                            print(
                                f"{Fore.RED}variedade sem cadastro: {id_variedade}{Style.RESET_ALL}"
                            )
                        
                        
                        # PEGANDO AREA PLANTADA DO EXTRATO DE PLANTIO, CASO NAO ACHE USE A AREA COMO PADRAO
                        get_plantio_area = plantio_extrato_area.filter(plantio__id_farmbox=id_plantio_farmbox).order_by('?').first()
                        if get_plantio_area:
                            total_area_plantada = get_plantio_area.get('total_area_plantada', 0)  # 0 as a fallback if not found
                            print('area planted here: ', total_area_plantada)
                        else:
                            total_area_plantada = area # If no result found
                        
                        if cultura_planejada or variety_id:
                            try:
                                field_to_update = Plantio.objects.filter(
                                    safra=safra, ciclo=ciclo, talhao=talhao_id
                                )[0]
                                field_to_update.id_farmbox = id_plantio_farmbox
                                if field_to_update.farmbox_update == True:
                                    if field_to_update.finalizado_colheita == False:
                                        # if planned_date:
                                        #     field_to_update.data_prevista_plantio = planned_date
                                        if area:
                                            field_to_update.area_colheita = total_area_plantada
                                            field_to_update.area_planejamento_plantio = area_planejamento

                                        if map_centro_id_farm:
                                            field_to_update.map_centro_id = map_centro_id_farm

                                        if map_geo_points_farm:
                                            field_to_update.map_geo_points = map_geo_points_farm

                                        if state == "active":
                                            if date_plantio:
                                                field_to_update.data_plantio = date_plantio
                                                field_to_update.finalizado_plantio = True
                                                field_to_update.area_colheita = total_area_plantada

                                            if emergence_date:
                                                field_to_update.data_emergencia = emergence_date

                                        if variety_id:
                                            id_variedade_done = [
                                                x
                                                for x in variedade_list
                                                if x.id_farmbox == variety_id
                                            ][0]
                                            field_to_update.variedade = id_variedade_done
                                        else:
                                            field_to_update.variedade = id_variedade
                                        field_to_update.save()
                                        print(
                                            f"{Fore.GREEN}Plantio Alterado com sucesso: {field_to_update} - {safra} - {ciclo} | {Fore.CYAN}{field_to_update.variedade} | {field_to_update.programa}{Style.RESET_ALL}"
                                        )
                                        print("\n")
                                        count_total += 1
                                else:    
                                    if map_centro_id_farm:
                                        field_to_update.map_centro_id = map_centro_id_farm
                                    if map_geo_points_farm:
                                        field_to_update.map_geo_points = map_geo_points_farm
                                    field_to_update.save()
                            except Exception as e:
                                print(
                                    f"{Fore.RED}Problema em salvar o plantio: {talhao_id} - {safra} - {ciclo}{Style.RESET_ALL}{e}"
                                )
                        else:
                            try:
                                field_to_update = Plantio.objects.filter(
                                    safra=safra, ciclo=ciclo, talhao=talhao_id
                                )[0]
                                if field_to_update.farmbox_update == True:
                                    if area:
                                        field_to_update.area_colheita = total_area_plantada
                                    if map_centro_id_farm:
                                        field_to_update.map_centro_id = map_centro_id_farm
                                    if map_geo_points_farm:
                                        field_to_update.map_geo_points = map_geo_points_farm
                                    field_to_update.variedade = None
                                    field_to_update.programa = None
                                    field_to_update.id_farmbox = id_plantio_farmbox
                                    field_to_update.save()
                                    print(
                                        f"{Fore.YELLOW}Plantio Alterado com sucesso para SEM VARIEDADE: {field_to_update}- {safra} - {ciclo}{Style.RESET_ALL}"
                                    )
                                    print("\n")
                                    count_total += 1
                                else:
                                    if map_centro_id_farm:
                                        field_to_update.map_centro_id = map_centro_id_farm
                                    if map_geo_points_farm:
                                        field_to_update.map_geo_points = map_geo_points_farm
                                    field_to_update.save()
                                    
                            except Exception as e:
                                print(
                                    f"{Fore.RED}Problema em salvar o plantio Não Planejado: {talhao_id} - {safra} - {ciclo}{Style.RESET_ALL}{e}"
                                )
                else:
                    print('dev mode')
                qs_plantio = Plantio.objects.filter(safra__safra="2023/2024")

                # serializer_plantio = PlantioSerializer(qs_plantio, many=True)
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs_plantio),
                    "Total de Talhões alterados": count_total,
                    # 'total soma plantado': plantio_extrato_area
                    # "dados": serializer_plantio.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                print(f"{Fore.RED}Problema ao ler o arquivo: {Style.RESET_ALL}{e}")
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API FROM FARMBOX END--------------------- ----------------------#

    # --------------------- ---------------------- UPDATE GEO POINNTS API FROM FARMBOX START --------------------- ----------------------#
    @action(detail=True, methods=["GET"])
    def update_geopoints(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                date_file = "2023-06-18 11:22- 2023-2024-geral-geo"
                with open(f"static/files/dataset-{date_file}.json") as user_file:
                    file_contents = user_file.read()
                    parsed_json = json.loads(file_contents)
                    new_list = parsed_json
                # DB CONSULT
                talhao_list = Talhao.objects.all()
                variedade_list = Variedade.objects.all()
                safra_list = Safra.objects.all()
                safra_list = [s for s in safra_list]
                ciclo_list = Ciclo.objects.all()
                projetos = Projeto.objects.all()

                count_total = 0
                for i in new_list:
                    state = i["state"]
                    date_plantio = i["date"]
                    emergence_date = i["emergence_date"]
                    activation_date = ["activation_date"]
                    parcela = i["name"].replace(" ", "")
                    farm_name = i["farm_name"]
                    variedade_name = i["variety_name"]
                    area = i["area"]

                    variedade_planejada = i["planned_variety_name"]
                    cultura_planejada = i["planned_culture_name"]

                    culture_id = i["culture_id"]
                    variety_id = i["variety_id"]
                    fazenda_id = i["farm"]["id"]
                    variedade_planejada_id = i["planned_variety_id"]
                    cultura_planejada_id = i["planned_culture_id"]

                    map_centro_id_farm = i["centroid"]
                    map_geo_points_farm = i["geo_points"]

                    safra_farm = i["harvest_name"]
                    ciclo_json = i["cycle"]

                    ciclo = ciclo_list[ciclo_json - 1]
                    safra = [s for s in safra_list if s.safra == safra_farm][0]

                    id_talhao = [
                        x.id_d for x in projetos if x.id_farmbox == fazenda_id
                    ][0]
                    id_variedade = [
                        x
                        for x in variedade_list
                        if x.id_farmbox == variedade_planejada_id
                    ][0]

                    if id_talhao:
                        try:
                            talhao_id = f"{id_talhao}{parcela}"
                            talhao_id = [
                                x for x in talhao_list if x.id_unico == talhao_id
                            ][0]
                        except Exception as e:
                            print(
                                f"{Fore.RED}id sem cadastro: {id_talhao}{parcela} - {farm_name} - Ciclo: {ciclo}{Style.RESET_ALL}"
                            )
                    else:
                        talhao_id = 0
                    try:
                        id_variedade = [
                            x
                            for x in variedade_list
                            if x.id_farmbox == variedade_planejada_id
                        ][0]
                    except Exception as e:
                        print(
                            f"{Fore.RED}variedade sem cadastro: {id_variedade}{Style.RESET_ALL}"
                        )
                    if cultura_planejada or variety_id:
                        try:
                            field_to_update = Plantio.objects.filter(
                                safra=safra, ciclo=ciclo, talhao=talhao_id
                            )[0]
                            if map_centro_id_farm:
                                field_to_update.map_centro_id = map_centro_id_farm

                            if map_geo_points_farm:
                                field_to_update.map_geo_points = map_geo_points_farm

                            field_to_update.save()
                            print(
                                f"{Fore.GREEN}Plantio Alterado com sucesso: {field_to_update}{Style.RESET_ALL}"
                            )
                            print("\n")
                            count_total += 1
                        except Exception as e:
                            print(
                                f"{Fore.RED}Problema em salvar o plantio: {talhao_id} - {safra} - {ciclo}{Style.RESET_ALL}{e}"
                            )

                qs_plantio = Plantio.objects.filter(safra__safra="2023/2024")

                serializer_plantio = PlantioSerializer(qs_plantio, many=True)
                response = {
                    "msg": f"Atualização de Área realizada com sucesso!!",
                    "total_return": len(qs_plantio),
                    "Total de Talhões alterados": count_total,
                    "dados": serializer_plantio.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API FROM FARMBOX END--------------------- ----------------------#

    @action(detail=True, methods=["GET"])
    def get_plantio_info(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                qs_plantio_date = Plantio.objects.filter(
                    safra__safra="2023/2024", ciclo="3", finalizado_plantio=True
                )
                data_plantada = {}
                area_total = 0
                for i in qs_plantio_date:
                    area_total += i.area_colheita
                    data_plantada.update({i.variedade.cultura.cultura: area_total})

                total_plantado = Plantio.objects.filter(
                    safra__safra="2023/2024", ciclo="3"
                ).aggregate(Sum("area_colheita"))

                serializer_plantio = PlantioSerializer(qs_plantio_date, many=True)
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs_plantio_date),
                    "Area Total dos Talhoes Plantados": total_plantado,
                    "data_plantada": data_plantada,
                    "dados": serializer_plantio.data,
                }
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- UPDATE PLANTIO API --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = None
                cicle_filter = None
                try:
                    safra_filter = request.data["safra"]
                    cicle_filter = request.data["ciclo"]
                except Exception as e:
                    print(e)

                if safra_filter is None or cicle_filter is None: 
                    print('safra ou filtro não informado')
                    current_safra = CicloAtual.objects.filter(nome="Colheita")[0]
                    safra_filter = current_safra.safra.safra
                    cicle_filter = current_safra.ciclo.ciclo

                safra_filter = "2024/2025" if safra_filter == None else safra_filter
                cicle_filter = "2" if cicle_filter == None else cicle_filter

                qs = (
                    Plantio.objects.values(
                        "safra__safra",
                        "pk",
                        "ciclo__ciclo",
                        "talhao__id_talhao",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__cultura__cultura",
                        "variedade__nome_fantasia",
                        "finalizado_plantio",
                        "finalizado_colheita",
                        "area_colheita",
                        "area_parcial",
                        "data_plantio",
                    )
                    .order_by("talhao__fazenda__nome", "talhao__id_talhao")
                    .filter(
                        safra__safra=safra_filter,
                        ciclo__ciclo=cicle_filter,
                        finalizado_plantio=True,
                    )
                    # .filter(~Q(data_plantio=None))
                )

                qs_annotate = Plantio.objects.filter(
                    safra__safra=safra_filter,
                    ciclo__ciclo=cicle_filter,
                    finalizado_plantio=True,
                )
                res1 = (
                    qs_annotate.values(
                        "talhao__fazenda__nome", "variedade__cultura__cultura"
                    )
                    .order_by("talhao__fazenda__nome")
                    .annotate(area_total=Sum("area_colheita"))
                )
                res2 = (
                    qs_annotate.values(
                        "talhao__fazenda__nome", "talhao__fazenda__fazenda__nome"
                    )
                    .annotate(count=Count("talhao__fazenda__nome"))
                    .order_by("talhao__fazenda__nome")
                )

                # ------------- ------------- START SEPARADO POR PROJETO ------------- -------------#
                resumo = {i["talhao__fazenda__nome"]: {} for i in qs}

                {
                    resumo[i["talhao__fazenda__nome"]].update(
                        {
                            i["talhao__id_talhao"]: {
                                "safra": i["safra__safra"],
                                "ciclo": i["ciclo__ciclo"],
                                "cultura": i["variedade__cultura__cultura"],
                                "variedade": i["variedade__nome_fantasia"],
                                "finalizado_colheita": i["finalizado_colheita"],
                                "id_plantio": i["pk"],
                            }
                        }
                    )
                    for i in qs
                }
                # ------------- ------------- END SEPARADO POR PROJETO ------------- -------------#

                # ------------- ------------- START SEPARADO POR FAZENDA > PROJETO ------------- -------------#
                # resumo = {}
                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]] = {
                #         i["talhao__fazenda__nome"]: {}
                #     }

                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]].update(
                #         {i["talhao__fazenda__nome"]: {}}
                #     )
                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]][
                #         i["talhao__fazenda__nome"]
                #     ].update(
                #         {
                #             i["talhao__id_talhao"]: {
                #                 "safra": i["safra__safra"],
                #                 "ciclo": i["ciclo__ciclo"],
                #                 "cultura": i["variedade__cultura__cultura"],
                #                 "variedade": i["variedade__nome_fantasia"],
                #                 "finalizado_colheita": i["finalizado_colheita"],
                #             }
                #         }
                #     )

                # ------------- ------------- END SEPARADO POR FAZENDA > PROJETO ------------- -------------#

                area_total = qs_annotate.aggregate(Sum("area_colheita"))
                projetos = Projeto.objects.values("nome", "fazenda__nome", "id_d")
                resumo_by_farm = []
                for projeto in projetos:
                    filter_array = list(
                        filter(
                            lambda x: x["fazenda"] == projeto["fazenda__nome"],
                            resumo_by_farm,
                        )
                    )
                    # print(filter_array)
                    if len(filter_array) > 0:
                        index_to_update = [
                            i
                            for i, _ in enumerate(resumo_by_farm)
                            if _["fazenda"] == projeto["fazenda__nome"]
                        ][0]
                        resumo_by_farm[index_to_update]["projetos"].append(
                            projeto["nome"]
                        )
                    else:
                        dict_to_insert = {
                            "fazenda": projeto["fazenda__nome"],
                            "projetos": [projeto["nome"]],
                        }
                        resumo_by_farm.append(dict_to_insert)

                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs),
                    "resumo_safra": res1,
                    "fazenda_grupo_projetos": resumo_by_farm,
                    "projetos": projetos,
                    "resumo_safra_fazenda": res2,
                    "Area Total dos Talhoes Plantados": area_total,
                    "dados": resumo,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- --------------------- END PLANTIO API --------------------- --------------------- #
    
    # --------------------- ---------------------- UPDATE PLANTIO API FARMTRUCK --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_farmtruck(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = None
                cicle_filter = None
                try:
                    safra_filter = request.data["safra"]
                    cicle_filter = request.data["ciclo"]
                except Exception as e:
                    print(e)

                if safra_filter is None or cicle_filter is None: 
                    print('safra ou filtro não informado')
                    current_safra = CicloAtual.objects.filter(nome="Colheita")[0]
                    safra_filter = current_safra.safra.safra
                    cicle_filter = current_safra.ciclo.ciclo

                safra_filter = "2024/2025" if safra_filter == None else safra_filter
                cicle_filter = "2" if cicle_filter == None else cicle_filter

                qs = (
                    Plantio.objects.values(
                        "safra__safra",
                        "pk",
                        "ciclo__ciclo",
                        "talhao__id_talhao",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__cultura__cultura",
                        "variedade__nome_fantasia",
                        "finalizado_plantio",
                        "finalizado_colheita",
                        "area_colheita",
                        "area_parcial",
                        "data_plantio",
                    )
                    .order_by("talhao__fazenda__nome", "talhao__id_talhao")
                    .filter(
                        safra__safra=safra_filter,
                        ciclo__ciclo=cicle_filter,
                        finalizado_plantio=True,
                    )
                    # .filter(~Q(data_plantio=None))
                )

                qs_annotate = Plantio.objects.filter(
                    safra__safra=safra_filter,
                    ciclo__ciclo=cicle_filter,
                    finalizado_plantio=True,
                )
                res1 = (
                    qs_annotate.values(
                        "talhao__fazenda__nome", "variedade__cultura__cultura"
                    )
                    .order_by("talhao__fazenda__nome")
                    .annotate(area_total=Sum("area_colheita"))
                )
                res2 = (
                    qs_annotate.values(
                        "talhao__fazenda__nome", "talhao__fazenda__fazenda__nome"
                    )
                    .annotate(count=Count("talhao__fazenda__nome"))
                    .order_by("talhao__fazenda__nome")
                )

                # ------------- ------------- START SEPARADO POR PROJETO ------------- -------------#
                resumo = {i["talhao__fazenda__nome"]: {} for i in qs}

                {
                    resumo[i["talhao__fazenda__nome"]].update(
                        {
                            i["talhao__id_talhao"]: {
                                "safra": i["safra__safra"],
                                "ciclo": i["ciclo__ciclo"],
                                "cultura": i["variedade__cultura__cultura"],
                                "variedade": i["variedade__nome_fantasia"],
                                "finalizado_colheita": i["finalizado_colheita"],
                                "id_plantio": i["pk"],
                            }
                        }
                    )
                    for i in qs
                }
                # ------------- ------------- END SEPARADO POR PROJETO ------------- -------------#

                # ------------- ------------- START SEPARADO POR FAZENDA > PROJETO ------------- -------------#
                # resumo = {}
                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]] = {
                #         i["talhao__fazenda__nome"]: {}
                #     }

                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]].update(
                #         {i["talhao__fazenda__nome"]: {}}
                #     )
                # for i in qs:
                #     resumo[i["talhao__fazenda__fazenda__nome"]][
                #         i["talhao__fazenda__nome"]
                #     ].update(
                #         {
                #             i["talhao__id_talhao"]: {
                #                 "safra": i["safra__safra"],
                #                 "ciclo": i["ciclo__ciclo"],
                #                 "cultura": i["variedade__cultura__cultura"],
                #                 "variedade": i["variedade__nome_fantasia"],
                #                 "finalizado_colheita": i["finalizado_colheita"],
                #             }
                #         }
                #     )

                # ------------- ------------- END SEPARADO POR FAZENDA > PROJETO ------------- -------------#

                area_total = qs_annotate.aggregate(Sum("area_colheita"))
                projetos = Projeto.objects.values("nome", "fazenda__nome", "id_d")
                resumo_by_farm = []
                for projeto in projetos:
                    filter_array = list(
                        filter(
                            lambda x: x["fazenda"] == projeto["fazenda__nome"],
                            resumo_by_farm,
                        )
                    )
                    # print(filter_array)
                    if len(filter_array) > 0:
                        index_to_update = [
                            i
                            for i, _ in enumerate(resumo_by_farm)
                            if _["fazenda"] == projeto["fazenda__nome"]
                        ][0]
                        resumo_by_farm[index_to_update]["projetos"].append(
                            projeto["nome"]
                        )
                    else:
                        dict_to_insert = {
                            "fazenda": projeto["fazenda__nome"],
                            "projetos": [projeto["nome"]],
                        }
                        resumo_by_farm.append(dict_to_insert)

                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs),
                    "resumo_safra": res1,
                    "fazenda_grupo_projetos": resumo_by_farm,
                    "projetos": projetos,
                    "resumo_safra_fazenda": res2,
                    "Area Total dos Talhoes Plantados": area_total,
                    "dados": resumo,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- --------------------- END PLANTIO API fARMTRUC --------------------- --------------------- #

    # --------------------- --------------------- START PLANTIO API DONE --------------------- --------------------- #

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_done(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]

                print(safra_filter)
                print(cicle_filter)
                # safra_filter = request.POST.get("safra")
                # cicle_filter = request.POST.get("ciclo")
                # print(safra_from_request, ciclo_from_request)
                safra_filter = "2023/2024" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter

                qs = (
                    Plantio.objects.values(
                        "safra__safra",
                        "ciclo__ciclo",
                        "talhao__id_talhao",
                        "talhao__id_unico",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__map_centro_id",
                        "talhao__fazenda__fazenda__nome",
                        "talhao__fazenda__fazenda__capacidade_plantio_ha_dia",
                        "programa__start_date",
                        "variedade__cultura__cultura",
                        "variedade__dias_ciclo",
                        "variedade__cultura__map_color",
                        "variedade__cultura__map_color_line",
                        "variedade__nome_fantasia",
                        "variedade__variedade",
                        "area_colheita",
                        "data_plantio",
                        "map_centro_id",
                        "map_geo_points",
                        "cronograma_programa__0",
                        "finalizado_plantio",
                        "finalizado_colheita",
                        "plantio_descontinuado",
                    )
                    .order_by(
                        "data_plantio", "talhao__fazenda__nome", "talhao__id_talhao"
                    )
                    .filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
                    # .filter(finalizado_plantio=True)
                    .filter(plantio_descontinuado=False)
                    .filter(~Q(programa_id=None))
                )

                qs_planejamento = (
                    PlannerPlantio.objects.select_related(
                        "projeto",
                        "cultura",
                        "ciclo",
                        "cultura",
                        "projeto__fazenda",
                        "safra",
                    )
                    .values(
                        "projeto",
                        "projeto__id",
                        "projeto__nome",
                        "cultura",
                        "variedade",
                        "safra__safra",
                        "ciclo__ciclo",
                        "start_date",
                        "area",
                    )
                    .filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
                )

                prev_date = {}

                for po, i in enumerate(qs):
                    if i["data_plantio"] == None:
                        capacidade_dia = i[
                            "talhao__fazenda__fazenda__capacidade_plantio_ha_dia"
                        ]
                        try:
                            # filtered_planner = qs_planejamento.filter(
                            #     projeto__nome=i["talhao__fazenda__nome"]
                            # )
                            # planner_date = False
                            # if filtered_planner:
                            #     # print("aqui temos planejamneto: , ", filtered_planner)
                            #     current_planner = filtered_planner[0]
                            #     inital_date_planner = current_planner["start_date"]
                            #     planner_date = True

                            # if planner_date:
                            #     start_date = get_base_date(inital_date_planner)
                            # else:
                            start_date = get_base_date(i["programa__start_date"])
                        except Exception as e:
                            print("error :", e)
                        if prev_date.get(i["talhao__fazenda__nome"]):
                            area_atual = prev_date.get(i["talhao__fazenda__nome"])[
                                "area"
                            ]
                            dias_necessarios = round(area_atual / capacidade_dia)
                            qs[po].update(
                                {
                                    "data_plantio": prev_date.get(
                                        i["talhao__fazenda__nome"]
                                    )["data_inicial"]
                                    + datetime.timedelta(days=dias_necessarios)
                                }
                            )
                            prev_date.update(
                                {
                                    i["talhao__fazenda__nome"]: {
                                        "area": area_atual + i["area_colheita"],
                                        "dias_necessários": 0,
                                        "data_inicial": start_date,
                                        "data_final": None,
                                    }
                                }
                            )
                        else:
                            prev_date.update(
                                {
                                    i["talhao__fazenda__nome"]: {
                                        "area": i["area_colheita"],
                                        "dias_necessários": 0,
                                        "data_inicial": start_date,
                                        "data_final": None,
                                    }
                                }
                            )
                            qs[po].update({"data_plantio": start_date})
                print(prev_date)

                # ln = [
                #     {**x, "data_plantio": "2024-01-01"}
                #     if x["data_plantio"] == None
                #     else x
                #     for x in qs
                # ]

                qs_by_day = (
                    Plantio.objects.values(
                        "data_plantio",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__cultura__cultura",
                    )
                    .annotate(area_total=Sum("area_colheita"))
                    .filter(finalizado_plantio=True, plantio_descontinuado=False)
                    .filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
                    .order_by("data_plantio")
                )

                response = {
                    "msg": f"Consulta realizada com sucesso GetPlantioDone API!! - Safra: {safra_filter} - Ciclo: {cicle_filter}",
                    "total_return": len(qs),
                    "data": qs,
                    "plantio_by_day": qs_by_day,
                    # "resume_by_farm": qsFarm,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- --------------------- END PLANTIO API DONE --------------------- --------------------- #

    @action(detail=True, methods=["POST"])
    def update_plantio_data(self, request, pk=None):
        if request.user.is_authenticated:
            print(request.data)
            if "plantio_arroz" in request.data:
                try:
                    file = request.FILES["plantio_arroz"]
                    entradas = openpyxl.load_workbook(file, data_only=True)
                    worksheet = entradas["Plantio"]
                    plantios = []

                    # DB CONSULT
                    talhao_list = Talhao.objects.all()
                    variedade_list = Variedade.objects.all()
                    safra_2022_2023 = 0
                    safra = Safra.objects.all()[safra_2022_2023]
                    ciclo_1 = 0
                    ciclo_2 = 1
                    ciclo_3 = 2
                    ciclo = Ciclo.objects.all()[ciclo_3]

                    for col in worksheet.iter_rows(min_row=1, max_col=14, max_row=3000):
                        if col[1].value != None and col[0].value != "ID":
                            id_talhao = col[0].value
                            finalizado = True if col[2].value == "Sim" else False
                            (
                                area_colher,
                                data_plantio,
                                dap,
                                id_variedade,
                            ) = (
                                col[8].value,
                                col[10].value,
                                col[11].value,
                                col[13].value,
                            )

                            if id_talhao:
                                try:
                                    talhao_id = [
                                        x
                                        for x in talhao_list
                                        if x.id_unico == id_talhao
                                    ][0]
                                except Exception as e:
                                    print(f"id sem cadastro: {id_talhao}")
                            else:
                                talhao_id = 0
                            try:
                                variedade_id = [
                                    x for x in variedade_list if x.id == id_variedade
                                ][0]
                                print(
                                    safra,
                                    ciclo,
                                    talhao_id,
                                    talhao_id.id_unico,
                                    variedade_id,
                                    area_colher,
                                    data_plantio,
                                    finalizado,
                                    dap,
                                )
                            except Exception as e:
                                print(f"variedade sem cadastro: {id_variedade}")
                            try:
                                Plantio.objects.filter(
                                    talhao__id_unico=talhao_id.id_unico,
                                    safra=safra,
                                    ciclo=ciclo,
                                ).update(finalizado_colheita=finalizado)

                            except Exception as e:
                                print(
                                    f"Problema em Atualiar o plantio: {talhao_id} - {e}"
                                )

                    qs_plantio_finalizado = Plantio.objects.filter(
                        finalizado_colheita=True
                    )
                    total_plantado = Plantio.objects.aggregate(Sum("area_colheita"))

                    total_plantado_finalizado = Plantio.objects.filter(
                        finalizado_colheita=True
                    ).aggregate(Sum("area_colheita"))

                    qs_plantio = Plantio.objects.all()
                    serializer_plantio = PlantioSerializer(qs_plantio, many=True)

                    response = {
                        "msg": f"Parcelas Atualizadas com successo!!",
                        "total_return": len(qs_plantio),
                        "Area Total dos Talhoes Plantados": total_plantado,
                        "Area Total dos Talhoes Plantados Finalizados": total_plantado_finalizado,
                        "Quantidade Total de Parcelas Finalizadas": len(
                            qs_plantio_finalizado
                        ),
                        "dados": serializer_plantio.data,
                    }

                    return Response(response, status=status.HTTP_200_OK)
                except Exception as e:
                    response = {"message": f"Ocorreu um Erro: {e}"}
                    return Response(response, status=status.HTTP_400_BAD_REQUEST)
            else:
                response = {"message": "Arquivo desconhecido"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # Plantio.objects.filter(talhao__id_unico="7B13").update(veiculos_carregados=4)

    # --------------------- ---------------------- UPDATE PLANTIO API --------------------- ----------------------#
    @action(detail=False, methods=["GET"])
    def get_plantio_cronograma_programa(self, request):
        if request.user.is_authenticated:
            try:
                farmer_filter_id = 11
                projeto = Projeto.objects.filter(id=farmer_filter_id)[0].nome

                qs = Plantio.objects.values(
                    "id",
                    "talhao__id_talhao",
                    "talhao_id",
                    "talhao__fazenda__nome",
                    "variedade__nome_fantasia",
                    "area_colheita",
                    "data_plantio",
                    # "get_cronograma_programa",
                ).filter(programa_id=6)
                ids_list = qs.values("id")

                qs_filtered = Plantio.objects.filter(id__in=ids_list)

                crono_list = [
                    (x.get_cronograma_programa, x.get_dap) for x in qs_filtered
                ]

                print(crono_list[0][1])

                final_return = []
                for i in crono_list:
                    dict_up = qs.filter(id=i[0][0]["id"])[0]
                    dict_up.update({"DAP": i[1]})
                    dict_up.update({"cronograma": i[0]})
                    final_return.append(dict_up)

                # final_return = [
                #     {
                #         "parcela": i.talhao.id_talhao,
                #         "data plantio": i.data_plantio,
                #         "area plantio": i.area_colheita,
                #         "cultura": i.variedade.cultura.cultura,
                #         "variedade": i.variedade.variedade,
                #         "Cronograma": i.get_cronograma_programa,
                #     }
                #     for i in Plantio.objects.filter(
                #         safra__safra="2023/2024",
                #         ciclo__ciclo="1",
                #         talhao__fazenda_id=farmer_filter_id,
                #     )
                # ]

                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs),
                    "projeto": projeto,
                    "dados": final_return,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO APLICACOES API START --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_operacoes_detail(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
                print(safra_filter)
                print(cicle_filter)
                safra_filter = "2023/2024" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter

                cache_key_qs_planejamento = f"get_plantio_operacoes_detail_qs_planejamento_{safra_filter}_{cicle_filter}"
                print('cache_key:', cache_key_qs_planejamento)
                qs_planejamento = cache.get(cache_key_qs_planejamento)
                if not qs_planejamento:
                    qs_planejamento = (
                        PlannerPlantio.objects.select_related(
                            "projeto",
                            "cultura",
                            "ciclo",
                            "cultura",
                            "projeto__fazenda",
                            "safra",
                        )
                        .values(
                            "projeto",
                            "projeto__id",
                            "projeto__nome",
                            "cultura",
                            "variedade",
                            "safra__safra",
                            "ciclo__ciclo",
                            "start_date",
                            "area",
                        )
                        .filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
                    )
                    cache.set(cache_key_qs_planejamento, qs_planejamento, timeout=60*5*6)  # cache for 5 minutes

                cache_key_qs_plantio_get_plantio_operacoes_detail = f"get_plantio_operacoes_detail_qs_plantio_{safra_filter}_{cicle_filter}"
                print('cache_key:', cache_key_qs_plantio_get_plantio_operacoes_detail)
                qs_plantio = cache.get(cache_key_qs_plantio_get_plantio_operacoes_detail)
                if not qs_plantio:
                    qs_plantio = (
                        Plantio.objects.select_related(
                            "safra",
                            "ciclo",
                            "talhao",
                            "fazenda",
                            "programa",
                            "variedade",
                            "variedade__cultura",
                            "talhao__fazenda__fazenda",
                        )
                        .values(
                            "id",
                            "talhao__id_talhao",
                            "talhao__id_unico",
                            "talhao_id",
                            "safra__safra",
                            "ciclo__ciclo",
                            "talhao__fazenda__nome",
                            "talhao__fazenda__id",
                            "talhao__fazenda__fazenda__nome",
                            "talhao__fazenda__fazenda__capacidade_plantio_ha_dia",
                            "variedade__nome_fantasia",
                            "variedade__cultura__cultura",
                            "area_colheita",
                            "data_plantio",
                            "data_prevista_plantio",
                            "finalizado_plantio",
                            "programa",
                            "programa_id",
                            "programa__start_date",
                            "programa__end_date",
                            "programa__nome",
                        )
                        .filter(~Q(programa_id=None))
                        .filter(safra=s_dict[safra_filter], ciclo=c_dict[cicle_filter])
                        .filter(Q(data_plantio=None))
                        .filter(plantio_descontinuado=False)
                        .filter(finalizado_colheita=False)
                    )
                    cache.set(cache_key_qs_plantio_get_plantio_operacoes_detail, qs_plantio, timeout=60*5*6)  # cache for 5 minutes

                qs_programas = Operacao.objects.values(
                    "estagio", "programa_id", "prazo_dap", "id"
                ).filter(ativo=True)

                cache_key_qs_aplicacoes = f"get_plantio_operacoes_detail_qs_aplicacoes_{safra_filter}_{cicle_filter}"
                print('cache_key:', cache_key_qs_aplicacoes)
                qs_aplicacoes = cache.get(cache_key_qs_aplicacoes)
                if not qs_aplicacoes:
                    qs_aplicacoes = (
                        Aplicacao.objects.select_related(
                            "defensivo",
                            "operacao",
                            "operacao__programa",
                            "operacao__programa__safra",
                            "operacao__programa__ciclo",
                            "operacao__programa__cultura",
                        )
                        .values(
                            "defensivo__produto",
                            "defensivo__tipo",
                            "defensivo__id_farmbox",
                            "dose",
                            "operacao",
                            "operacao__estagio",
                            "operacao__prazo_dap",
                            "operacao__programa",
                            "operacao__programa__nome",
                            "operacao__programa__safra__safra",
                            "operacao__programa__ciclo__ciclo",
                            "operacao__programa__cultura__cultura",
                        )
                        .filter(~Q(operacao__programa_id=None))
                        .filter(
                            operacao__programa__safra=s_dict[safra_filter],
                            operacao__programa__ciclo=c_dict[cicle_filter],
                            ativo=True,
                        )
                    )
                    cache.set(cache_key_qs_aplicacoes, qs_aplicacoes, timeout=60*5*6)  # cache for 5 minutes

                final_result = {i["talhao__fazenda__nome"]: {} for i in qs_plantio}
                try:
                    {
                        final_result[i["talhao__fazenda__nome"]].update(
                            {
                                i["talhao__id_talhao"]: {
                                    "safra": i["safra__safra"],
                                    "ciclo": i["ciclo__ciclo"],
                                    "cultura": i["variedade__cultura__cultura"],
                                    "variedade": i["variedade__nome_fantasia"],
                                    "plantio_id": i["id"],
                                    "fazenda_grupo": i[
                                        "talhao__fazenda__fazenda__nome"
                                    ],
                                    "projeto_id": i["talhao__fazenda__id"],
                                    "talhao_id_unico": i["talhao__id_unico"],
                                    "plantio_finalizado": i["finalizado_plantio"],
                                    "area_colheita": i["area_colheita"],
                                    "data_plantio": i["data_plantio"],
                                    "data_prevista_plantio": i["data_prevista_plantio"],
                                    "dap": get_dap(i["data_plantio"]),
                                    "programa_id": i["programa"],
                                    "programa": i["programa__nome"],
                                    "programa_start_date": i["programa__start_date"],
                                    "programa_end_date": i["programa__end_date"],
                                    "capacidade_plantio_dia": i[
                                        "talhao__fazenda__fazenda__capacidade_plantio_ha_dia"
                                    ],
                                    "cronograma": [
                                        {
                                            "estagio": x["estagio"],
                                            "dap": x["prazo_dap"],
                                            "data prevista": get_prev_app_date(
                                                i["data_plantio"], x["prazo_dap"]
                                            ),
                                            "produtos": [
                                                {
                                                    "produto": y["defensivo__produto"],
                                                    "tipo": y["defensivo__tipo"],
                                                    "dose": y["dose"],
                                                    "id_farmbox": y["defensivo__id_farmbox"],
                                                    "quantidade aplicar": get_quantidade_aplicar(
                                                        y["dose"], i["area_colheita"]
                                                    ),
                                                }
                                                for y in qs_aplicacoes
                                                if x["programa_id"] == i["programa"]
                                                and y["operacao"] == x["id"]
                                            ],
                                        }
                                        for x in qs_programas
                                        if x["programa_id"] == i["programa"]
                                    ],
                                }
                            }
                        )
                        for i in qs_plantio
                    }
                except Exception as e:
                    print("erro ao gerar", e)

                prev_date = {}
                # 50 ha por dia
                # max_day = 50
                today_check = datetime.datetime.now().date()
                # PROGRAMA PARA GERAR DATAS FUTURAS DE ACORDO COM A LÓGICA PARA
                for k, v in final_result.items():
                    prev_date.update(
                        {
                            k: {
                                "area": 0,
                                "dias_necessários": 0,
                                "data_inicial": None,
                                "data_final": None,
                            }
                        }
                    )
                    # print("inside loop : ", k)
                    filtered_planner = qs_planejamento.filter(projeto__nome=k)
                    planner_date = False
                    if filtered_planner:
                        # print("aqui temos planejamneto: , ", filtered_planner)
                        current_planner = filtered_planner[0]
                        inital_date_planner = current_planner["start_date"]
                        planner_date = True

                    for kk, vv in v.items():
                        # print("VVV :", vv)
                        # print("\n")
                        projeto_id = vv["projeto_id"]

                        data_plantio = vv["data_plantio"]
                        data_prevista_plantio = vv["data_prevista_plantio"]
                        area_colheita = vv["area_colheita"]
                        start_date = vv["programa_start_date"]
                        end_date = vv["programa_end_date"]
                        cronograma = vv["cronograma"]
                        capacidade_dia = vv["capacidade_plantio_dia"]

                        prev_date[k]["area"] += vv["area_colheita"]
                        prev_date[k]["dias_necessários"] = round(
                            prev_date[k]["area"] / capacidade_dia
                        )
                        if planner_date:
                            prev_date[k]["data_inicial"] = get_base_date(
                                inital_date_planner
                            )
                        else:
                            prev_date[k]["data_inicial"] = get_base_date(start_date)
                        prev_date[k]["data_final"] = prev_date[k][
                            "data_inicial"
                        ] + datetime.timedelta(days=prev_date[k]["dias_necessários"])
                        # ------------HERE is the challenge-------------------#
                        # print('data Prevista: ', data_prevista_plantio)
                        # print('data Prevista: ', type(data_prevista_plantio))
                        # print('data Prevista check: ', today_check)
                        # print('data Prevista check: ', type(today_check))
                        # print(data_prevista_plantio > today_check)
                        print('\n')
                        if data_plantio is None:
                            final_result[k][kk].update(
                                {
                                    "data_plantio": data_prevista_plantio if data_prevista_plantio and data_prevista_plantio > today_check else  prev_date[k]["data_inicial"]
                                    + datetime.timedelta(
                                        days=prev_date[k]["dias_necessários"]
                                    )
                                }
                            )
                        prev_date[k]["dias_necessários"] = round(
                            prev_date[k]["area"] / capacidade_dia
                        )
                        index = 0
                        for vvv in final_result[k][kk]["cronograma"]:
                            final_result[k][kk]["cronograma"][index].update(
                                {
                                    "data prevista": final_result[k][kk]["data_plantio"]
                                    + datetime.timedelta(days=vvv["dap"] - 1),
                                    "aplicado": False,
                                }
                            )
                            index += 1
                print(qs_planejamento)

                # RESUMO POR DIA COM TOTAIS DE APLICACOES
                final_by_day = []

                final = []

                def find_exs(x):
                    if x["data"]:
                        return True
                    else:
                        return False

                for k, v in final_result.items():
                    for kk, vv in v.items():
                        final.append({"fazenda": k, "parcela": kk, "dados": vv})
                        cronograma = vv["cronograma"]
                        for kkk in cronograma:
                            data_prev_cronograma = kkk["data prevista"]
                            produtos_cronograma = kkk["produtos"]
                            exists = [
                                [x["data"], ind]
                                for ind, x in enumerate(final_by_day)
                                if x["data"] == data_prev_cronograma
                            ]
                            if exists:
                                indice_data_existente = exists[0][1]
                                dict_to_update = final_by_day[exists[0][1]]["produtos"]

                                for i in produtos_cronograma:
                                    filtered = [
                                        (index, x["produto"], x["quantidade"])
                                        for index, x in enumerate(
                                            final_by_day[exists[0][1]]["produtos"]
                                        )
                                        if x["produto"] == i["produto"]
                                    ]

                                    # add quantity to product that already exists
                                    if filtered:
                                        index_of = filtered[0][0]
                                        value_of_upda = filtered[0][2]
                                        final_by_day[exists[0][1]]["produtos"][
                                            index_of
                                        ].update(
                                            {
                                                "produto": i["produto"],
                                                "tipo": i["tipo"],
                                                "id_farmbox": i["id_farmbox"],
                                                "quantidade": i["quantidade aplicar"]
                                                + value_of_upda,
                                            }
                                        )
                                    else:
                                        dict_to_update.append(
                                            {
                                                "produto": i["produto"],
                                                "tipo": i["tipo"],
                                                "id_farmbox": i["id_farmbox"],
                                                "quantidade": i["quantidade aplicar"],
                                            }
                                        )
                            else:
                                final_by_day.append(
                                    {
                                        "data": data_prev_cronograma,
                                        "produtos": [
                                            {
                                                "produto": x["produto"],
                                                "quantidade": x["quantidade aplicar"],
                                                "tipo": x["tipo"],
                                                "id_farmbox": x["id_farmbox"],
                                            }
                                            for x in produtos_cronograma
                                        ],
                                    }
                                )

                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "prev_date": prev_date,
                    "app_date": final_by_day,
                    "total_query_plantio": qs_plantio.count(),
                    "total_return_plantio": len(final_result),
                    "dados": final,
                    # "dados_plantio": qs_plantio,
                    # "total_return_aplicacoes": len(qs_aplicacoes),
                    # "dados_aplicacoes": qs_aplicacoes,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO APLICACOES API END --------------------- ----------------------#

    # --------------------- ---------------------- PLANTIO BIO APLICACOES API START --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_operacoes_detail_biologico_api(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
                print(safra_filter)
                print(cicle_filter)
                safra_filter = "2023/2024" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter
                cicle_fitler = [cicle_filter]

                # cicle_fitler = ["1", "2", "3"]
                qs_planejamento = (
                    PlannerPlantio.objects.select_related(
                        "projeto",
                        "cultura",
                        "ciclo",
                        "cultura",
                        "projeto__fazenda",
                        "safra",
                    )
                    .values(
                        "projeto",
                        "projeto__id",
                        "projeto__nome",
                        "cultura",
                        "variedade",
                        "safra__safra",
                        "ciclo__ciclo",
                        "start_date",
                        "area",
                    )
                    .filter(safra__safra=safra_filter)
                    .filter(ciclo__ciclo__in=cicle_fitler)
                )

                qs_plantio = (
                    Plantio.objects.select_related(
                        "safra",
                        "ciclo",
                        "talhao",
                        "fazenda",
                        "programa",
                        "variedade",
                        "variedade__cultura",
                        "talhao__fazenda__fazenda",
                    )
                    .values(
                        "id",
                        "talhao__id_talhao",
                        "talhao__id_unico",
                        "talhao_id",
                        "safra__safra",
                        "ciclo__ciclo",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__id",
                        "talhao__fazenda__fazenda__nome",
                        "talhao__fazenda__fazenda__capacidade_plantio_ha_dia",
                        "variedade__nome_fantasia",
                        "variedade__cultura__cultura",
                        "area_colheita",
                        "data_plantio",
                        "data_prevista_plantio",
                        "finalizado_plantio",
                        "programa",
                        "programa_id",
                        "programa__start_date",
                        "programa__end_date",
                        "programa__nome",
                    )
                    .filter(~Q(programa_id=None))
                    .filter(safra=s_dict[safra_filter])
                    .filter(ciclo__ciclo__in=cicle_fitler)
                    .filter(Q(data_plantio=None))
                    .filter(plantio_descontinuado=False)
                    .filter(finalizado_colheita=False)
                )

                qs_programas = Operacao.objects.values(
                    "estagio", "programa_id", "prazo_dap", "id"
                ).filter(ativo=True)

                qs_aplicacoes = (
                    Aplicacao.objects.select_related(
                        "defensivo",
                        "operacao",
                        "operacao__programa",
                        "operacao__programa__safra",
                        "operacao__programa__ciclo",
                        "operacao__programa__cultura",
                    )
                    .values(
                        "defensivo__produto",
                        "defensivo__tipo",
                        "defensivo__id_farmbox",
                        "dose",
                        "operacao",
                        "operacao__estagio",
                        "operacao__prazo_dap",
                        "operacao__programa",
                        "operacao__programa__nome",
                        "operacao__programa__safra__safra",
                        "operacao__programa__ciclo__ciclo",
                        "operacao__programa__cultura__cultura",
                    )
                    .filter(~Q(operacao__programa_id=None))
                    .filter(
                        operacao__programa__safra=s_dict[safra_filter],
                        ativo=True,
                    )
                    .filter(operacao__programa__ciclo__ciclo__in=cicle_fitler)
                    .filter(defensivo__tipo='biologico')
                )
                # TODO
                print('tamanho do qs plantio: ', len(qs_plantio))
                final_result = {i["talhao__fazenda__nome"]: {} for i in qs_plantio}
                try:
                    {
                        final_result[i["talhao__fazenda__nome"]].update(
                            {
                                i["talhao__id_talhao"]: {
                                    "safra": i["safra__safra"],
                                    "ciclo": i["ciclo__ciclo"],
                                    "cultura": i["variedade__cultura__cultura"],
                                    "variedade": i["variedade__nome_fantasia"],
                                    "plantio_id": i["id"],
                                    "fazenda_grupo": i[
                                        "talhao__fazenda__fazenda__nome"
                                    ],
                                    "projeto_id": i["talhao__fazenda__id"],
                                    "talhao_id_unico": i["talhao__id_unico"],
                                    "plantio_finalizado": i["finalizado_plantio"],
                                    "area_colheita": i["area_colheita"],
                                    "data_plantio": i["data_plantio"],
                                    "data_prevista_plantio": i["data_prevista_plantio"],
                                    "dap": get_dap(i["data_plantio"]),
                                    "programa_id": i["programa"],
                                    "programa": i["programa__nome"],
                                    "programa_start_date": i["programa__start_date"],
                                    "programa_end_date": i["programa__end_date"],
                                    "capacidade_plantio_dia": i[
                                        "talhao__fazenda__fazenda__capacidade_plantio_ha_dia"
                                    ],
                                    "cronograma": [
                                        {
                                            "estagio": x["estagio"],
                                            "dap": x["prazo_dap"],
                                            "data prevista": get_prev_app_date(
                                                i["data_plantio"], x["prazo_dap"]
                                            ),
                                            "produtos": [
                                                {
                                                    "produto": y["defensivo__produto"],
                                                    "tipo": y["defensivo__tipo"],
                                                    "dose": y["dose"],
                                                    "id_farmbox": y["defensivo__id_farmbox"],
                                                    "quantidade aplicar": get_quantidade_aplicar(
                                                        y["dose"], i["area_colheita"]
                                                    ),
                                                }
                                                for y in qs_aplicacoes
                                                if x["programa_id"] == i["programa"]
                                                and y["operacao"] == x["id"]
                                            ],
                                        }
                                        for x in qs_programas
                                        if x["programa_id"] == i["programa"]
                                    ],
                                }
                            }
                        )
                        for i in qs_plantio
                    }
                except Exception as e:
                    print("erro ao gerar", e)

                prev_date = {}
                # 50 ha por dia
                # max_day = 50

                # print('final result :', final_result)

                # PROGRAMA PARA GERAR DATAS FUTURAS DE ACORDO COM A LÓGICA PARA
                for k, v in final_result.items():
                    # print('chave: ', k)
                    # for parcela, dados in v.items():
                    #     print(f'parcela: {parcela} - Cronograma: ', dados['cronograma'])
                    # print('\n')
                    prev_date.update(
                        {
                            k: {
                                "area": 0,
                                "dias_necessários": 0,
                                "data_inicial": None,
                                "data_final": None,
                            }
                        }
                    )
                    filtered_planner = qs_planejamento.filter(projeto__nome=k)
                    planner_date = False
                    if filtered_planner:
                        # print("aqui temos planejamneto: , ", filtered_planner)
                        current_planner = filtered_planner[0]
                        inital_date_planner = current_planner["start_date"]
                        planner_date = True

                    for kk, vv in v.items():
                        # print("VVV :", vv)
                        # print("\n")
                        projeto_id = vv["projeto_id"]

                        data_plantio = vv["data_plantio"]
                        data_prevista_plantio = vv["data_prevista_plantio"]
                        area_colheita = vv["area_colheita"]
                        start_date = vv["programa_start_date"]
                        end_date = vv["programa_end_date"]
                        cronograma = vv["cronograma"]
                        capacidade_dia = vv["capacidade_plantio_dia"]

                        prev_date[k]["area"] += vv["area_colheita"]
                        prev_date[k]["dias_necessários"] = round(
                            prev_date[k]["area"] / capacidade_dia
                        )
                        if planner_date:
                            prev_date[k]["data_inicial"] = get_base_date(
                                inital_date_planner
                            )
                        else:
                            prev_date[k]["data_inicial"] = get_base_date(start_date)
                        prev_date[k]["data_final"] = prev_date[k][
                            "data_inicial"
                        ] + datetime.timedelta(days=prev_date[k]["dias_necessários"])
                        # ------------HERE is the challenge-------------------#
                        if data_plantio is None:
                            final_result[k][kk].update(
                                {
                                    "data_plantio": data_prevista_plantio if data_prevista_plantio else  prev_date[k]["data_inicial"]
                                    + datetime.timedelta(
                                        days=prev_date[k]["dias_necessários"]
                                    )
                                }
                            )
                        prev_date[k]["dias_necessários"] = round(
                            prev_date[k]["area"] / capacidade_dia
                        )
                        index = 0
                        for vvv in final_result[k][kk]["cronograma"]:
                            final_result[k][kk]["cronograma"][index].update(
                                {
                                    "data prevista": final_result[k][kk]["data_plantio"]
                                    + datetime.timedelta(days=vvv["dap"] - 1),
                                    "aplicado": False,
                                }
                            )
                            index += 1

                # RESUMO POR DIA COM TOTAIS DE APLICACOES
                final_by_day = []

                final = []

                def find_exs(x):
                    if x["data"]:
                        return True
                    else:
                        return False
                count_prod = 0
                for k, v in final_result.items():
                    for kk, vv in v.items():
                        final.append({"fazenda": k, "parcela": kk, "dados": vv})
                        cronograma = vv["cronograma"]
                        for kkk in cronograma:
                            data_prev_cronograma = kkk["data prevista"]
                            produtos_cronograma = kkk["produtos"]
                            exists = [
                                [x["data"], ind]
                                for ind, x in enumerate(final_by_day)
                                if x["data"] == data_prev_cronograma
                            ]
                            if exists:
                                indice_data_existente = exists[0][1]
                                dict_to_update = final_by_day[exists[0][1]]["produtos"]

                                for i in produtos_cronograma:
                                    filtered = [
                                        (index, x["produto"], x["quantidade"])
                                        for index, x in enumerate(
                                            final_by_day[exists[0][1]]["produtos"]
                                        )
                                        if x["produto"] == i["produto"]
                                    ]

                                    # add quantity to product that already exists
                                    if filtered:
                                        index_of = filtered[0][0]
                                        value_of_upda = filtered[0][2]
                                        final_by_day[exists[0][1]]["produtos"][
                                            index_of
                                        ].update(
                                            {
                                                "produto": i["produto"],
                                                "tipo": i["tipo"],
                                                "id_farmbox": i["id_farmbox"],
                                                "quantidade": i["quantidade aplicar"] + value_of_upda,
                                            }
                                        )
                                    else:
                                        dict_to_update.append(
                                            {
                                                "produto": i["produto"],
                                                "tipo": i["tipo"],
                                                "id_farmbox": i["id_farmbox"],
                                                "quantidade": i["quantidade aplicar"],
                                            }
                                        )
                            else:
                                final_by_day.append(
                                    {
                                        "data": data_prev_cronograma,
                                        "produtos": [
                                            {
                                                "produto": x["produto"],
                                                "quantidade": x["quantidade aplicar"],
                                                "tipo": x["tipo"],
                                                "id_farmbox": x["id_farmbox"],
                                            }
                                            for x in produtos_cronograma
                                        ],
                                    }
                                )
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "prev_date": prev_date,
                    "app_date": final_by_day,
                    "total_query_plantio": qs_plantio.count(),
                    "total_return_plantio": len(final_result),
                    # "dados": final,
                    # "dados_plantio": qs_plantio,
                    # "total_return_aplicacoes": len(qs_aplicacoes),
                    # "dados_aplicacoes": qs_aplicacoes,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO APLICACOES API END --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_operacoes_detail_json_program(self, request):
        if request.user.is_authenticated:

            start_time = time.time()  # Start time
            try:
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
                device = None
                if request.data.get("device") is not None:
                    device = request.data['device']
                print('device:', device)
                print('safra filter: ', safra_filter)
                print('cicle filter: ',cicle_filter)
                if device is None:
                    safra_filter = ['2024/2025','2025/2026']    
                    cicle_filter = ['1','2','3']
                else:
                    safra_filter = ["2024/2025"] if safra_filter == None else [safra_filter]
                    cicle_filter = "1" if cicle_filter == None else cicle_filter
                qs_start_time = time.time()
                if device == 'WEB':
                    cache_key = f"get_plantio_operacoes_detail_json_program_qs_plantio_web_{safra_filter[0]}_{cicle_filter}"
                    print('cache_key web:', cache_key)
                    qs_plantio = cache.get(cache_key)
                    if not qs_plantio:
                        qs_plantio = (
                            Plantio.objects.select_related(
                                "safra",
                                "ciclo",
                                "talhao",
                                "fazenda",
                                "programa",
                                "variedade",
                                "variedade__cultura",
                                "talhao__fazenda",
                                "talhao__fazenda__fazenda",
                            )
                            .values(
                                "id",
                                "talhao__id_talhao",
                                "talhao__id_unico",
                                "talhao_id",
                                "id_farmbox",
                                "safra__safra",
                                "safra__id_farmbox",
                                "ciclo__ciclo",
                                "talhao__fazenda__nome",
                                "talhao__fazenda__id_farmbox",
                                "talhao__fazenda__fazenda__id_responsavel_farmbox",
                                "talhao__fazenda__fazenda__id_encarregado_farmbox",
                                "talhao__fazenda__fazenda__nome",
                                "talhao__fazenda__fazenda__capacidade_plantio_ha_dia",
                                "variedade__nome_fantasia",
                                "variedade__cultura__cultura",
                                "area_colheita",
                                "data_plantio",
                                "finalizado_plantio",
                                "programa",
                                "programa_id",
                                "programa__start_date",
                                "programa__end_date",
                                "programa__nome",
                                "programa__nome_fantasia",
                                "cronograma_programa",
                            )
                            .filter(~Q(programa_id=None))
                            .filter(safra__safra__in=safra_filter)
                            .filter(ciclo=c_dict[cicle_filter])
                            .filter(data_plantio__isnull=False)
                            .filter(plantio_descontinuado=False)
                            .filter(finalizado_colheita=False)
                        )
                        cache.set(cache_key, qs_plantio, timeout=60*5*6)  # cache for 60 minutes
                else:
                    cache_key = f"get_plantio_operacoes_detail_json_program_qs_plantio_{safra_filter}_{cicle_filter}"
                    print('cache_key:', cache_key)
                    qs_plantio = cache.get(cache_key)
                    if not qs_plantio:
                        qs_plantio = (
                            Plantio.objects.select_related(
                                "safra",
                                "ciclo",
                                "talhao",
                                "fazenda",
                                "programa",
                                "variedade",
                                "variedade__cultura",
                                "talhao__fazenda",
                                "talhao__fazenda__fazenda",
                            )
                            .values(
                                "id",
                                "talhao__id_talhao",
                                "talhao__id_unico",
                                "talhao_id",
                                "id_farmbox",
                                "safra__safra",
                                "safra__id_farmbox",
                                "ciclo__ciclo",
                                "talhao__fazenda__nome",
                                "talhao__fazenda__id_farmbox",
                                "talhao__fazenda__fazenda__id_responsavel_farmbox",
                                "talhao__fazenda__fazenda__id_encarregado_farmbox",
                                "talhao__fazenda__fazenda__nome",
                                "talhao__fazenda__fazenda__capacidade_plantio_ha_dia",
                                "variedade__nome_fantasia",
                                "variedade__cultura__cultura",
                                "area_colheita",
                                "data_plantio",
                                "finalizado_plantio",
                                "programa",
                                "programa_id",
                                "programa__start_date",
                                "programa__end_date",
                                "programa__nome",
                                "programa__nome_fantasia",
                                "cronograma_programa",
                            )
                            .filter(~Q(programa_id=None))
                            .filter(safra__safra__in=safra_filter)
                            .filter(data_plantio__isnull=False)
                            .filter(plantio_descontinuado=False)
                            .filter(finalizado_colheita=False)
                        )
                        cache.set(cache_key, qs_plantio, timeout=60*5*12)  # cache for 60 minutes
                qs_end_time = time.time()
                print(f"Time for database query: {qs_end_time - qs_start_time:.2f} seconds")
                process_start_time = time.time()
                try:
                    result = [
                        {
                            "fazenda": i["talhao__fazenda__nome"],
                            "parcela": i["talhao__id_talhao"],
                            "plantio_id_farmbox": i["id_farmbox"],
                            "dados": {
                                "safra": i["safra__safra"],
                                "safra_id_farmbox": i["safra__id_farmbox"],
                                "ciclo": i["ciclo__ciclo"],
                                "cultura": i["variedade__cultura__cultura"],
                                "variedade": i["variedade__nome_fantasia"],
                                "plantio_id": i["id"],
                                "projeto_id_farmbox": i["talhao__fazenda__id_farmbox"],
                                "responsavel_id_farmbox": i["talhao__fazenda__fazenda__id_responsavel_farmbox"],
                                "encarregado_id_farmbox": i["talhao__fazenda__fazenda__id_encarregado_farmbox"],
                                "fazenda_grupo": i["talhao__fazenda__fazenda__nome"],
                                "talhao_id_unico": i["talhao__id_unico"],
                                "plantio_finalizado": i["finalizado_plantio"],
                                "area_colheita": i["area_colheita"],
                                "data_plantio": i["data_plantio"],
                                "data_inicio_plantio": i["cronograma_programa"][0][
                                    "Data Plantio"
                                ],
                                "dap": get_dap(
                                    i["cronograma_programa"][0]["Data Plantio"]
                                ),
                                "programa_id": i["programa"],
                                "programa": i["programa__nome"],
                                "programa_start_date": i["programa__start_date"],
                                "programa_end_date": i["programa__end_date"],
                                "capacidade_plantio_dia": i[
                                    "talhao__fazenda__fazenda__capacidade_plantio_ha_dia"
                                ],
                                "cronograma": [
                                    {
                                        **x,
                                        "estagio": x["estagio"]
                                        + "|"
                                        + i["programa__nome_fantasia"],
                                    }
                                    for x in i["cronograma_programa"][1:]
                                ],
                            },
                        }
                        for i in qs_plantio
                    ]
                    process_end_time = time.time()
                    print(f"Time for data processing: {process_end_time - process_start_time:.2f} seconds")
                except Exception as e:
                    print(f"erro ao gerar o cronograma {e}")
                end_time = time.time()
                print(f"Total time: {end_time - start_time:.2f} seconds")
                response = {
                    "msg": f"Retorno com os arquivos Json com Sucesso!!",
                    "total_query_plantio": qs_plantio.count(),
                    "dados_plantio": result,
                }
                end_time = time.time()  # End time
                print(f"Tempo total da requisição: {end_time - start_time:.2f} segundos")
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO APLICACOES TESTE JSON FIELD API START --------------------- ----------------------#

    # --------------------- ---------------------- PLANTIO MAP GEO API END --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_detail_map(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = None
                cicle_filter = None
                try:
                    safra_filter = request.data["safra"]
                    cicle_filter = request.data["ciclo"]
                except Exception as e:
                    print(e)
                print(safra_filter)
                print(cicle_filter)
                safra_filter = "2023/2024" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter

                cache_key_qs_plantio_map = f"get_plantio_map_{safra_filter}_{cicle_filter}"
                print('cache_key:', cache_key_qs_plantio_map)
                qs_plantio = cache.get(cache_key_qs_plantio_map)
                if not qs_plantio:
                    qs_plantio = (
                        Plantio.objects.select_related(
                            "safra",
                            "ciclo",
                            "talhao",
                            "fazenda",
                            "programa",
                            "variedade",
                            "variedade__cultura",
                            "talhao__fazenda",
                            "talhao__fazenda__fazenda",
                        )
                        .values(
                            "id",
                            "modificado",
                            "talhao__id_talhao",
                            "talhao__id_unico",
                            "talhao_id",
                            "safra__safra",
                            "ciclo__ciclo",
                            "talhao__fazenda__nome",
                            "talhao__fazenda__map_centro_id",
                            "talhao__fazenda__map_zoom",
                            "talhao__fazenda__fazenda__nome",
                            "variedade__nome_fantasia",
                            "variedade__cultura__cultura",
                            "variedade__cultura__map_color",
                            "variedade__cultura__map_color_line",
                            "finalizado_plantio",
                            "finalizado_colheita",
                            "plantio_descontinuado",
                            "area_colheita",
                            "map_centro_id",
                            "map_geo_points",
                        )
                        .filter(safra=s_dict[safra_filter], ciclo=c_dict[cicle_filter])
                        .filter(plantio_descontinuado=False)
                        # .filter(variedade__cultura__isnull=False)
                    )
                    cache.set(cache_key_qs_plantio_map, qs_plantio, timeout=60*5*6)  # cache for 5 minutes
                result = [
                    {
                        "fazenda": i["talhao__fazenda__nome"],
                        "parcela": i["talhao__id_talhao"],
                        "dados": {
                            "safra": i["safra__safra"],
                            "ciclo": i["ciclo__ciclo"],
                            "cultura": i["variedade__cultura__cultura"],
                            "variedade": i["variedade__nome_fantasia"],
                            "plantio_id": i["id"],
                            "finalizado_plantio": i["finalizado_plantio"],
                            # "finalizado_colheita": True if i["modificado"] < datetime.datetime.strptime("2023-12-30", '%Y-%m-%d') else False,
                            "finalizado_colheita": i["finalizado_colheita"],
                            "plantio_descontinuado": i["plantio_descontinuado"],
                            "fazenda_grupo": i["talhao__fazenda__fazenda__nome"],
                            "talhao_id_unico": i["talhao__id_unico"],
                            "area_colheita": i["area_colheita"],
                            "map_geo_points": i["map_geo_points"],
                            "map_geo_points_center": i["map_centro_id"],
                            "variedade_color": i["variedade__cultura__map_color"],
                            "variedade_color_line": i[
                                "variedade__cultura__map_color_line"
                            ],
                            "projeto_map_centro_id": i[
                                "talhao__fazenda__map_centro_id"
                            ],
                            "projeto_map_zoom": i["talhao__fazenda__map_zoom"],
                        },
                    }
                    for i in qs_plantio
                ]

                response = {
                    "msg": f"Retorno com os arquivos DE MAPAS com Sucesso!!",
                    "total_query_plantio": qs_plantio.count(),
                    "dados_plantio": result,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO UPDATE APLICATION FIELD API START --------------------- ----------------------#

    # --------------------- ---------------------- PLANTIO PRODUTIVIDADE MAP API START --------------------- ----------------------#
    @action(detail=False, methods=["GET", "POST"])
    def get_produtividade_plantio(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = None
                cicle_filter = None
                try:
                    safra_filter = request.data["safra"]
                    cicle_filter = request.data["ciclo"]
                except Exception as e:
                    print(e)
                safra_filter = "2023/2024" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter
                qs_colheita = Colheita.objects.values(
                    "plantio__id",
                    "plantio__talhao__fazenda__nome",
                    "plantio__talhao__id_talhao",
                    "plantio__area_colheita",
                    "plantio__finalizado_colheita",
                    "plantio__area_parcial",
                ).annotate(
                    peso_kg=Round((Sum("peso_scs_limpo_e_seco") * 60), precision=2),
                    peso_scs=Round(Sum("peso_scs_limpo_e_seco"), precision=2),
                    # produtividade=Round(
                    #     ("peso_scs" / "plantio__area_colheita"),
                    #     precision=2,
                    #     output_field=DecimalField(),
                    # ),
                )
                qs_plantio = Plantio.objects.values(
                    "id",
                    "modificado",
                    "talhao__id_talhao",
                    "talhao__id_unico",
                    "talhao_id",
                    "safra__safra",
                    "ciclo__ciclo",
                    "talhao__fazenda__nome",
                    "talhao__fazenda__map_centro_id",
                    "talhao__fazenda__map_zoom",
                    "talhao__fazenda__fazenda__nome",
                    "variedade__nome_fantasia",
                    "variedade__cultura__cultura",
                    "variedade__cultura__map_color",
                    "variedade__cultura__map_color_line",
                    "finalizado_plantio",
                    "finalizado_colheita",
                    "area_colheita",
                    "area_parcial",
                    "map_centro_id",
                    "map_geo_points",
                ).filter(
                    safra=s_dict[safra_filter],
                    ciclo=c_dict[cicle_filter],
                    plantio_descontinuado=False,
                ).filter(variedade__cultura__isnull=False)

                result = [x for x in qs_plantio]
                for i in qs_colheita:
                    for j in result:
                        # dateFilt = "2023-12-30"
                        # j["finalizado_colheita"] = True if j["modificado"] < datetime.datetime.strptime(dateFilt, '%Y-%m-%d') else False
                        if i["plantio__id"] == j["id"]:
                            j["peso_kg"] = i["peso_kg"]
                            j["peso_scs"] = i["peso_scs"]
                            if i["plantio__finalizado_colheita"] == True:
                                j["produtividade"] = (
                                    i["peso_scs"] / i["plantio__area_colheita"]
                                )
                            else:
                                area_parcial = (
                                    i["plantio__area_parcial"]
                                    if i["plantio__area_parcial"]
                                    else None
                                )
                                j["produtividade"] = (
                                    i["peso_scs"] / area_parcial if area_parcial else 0
                                )

                response = {
                    "msg": f"Retorno com os arquivos DE MAPAS e Produtividades com Sucesso!!",
                    # "total_query_plantio": qs_plantio.count(),
                    # "dados_colheita": qs_colheita,
                    "dados_plantio": result,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO PRODUTIVIDADE MAP API END --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST", "PUT"])
    def update_aplication_plantio(sef, request, pk=None):
        if request.user.is_authenticated:
            try:
                params = request.data["data"]
                print(params)
                id_list = [x["id"] for x in params]
                print("list of IDS: ", id_list)
                list_updated = []
                updated_instances = []
                for i in params:
                    try:
                        update_field = Plantio.objects.get(pk=i["id"])
                        index = get_index_dict_estagio(
                            update_field.cronograma_programa, i["estagio"]
                        )
                        print(
                            f"{update_field.talhao.fazenda.nome} - {update_field.talhao.id_talhao}"
                        )
                        print(update_field.cronograma_programa[index]["estagio"])
                        print(update_field.cronograma_programa[index]["aplicado"])
                        print("\n")
                        field_to_update = update_field.cronograma_programa[index][
                            "aplicado"
                        ]

                        # new_value = None
                        # if field_to_update == True:
                        #     new_value = False

                        # if field_to_update == False:
                        #     new_value = True

                        new_value = not field_to_update

                        update_field.cronograma_programa[index]["aplicado"] = new_value

                        update_field.save()
                        updated_instances.append(update_field)

                        updated = {
                            "talhao": update_field.talhao.id_talhao,
                            "estagio": update_field.cronograma_programa[index][
                                "estagio"
                            ],
                            "situação": update_field.cronograma_programa[index][
                                "aplicado"
                            ],
                        }
                        print("updated at: ", updated)
                        list_updated.append(updated)
                    except Exception as e:
                        print("Erro ao atualizar a Ap no DB", e)
                # Ensure atomicity using transaction.atomic
                
                with transaction.atomic():        
                    # Plantio.objects.bulk_update(updated_instances, ['cronograma_programa'])
                    for instance in updated_instances[0:1]:
                        post_save.send(sender=Plantio, instance=instance, created=False)
                        print('signals sent!!')

                response = {
                    "msg": "Atualização realizada com sucesso!",
                    "data": params,
                    "result": list_updated,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    def save_app_farmbox(self, parsed_json):
        try:
            app_id_farm = parsed_json["plantations"][0]['plantation']['farm_name']
            number_app = parsed_json['code']
            name_app_op = parsed_json['inputs'][0]['input']['name']
            app_name = f'{number_app} - {name_app_op}'
            nova_app = AppFarmboxIntegration(
                app_nuumero=app_name,
                app_fazenda=app_id_farm,
                app=parsed_json
            )
            nova_app.save()
            print("AppFarmboxIntegration saved successfully!")
        except Exception as e:
            print('Problem saving new Application:', e)

    # TODO
    @action(detail=False, methods=["GET", "POST", "PUT"])
    def open_app_farmbox(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                start_total = time.perf_counter()
                params = request.data["data"]
                print('Abrindo Aplicação')
                # print('Params from Farmbox: ', params)
                start_inputs = time.perf_counter()
                old_inputs = params.get('inputs')
                print('old paramsss::', params)
                print('\n')
                new_inputs = [
                    {**item, "input_id": 156297} if item.get("input_id") == 154220 else item
                    for item in old_inputs
                ]
                params['inputs'] = new_inputs
                logger.info('inputs processing time: %.4fs', time.perf_counter() - start_inputs)
                # logger.info('using new inputs:', params)
                
                url = "https://farmbox.cc/api/v1/applications"
                payload = params
                headers = {
                    "content-type": "application/json",
                    "Authorization": FARMBOX_ID,
                }
                start_request = time.perf_counter()
                response_farm = requests.post(url, data=json.dumps(payload), headers=headers)
                logger.info('request time: %.4fs', time.perf_counter() - start_request)
                
                logger.info(f'responseAll from farmbox: {response_farm}')
                print('\n\n')
                logger.info(f'response: {response_farm.status_code}')

                if response_farm.status_code == 201:
                    parsed_json = json.loads(response_farm.text)
                    
                    # START SAVE HERE
                    logger.info('Start save app from  FarmBox')
                    Thread(target=self.save_app_farmbox, args=(parsed_json,)).start()
                    logger.info('Finish off save app from  FarmBox')
                    logger.info('total time: %.4fs', time.perf_counter() - start_total)
                    response = {
                        "msg": "APP Aberta com sucesso!!",
                        "data": params,
                        "status": response_farm.status_code,
                        "result": response_farm.text,
                    }
                    return Response(response, status=status.HTTP_201_CREATED)
                else:
                    response = {
                        "msg": "Problema em abrir a Aplicação!!",
                        "data": params,
                        "status": response_farm.status_code,
                        "result": response_farm.text,
                    }
                    return Response(response, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET", "POST"])
    def get_plantio_calendar_done(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                safra_filter = "2023/2024"
                cicle_filter = "3"
                if request.data["safra"]:
                    safra_filter = request.data["safra"]
                if request.data["ciclo"]:
                    cicle_filter = request.data["ciclo"]
                qs = (
                    Plantio.objects.filter(
                        safra__safra=safra_filter,
                        finalizado_plantio=True,
                        plantio_descontinuado=False,
                    )
                    .filter(~Q(variedade__cultura__cultura="Milheto"))
                    .values(
                        "talhao__fazenda__nome",
                        "variedade__cultura__cultura",
                        "variedade__variedade",
                        "ciclo__ciclo",
                    )
                    .annotate(
                        month=ExtractMonth("data_plantio"),
                        year=ExtractYear("data_plantio"),
                        area_total=Sum("area_colheita"),
                    )
                    .order_by("year", "ciclo__ciclo", "month", "talhao__fazenda__nome")
                )
                response = {
                    "msg": "Consulta realizada com sucesso!!",
                    "data": qs,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- PLANTIO UPDATE APLICATION FIELD API END --------------------- ----------------------#

    # @action(detail=False, methods=["GET"])
    # def get_new_information(self, request, pk=None):
    #     if request.user.is_authenticated:
    #         try:
    #             print("successfull")
    # data = mongo_api.read_data_from_db(db_name)
    #             response = {
    #                 "msg": f"Consulta realizada com sucesso!!",
    #                 "dados": data,
    #             }
    #             return Response(response, status=status.HTTP_200_OK)
    #         except Exception as e:
    #             print("error")
    #             response = f"erro na requisição: {e}"
    #             return Response(response, status=status.HTTP_400_BAD_REQUEST)
    #     else:
    #         response = {"message": "Você precisa estar logado!!!"}
    #         return Response(response, status=status.HTTP_400_BAD_REQUEST)

    # --------------------- ---------------------- DEFENSIVOS API START --------------------- ----------------------#

    @action(detail=False, methods=["GET", "POST"])
    def get_colheita_plantio_info(self, request, pk=None):
        if request.user.is_authenticated:
            print('pegando dados da colheira: ')
            safra_filter = None
            cicle_filter = None
            try:
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
            except Exception as e:
                print(e)
            safra_filter = "2023/2024" if safra_filter == None else safra_filter
            cicle_filter = "1" if cicle_filter == None else cicle_filter
            print('safra e ciclo filtr', safra_filter, cicle_filter)
            total_dias_plantado_acompanhamento = {
                "arroz": 117,
                "soja_feijao": 10
            }
            try:
                cargas_query = (
                    Colheita.objects.select_related(
                        "plantio__talhao__fazenda",
                        "plantio__talhao",
                    )
                    .values(
                        "plantio__talhao__id_talhao",
                        "plantio__id",
                        "plantio__talhao__fazenda__nome",
                    )
                    .annotate(
                        total_peso_liquido=Sum("peso_liquido"),
                        total_romaneio=Count("romaneio"),
                    )
                    .annotate(
                        totaldays=(
                            datetime.datetime.now().date() - F("plantio__data_plantio")
                        )
                    )
                    .filter(
                        plantio__safra=s_dict[safra_filter],
                        plantio__ciclo=c_dict[cicle_filter],
                        plantio__finalizado_plantio=True,
                        # plantio__finalizado_colheita=False,
                        plantio__plantio_descontinuado=False,
                        totaldays__gte=datetime.timedelta(days=total_dias_plantado_acompanhamento["soja_feijao"]),
                        plantio__acompanhamento_medias=True
                        # ARROZ = 117
                        # totaldays__gte=datetime.timedelta(days=117),
                    )
                )
                qs = (
                    Plantio.objects.select_related(
                        "talhao__fazenda",
                        "safra",
                        "ciclo",
                        "variedade",
                        "variedade__cultura",
                    )
                    .values(
                        "id",
                        "talhao__id_talhao",
                        "talhao__id_unico",
                        "data_plantio",
                        "safra__safra",
                        "ciclo__ciclo",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__map_centro_id",
                        "talhao__fazenda__map_zoom",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__nome_fantasia",
                        "variedade__cultura__cultura",
                        "variedade__cultura__map_color",
                        "variedade__cultura__map_color_line",
                        "finalizado_plantio",
                        "finalizado_colheita",
                        "area_colheita",
                        "area_parcial",
                    )
                    .annotate(
                        totaldays=(datetime.datetime.now().date() - F("data_plantio"))
                    )
                    .filter(
                        safra=s_dict[safra_filter],
                        ciclo=c_dict[cicle_filter],
                        finalizado_plantio=True,
                        # finalizado_colheita=False,
                        plantio_descontinuado=False,
                        totaldays__gte=datetime.timedelta(days=total_dias_plantado_acompanhamento["soja_feijao"]),
                        acompanhamento_medias=True
                    )
                    .order_by("talhao__id_unico")
                )

                response = {
                    "msg": "Consulta realizada com sucesso!!",
                    "len_data": qs.count(),
                    "data": qs,
                    "len_cargas": cargas_query.count(),
                    "cargas": cargas_query,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_401_UNAUTHORIZED)
    
    def group_data(self, data):
        grouped_data = []
        data = sorted(data, key=lambda x: x["talhao__fazenda__nome"])  # Group by farm name
        
        for farm_name, items in groupby(data, key=lambda x: x["talhao__fazenda__nome"]):
            items = list(items)
            farm_group = {
                "farm": farm_name,
                "colheita": sum(item["area_colheita"] or 0 for item in items),
                "parcial": sum(item["area_parcial"] or 0 for item in items),
                "variedades": [],
                "culturas": [],  # New key to aggregate by culture
                "peso_liquido": sum(item['cargas'][0]['total_peso_liquido'] for item in items if item.get('cargas') or 0)
            }
            
            # Group by variety
            variety_totals = {}
            for item in items:
                variety = item["variedade__nome_fantasia"]
                if variety not in variety_totals:
                    variety_totals[variety] = {
                        "colheita": 0,
                        "parcial": 0,
                        "cultura": item["variedade__cultura__cultura"],
                    }
                variety_totals[variety]["colheita"] += item["area_colheita"] or 0
                variety_totals[variety]["parcial"] += item["area_parcial"] or 0

            # Add variety totals to the farm group
            farm_group["variedades"] = [
                {
                    "variedade": v,
                    "colheita": t["colheita"],
                    "parcial": t["parcial"],
                    "cultura": t["cultura"],
                    "percent": round((t["parcial"] / t["colheita"] * 100), 2) if t["colheita"] > 0 else 0,
                    
                }
                for v, t in variety_totals.items()
            ]
            
            # Group by culture
            cultura_totals = {}
            for item in items:
                cultura = item["variedade__cultura__cultura"]
                if cultura not in cultura_totals:
                    cultura_totals[cultura] = {"colheita": 0, "parcial": 0}
                cultura_totals[cultura]["colheita"] += item["area_colheita"] or 0
                cultura_totals[cultura]["parcial"] += item["area_parcial"] or 0

            # Add culture totals with percentage to the farm group
            farm_group["culturas"] = [
                {
                    "cultura": c,
                    "colheita": t["colheita"],
                    "parcial": t["parcial"],
                    "percent": round((t["parcial"] / t["colheita"] * 100), 2) if t["colheita"] > 0 else 0,
                }
                for c, t in cultura_totals.items()
            ]

            grouped_data.append(farm_group)

        return grouped_data


    @action(detail=False, methods=["GET", "POST"])
    def get_colheita_plantio_info_react_native(self, request, pk=None):
        if request.user.is_authenticated:
            print('pegando dados da colheira: ')
            safra_filter = None
            cicle_filter = None
            try:
                # safra_filter = request.data["safra"]
                # cicle_filter = request.data["ciclo"]
                # print('safra filter: ', safra_filter)
                # print('cicle filter: ', cicle_filter)
                
                safracicle_filter = CicloAtual.objects.filter(nome="Colheita")[0]
                safra_filter = safracicle_filter.safra.safra
                cicle_filter = str(safracicle_filter.ciclo.ciclo)
                
            except Exception as e:
                print(e)
            safra_filter = "2023/2024" if safra_filter == None else safra_filter
            cicle_filter = "1" if cicle_filter == None else cicle_filter
            print('safra e ciclo filtr', safra_filter, cicle_filter)
            total_dias_plantado_acompanhamento = {
                "arroz": 117,
                "soja_feijao": 10
            }
            try:
                cargas_query = (
                    Colheita.objects.select_related(
                        "plantio__talhao__fazenda",
                        "plantio__talhao",
                    )
                    .values(
                        "plantio__talhao__id_talhao",
                        "plantio__id",
                        "plantio__talhao__fazenda__nome",
                    )
                    .annotate(
                        total_peso_liquido=Sum("peso_liquido"),
                        total_romaneio=Count("romaneio"),
                    )
                    .annotate(
                        totaldays=(
                            datetime.datetime.now().date() - F("plantio__data_plantio")
                        )
                    )
                    .filter(
                        plantio__safra=s_dict[safra_filter],
                        plantio__ciclo=c_dict[cicle_filter],
                        plantio__finalizado_plantio=True,
                        # plantio__finalizado_colheita=False,
                        plantio__plantio_descontinuado=False,
                        plantio__acompanhamento_medias=True,
                        # totaldays__gte=datetime.timedelta(days=total_dias_plantado_acompanhamento["soja_feijao"]),
                        # ARROZ = 117
                        # totaldays__gte=datetime.timedelta(days=117),
                    )
                )
                qs = (
                    Plantio.objects.select_related(
                        "talhao__fazenda",
                        "safra",
                        "ciclo",
                        "variedade",
                        "variedade__cultura",
                    )
                    .values(
                        "id",
                        "talhao__id_talhao",
                        "talhao__id_unico",
                        "data_plantio",
                        "safra__safra",
                        "ciclo__ciclo",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__map_centro_id",
                        "talhao__fazenda__map_zoom",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__nome_fantasia",
                        "variedade__dias_ciclo",
                        "variedade__cultura__cultura",
                        "variedade__cultura__map_color",
                        "variedade__cultura__map_color_line",
                        "finalizado_plantio",
                        "finalizado_colheita",
                        "area_colheita",
                        "area_parcial",
                    )
                    .annotate(
                        totaldays=(datetime.datetime.now().date() - F("data_plantio"))
                    )
                    .filter(
                        safra=s_dict[safra_filter],
                        ciclo=c_dict[cicle_filter],
                        finalizado_plantio=True,
                        # finalizado_colheita=False,
                        plantio_descontinuado=False,
                        acompanhamento_medias=True
                        # totaldays__gte=datetime.timedelta(days=total_dias_plantado_acompanhamento["soja_feijao"]),
                    )
                    .order_by("talhao__id_unico")
                )
                
                # Iterate through the 'data' array
                for item in qs:
                    # Find matching 'cargas' based on the 'id' in 'data' and 'plantio__id' in 'cargas'
                    matching_cargas = [carga for carga in cargas_query if carga["plantio__id"] == item["id"]]
                    
                    # If there are matching 'cargas', insert them into the 'data' object
                    if matching_cargas:
                        item["cargas"] = matching_cargas
                
                filter_data = {}
                filter_data_farm = [x['talhao__fazenda__fazenda__nome'] for x in qs]
                filter_data_proj = [x['talhao__fazenda__nome'] for x in qs]
                filter_data_variety = [{'variety': x['variedade__nome_fantasia'] ,'culture': x['variedade__cultura__cultura'] } for x in qs]
                
                
                filter_data['farm'] = list(set(filter_data_farm))
                filter_data['proj'] = list(set(filter_data_proj))
                filter_data['variety'] = list({tuple(d.items()): d for d in filter_data_variety}.values())
                
                grouped_data = self.group_data(qs)
                
                response = {
                    "msg": "Consulta realizada com sucesso!!",
                    "data": qs,
                    "grouped_data": grouped_data,
                    "filter_data": filter_data
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_401_UNAUTHORIZED)
        
    def create_kml(self, queryset):
        kml_header = '''<?xml version="1.0" encoding="UTF-8"?>
                    <kml xmlns="http://www.opengis.net/kml/2.2">
                        <Document>
                            <Style id="style1">
                                <LineStyle>
                                    <color>80000000</color>  <!-- Black fill, fully opaque -->
                                    <width>2</width>
                                </LineStyle>
                                <PolyStyle>
                                    <color>80ffffff</color>  <!-- White line with 80% opacity -->
                                    
                                </PolyStyle>
                                <IconStyle>
                                    <scale>0</scale>  <!-- Set scale to 0 to hide the icon -->
                                </IconStyle>
                            </Style>
                    '''
        kml_footer = '</Document>\n</kml>\n'

        placemarks = ""

        # Loop through each item in the queryset
        for item in queryset:
            # Extract the points for each item
            coordinates = []
            for point in item['map_geo_points']:  # Assuming `map_geo_points` is a JSON field or similar
                lat = point['latitude']
                lng = point['longitude']
                coordinates.append(f"{lng},{lat}")  # KML format is lng,lat
            coordinates.append(f"{item['map_geo_points'][0]['longitude']},{item['map_geo_points'][0]['latitude']}")  # KML format is lng,lat
            
            # Calculate the center of the polygon for the label
            center_lat = sum(float(point['latitude']) for point in item['map_geo_points']) / len(item['map_geo_points'])
            center_lng = sum(float(point['longitude']) for point in item['map_geo_points']) / len(item['map_geo_points'])
            
            
            # Create a placemark for each item with additional details
            # Create a placemark for the polygon
            polygon_placemark = f'''
            <Placemark>
                <name>{item['talhao__id_talhao']}</name> <!-- This is the polygon name -->
                <description>Farmbox ID: {item['id_farmbox']}</description>
                <styleUrl>#style1</styleUrl>
                <Polygon>
                    <outerBoundaryIs>
                        <LinearRing>
                            <coordinates>{" ".join(coordinates)}</coordinates>
                        </LinearRing>
                    </outerBoundaryIs>
                </Polygon>
            </Placemark>
            '''

            # Create a separate placemark for the label
            label_placemark = f'''
            <Placemark>
                <name>{item['talhao__id_talhao']}</name> <!-- This is the label -->
                <styleUrl>#style1</styleUrl>  <!-- You can use the same style or define a new one -->
                <Point>
                    <coordinates>{center_lng},{center_lat}</coordinates> <!-- Center coordinates for the label -->
                </Point>
                <LabelStyle>
                    <color>ff0000ff</color>  <!-- Red label (fully opaque) -->
                    <scale>1.2</scale>  <!-- Adjust the label size -->
                </LabelStyle>
            </Placemark>
            '''

            placemarks += polygon_placemark + label_placemark

        return kml_header + placemarks + kml_footer
    
    @action(detail=False, methods=["GET", "POST"])
    def get_kmls_aviacao(self, request, pk=None):
        # get id_farmbox
        projeto_filter = request.data["projeto"]
        parcelas_filter = request.data["parcelas"]
        safra_filter = "2024/2025"
        ciclo_filter = "1"
        try:
            filter_safra_and_ciclo = parcelas_filter[0]
            refer_plantio = Plantio.objects.get(id_farmbox=filter_safra_and_ciclo)
            print('refer plantio Safra: ', refer_plantio.safra.safra)
            print('refer plantio ciclo: ', refer_plantio.ciclo.ciclo)
            safra_filter = refer_plantio.safra.safra
            ciclo_filter = refer_plantio.ciclo.ciclo
        except Exception as e:
            print('error em filtrar safra e ciclo: ', e)
        for i in parcelas_filter:
            print(i)

        start_time = time.time()
        print(start_time)
        try:
            plantio_map = Plantio.objects.values(
                "map_geo_points", "map_centro_id", "talhao__id_talhao", "id_farmbox"
            ).filter(
                safra__safra=safra_filter,
                ciclo__ciclo=ciclo_filter,
                # finalizado_plantio=True,
                # programa__isnull=False,
                talhao__fazenda__id_farmbox=projeto_filter,
            ).order_by('data_prevista_plantio')
            
            if parcelas_filter:
                poligons_to_export = plantio_map.filter(id_farmbox__in=parcelas_filter)
                print('poligons to exportL ', poligons_to_export)
                try:
                    kml_content = self.create_kml(poligons_to_export)
                except Exception as e:
                    print('erro ao gerar o KML: ', e)


            # Prepare the KML content as a base64 encoded string if needed
            kml_data = base64.b64encode(kml_content.encode()).decode()
            kml_data_uri = f"data:application/vnd.google-earth.kml+xml;base64,{kml_data}"

            response = {
                "msg": "Mapa gerado com sucesso!!",
                "data": {
                    # Include the image data URI
                    "kml": kml_data_uri   
                },
            }
            return Response(
                response,
                content_type="application/json",  # Use JSON response type
                status=status.HTTP_200_OK,
            )

            # POSTMAN
            # return HttpResponse(
            #     img_buffer.getvalue(),
            #     content_type="image/png",
            #     status=status.HTTP_200_OK,
            # )
        except Exception as e:
            response = {"message": f"Ocorreu um Erro: {e}"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET", "POST"])
    def get_matplot_draw(self, request, pk=None):
        # get id_farmbox
        projeto_filter = request.data["projeto"]
        parcelas_filter = request.data["parcelas"]
        safra_filter = "2024/2025"
        ciclo_filter = "3"
        planejamento_plantio = False

        try:
            filter_safra_and_ciclo = parcelas_filter[0]
            refer_plantio = Plantio.objects.get(id_farmbox=filter_safra_and_ciclo)
            print('refer plantio Safra: ', refer_plantio.safra.safra)
            print('refer plantio ciclo: ', refer_plantio.ciclo.ciclo)
            safra_filter = refer_plantio.safra.safra
            ciclo_filter = refer_plantio.ciclo.ciclo
        except Exception as e:
            print('error em filtrar safra e ciclo: ', e)
        print(parcelas_filter)
        for i in parcelas_filter:
            print(i)

        start_time = time.time()
        print(start_time)
        try:
            if planejamento_plantio == True:
                safra_filter = "2024/2025"
                ciclo_filter = '3'
                filter_farm_id= "6"

                plantio_map = Plantio.objects.values(
                    "map_geo_points", "map_centro_id", "talhao__id_talhao", "id_farmbox"
                ).filter(
                    safra__safra=safra_filter,
                    ciclo__ciclo=ciclo_filter,
                    # finalizado_plantio=True,
                    # programa__isnull=False,
                    talhao__fazenda__fazenda__id_d=filter_farm_id,
                ).order_by('data_prevista_plantio')

                plantio_ids = Plantio.objects.values(
                    "id_farmbox", "talhao__id_talhao"
                ).filter(
                    safra__safra=safra_filter,
                    # finalizado_plantio=True,
                    # programa__isnull=False,
                    talhao__fazenda__fazenda__id_d=filter_farm_id,
                ).order_by('data_prevista_plantio')

                gp_date = ( 
                        Plantio.objects.filter(safra__safra=safra_filter, ciclo__ciclo=ciclo_filter, programa__isnull=False)
                                        .filter(talhao__fazenda__fazenda__id_d=filter_farm_id)
                                        .annotate(date_only=TruncDate('data_prevista_plantio'))
                                        .values('date_only', 'id_farmbox')
                                        .order_by('date_only')
                )
                grouped_by_date = [
                    {
                        'index': idx, 'date': date, 'list': [item['id_farmbox'] for item in group]
                    }
                    for idx, (date, group) in enumerate(groupby(gp_date, key=itemgetter('date_only')), start=1)
                ]

                print('Resultado dos campos consolidados')
                for i in grouped_by_date:
                    print(i)
                    print('\n')

            else:
                grouped_by_date=[]
                plantio_map = Plantio.objects.values(
                    "map_geo_points", "map_centro_id", "talhao__id_talhao", "id_farmbox"
                ).filter(
                    safra__safra=safra_filter,
                    ciclo__ciclo=ciclo_filter,
                    # finalizado_plantio=True,
                    # programa__isnull=False,
                    talhao__fazenda__id_farmbox=projeto_filter,
                ).order_by('data_prevista_plantio')

                plantio_ids = Plantio.objects.values(
                    "id_farmbox", "talhao__id_talhao"
                ).filter(
                    safra__safra=safra_filter,
                    # finalizado_plantio=True,
                    # programa__isnull=False,
                    talhao__fazenda__id_farmbox=projeto_filter,
                ).order_by('data_prevista_plantio')

            print("depois de fazer a query")
            print(time.time() - start_time)
            center_id = []
            polygons = []
            labels = []
            farm_ids = []
            ids_farmbox= []
            print("entrando no loop")
            print(time.time() - start_time)
            for i in plantio_map:
                center = [
                    float(i["map_centro_id"]["lat"]),
                    (float(i["map_centro_id"]["lng"]) * -1),
                ]
                center_id.append(center)

                np = [
                    [float(x["latitude"]), (float(x["longitude"]) * -1)]
                    for x in i["map_geo_points"]
                ]
                polygons.append(np)

                labels.append(i["talhao__id_talhao"])
                farm_ids.append(i["talhao__id_talhao"])
                ids_farmbox.append(i["id_farmbox"])

            print("saindo do loop")
            print(time.time() - start_time)

            parcelas_filter = [
                x["talhao__id_talhao"]
                for x in plantio_ids
                if x["id_farmbox"] in parcelas_filter
            ]
            print('parcelas filter', parcelas_filter)

            img_buffer = draw_cartoon_map(
                polygons=polygons,
                labels=labels,
                centerid=center_id,
                ids_farmbox=farm_ids,
                filled_polygon_index=parcelas_filter,
                filled_color=(0, 0.96, 0, 0.7),
                # filled_color="#4191C4",
                fontsize=3,
                planejamento_plantio=planejamento_plantio,
                grouped_by_date=grouped_by_date,
                ids_farmbox_planner=ids_farmbox,
            )

            data_img = base64.b64encode(img_buffer.getvalue()).decode()
            data_uri = f"data:image/png;base64,{data_img}"

            response = {
                "msg": "Mapa gerado com sucesso!!",
                "data": {
                    "center_ids": center_id,
                    "labels": labels,
                    "polygons": polygons,
                },
            }
            return HttpResponse(
                data_uri,
                content_type="image/png",
                status=status.HTTP_200_OK,
            )

            # POSTMAN
            # return HttpResponse(
            #     img_buffer.getvalue(),
            #     content_type="image/png",
            #     status=status.HTTP_200_OK,
            # )
        except Exception as e:
            response = {"message": f"Ocorreu um Erro: {e}"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET", "POST"])
    def get_matplot_draw_application(self, request, pk=None):
        # get id_farmbox
        projeto_filter = request.data["projeto"]
        parcelas_filter = request.data["parcelas"]
        print(parcelas_filter)
        for i in parcelas_filter:
            print(i)

        start_time = time.time()
        print(start_time)
        try:
            plantio_map = Plantio.objects.values(
                "map_geo_points",
                "map_centro_id",
                "talhao__id_talhao",
                "id_farmbox",
                "id",
            ).filter(
                safra__safra="2023/2024",
                ciclo__ciclo="3",
                finalizado_plantio=True,
                talhao__fazenda__nome=projeto_filter,
            )

            print("depois de fazer a query")
            print(time.time() - start_time)
            center_id = []
            polygons = []
            labels = []
            farm_ids = []
            print("entrando no loop")
            print(time.time() - start_time)
            for i in plantio_map:
                center = [
                    float(i["map_centro_id"]["lat"]),
                    (float(i["map_centro_id"]["lng"]) * -1),
                ]
                center_id.append(center)

                np = [
                    [float(x["latitude"]), (float(x["longitude"]) * -1)]
                    for x in i["map_geo_points"]
                ]
                polygons.append(np)

                labels.append(i["talhao__id_talhao"])
                farm_ids.append(i["id"])

            print("saindo do loop")
            print(time.time() - start_time)

            img_buffer = draw_cartoon_map(
                polygons=polygons,
                labels=labels,
                centerid=center_id,
                ids_farmbox=farm_ids,
                filled_polygon_index=parcelas_filter,
                filled_color=(0, 0.96, 0, 0.7),
                fontsize=3,
            )

            data_img = base64.b64encode(img_buffer.getvalue()).decode()
            data_uri = f"data:image/png;base64,{data_img}"

            response = {
                "msg": "Mapa gerado com sucesso!!",
                "data": {
                    "center_ids": center_id,
                    "labels": labels,
                    "polygons": polygons,
                },
            }
            return HttpResponse(
                data_uri,
                content_type="image/png",
                status=status.HTTP_200_OK,
            )

            # POSTMAN
            # return HttpResponse(
            #     img_buffer.getvalue(),
            #     content_type="image/png",
            #     status=status.HTTP_200_OK,
            # )
        except Exception as e:
            response = {"message": f"Ocorreu um Erro: {e}"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET", "POST"])
    def get_bio_prods_open_and_planted(self, request, pk=None):
        # TODO
        # alterar a safra e ciclo, recebendo direto da requisição
        # verificar possibiliade de filtar somente safra e não ciclo
        if request.user.is_authenticated:
            try:
                safra_filter = request.data["safra"]
                safra_filter = safra_filter if safra_filter else "2024/2025"
                cicle_filter = request.data["ciclo"]
                print('cicle filter here: ', cicle_filter)
                # cicle_filter = cicle_filter if cicle_filter else "2"
                qs = (
                    Plantio.objects
                    .select_related(
                            "talhao__fazenda__fazenda",
                        )
                    .filter(
                        safra__safra=safra_filter,
                        # ciclo__ciclo=cicle_filter,
                        finalizado_plantio=True,
                        plantio_descontinuado=False,
                        programa__isnull=False,
                        finalizado_colheita=False
                    )
                    .filter(~Q(variedade__cultura__cultura="Milheto"))
                    .values(
                        "cronograma_programa",
                        "area_colheita",
                        "talhao__id_talhao",
                        "talhao__fazenda__nome"
                    )
                )
                bio_prods = []
                bio_prods_geral = []
                today = datetime.datetime.today()
                check_date = today + datetime.timedelta(days=15)

                # print('check_date: ', check_date)
                for i in qs:
                    cronograma = i['cronograma_programa']
                    area = i["area_colheita"]
                    parcela = i['talhao__id_talhao']
                    projeto = i['talhao__fazenda__nome']
                    # print(i)
                    # print('\n')
                    for program in cronograma[1:]:
                        if program['aplicado'] == False:
                            datetime_object = datetime.datetime.strptime(program['data prevista'], '%Y-%m-%d')
                            if datetime_object <=check_date:
                                check_prods = program['produtos']
                                for prods in check_prods:
                                    if prods['tipo'] == 'biologico':
                                        estagio = program['estagio']
                                        # print('estagio: ', estagio)
                                        # print('data Prevista: ',program["data prevista"])
                                        # print('check_date: ', check_date)
                                        # print('\n')
                                        # print(prods)
                                        prods_formated = prods
                                        prods_formated["quantidade aplicar"] = float(area) * float(prods["dose"])
                                        prods_formated["parcela"] = parcela
                                        prods_formated["projeto"] = projeto
                                        prods_formated["area"] = area
                                        prods_formated["estagio"] = estagio
                                        bio_prods.append(prods_formated)
                            check_prods_geral = program['produtos']
                            for prods in check_prods_geral:
                                if prods['tipo'] == 'biologico':
                                    estagio = program['estagio']
                                    # print('estagio: ', estagio)
                                    # print('data Prevista: ',program["data prevista"])
                                    # print('check_date: ', check_date)
                                    # print('\n')
                                    # print(prods)
                                    prods_formated_geral = prods
                                    prods_formated_geral["quantidade aplicar"] = float(area) * float(prods["dose"])
                                    prods_formated_geral["parcela"] = parcela
                                    prods_formated_geral["projeto"] = projeto
                                    prods_formated_geral["area"] = area
                                    prods_formated_geral["estagio"] = estagio
                                    bio_prods_geral.append(prods_formated_geral)

                # summed_quantities = defaultdict(int)
                # for obj in bio_prods:
                #     summed_quantities[obj['id_farmbox']] += obj['quantidade aplicar']
                # total_prods = [{'id': key, 'quantity': value} for key, value in summed_quantities.items()]
                # cout_total = 0
                # for i in total_prods:
                #     cout_total += i['quantity']
                # print('total: ', cout_total)
                # print(len(bio_prods))

                # summed_quantities_geral = defaultdict(int)
                # for obj in bio_prods_geral:
                #     summed_quantities_geral[obj['id_farmbox']] += obj['quantidade aplicar']
                # total_prods_geral = [{'id': key, 'quantity': value} for key, value in summed_quantities_geral.items()]
                # cout_total_geral = 0
                # for i in total_prods_geral:
                #     cout_total_geral += i['quantity']
                # print('total Geral: ', cout_total_geral)
                # print(len(bio_prods))
                response = {
                    "msg": "Consulta realizada com sucesso dos produtos previstos de Biológicos!!",
                    "data": bio_prods,
                    "data_geral": bio_prods_geral,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['GET', 'POST'])
    def save_planejamento_protheus(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                projetos_response=[]
                format_data_to_send=[]
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
                query_plantio = Plantio.objects.values(
                        "safra__safra",
                        "pk",
                        "ciclo__ciclo",
                        "talhao__id_talhao",
                        "talhao__fazenda__nome",
                        "talhao__fazenda__id_d",
                        "talhao__fazenda__fazenda__nome",
                        "variedade__cultura__cultura",
                        "variedade__cultura__id_protheus_planejamento",
                        "variedade__id_protheus_planejamento_second_option",
                        "variedade__variedade",
                        "variedade__id_farmbox",
                        "variedade__id_cultura_dif_protheus_planejamento",
                        "area_colheita",
                        'id_farmbox'
                    ).filter(
                        safra__safra=safra_filter,
                        ciclo__ciclo=cicle_filter,
                        plantio_descontinuado=False
                    ).filter(~Q(variedade=None) & ~Q(variedade__cultura=None))

                list_returned = [
                    {
                        'codigo_planejamento' : '??',
                        'codigo_cultura': x['variedade__cultura__id_protheus_planejamento'] if x['variedade__cultura__cultura'] != 'Feijão' else x['variedade__id_cultura_dif_protheus_planejamento'],
                        'variedade': x['variedade__variedade'],
                        "id_variedade": x['variedade__id_farmbox'],
                        'parcela': x['talhao__id_talhao'],
                        'area_parcela': x['area_colheita'],
                        'id_talhao_farm': x['id_farmbox'],
                        'projeto': x['talhao__fazenda__id_d'],
                        "projeto_nome": x["talhao__fazenda__nome"],
                        'index': index
                    }
                    for index, x in enumerate(query_plantio, start=1)
                ]

                # TRATAR PARA ENVIAR PROJETOS PARA ABRIR OS HEADERS DO PLANEJAMENTO
                farm_list = list(set([x['talhao__fazenda__nome'] for x in query_plantio]))
                query_projetos = Projeto.objects.values(
                    'nome',
                    'fazenda__nome',
                    'fazenda__id_d',
                    'id_d',
                )
                filtered_farm_list = [{**x, 'safra': safra_filter, 'ciclo': cicle_filter, 'projeto': x['id_d'], 'fazenda_planejamento': x['fazenda__nome'], 'id_fazenda_planejamento' : x['fazenda__id_d']} for x in query_projetos if x['nome'] in farm_list]

                print('dados para enviar', filtered_farm_list)

                # send to protheus saving headers
                headers = {
                    "Accept": "application/json",
                    "Content-Type": "application/json",
                    "Authorization": f'Basic ${PROTHEUS_TOKEN}',
                    "Access-Control-Allow-Origin": "*"
                }
                url = "https://api.diamanteagricola.com.br:8089/rest/planejamento/cabecalho"

                payload = {
                    "projetos": filtered_farm_list
                }
                if len(farm_list) > 0:
                    try:
                        response_headers = requests.post(url, data=json.dumps(payload), headers=headers, verify=False, auth=HTTPBasicAuth('api', PROTHEUS_TOKEN))

                        print('response headers from protheus:', response_headers)
                        print('\n\n')
                        print('response:', response_headers.status_code, response_headers.text)
                        if response_headers.status_code == 201:
                            safra_id = Safra.objects.all()
                            cicle_id = Ciclo.objects.all()
                            projetos_to_query = Projeto.objects.all()
                            parsed_json = json.loads(response_headers.text)
                            print('\n')
                            print('Projetos of Response:', parsed_json['Projetos'])
                            projetos_response = parsed_json['Projetos']
                            for resp in projetos_response:
                                if resp.get('codigo_planejamento') != None:
                                    try:
                                        ciclo_to_save = [x for x in cicle_id if x.ciclo == int(cicle_filter)][0]
                                        safra_to_save = [ x for x in safra_id if x.safra == safra_filter][0]
                                        projeto_to_save = [x for x in projetos_to_query if x.id_d == resp['projeto']][0]
                                        codigo_to_save = resp['codigo_planejamento']
                                        if not HeaderPlanejamentoAgricola.objects.filter(
                                            projeto=projeto_to_save,
                                            # codigo_planejamento=codigo_to_save,
                                            safra=safra_to_save,
                                            ciclo=ciclo_to_save
                                        ).exists():
                                            new_planner = HeaderPlanejamentoAgricola(
                                                projeto=projeto_to_save,
                                                codigo_planejamento=codigo_to_save,
                                                safra=safra_to_save,
                                                ciclo=ciclo_to_save
                                            )
                                            new_planner.save()
                                        print(f'{Fore.GREEN}Novo Planejamento incluido com sucesso!! - {Fore.BLUE} {new_planner}{Style.RESET_ALL}')
                                    except Exception as e:
                                        print(f'{Fore.LIGHTYELLOW_EX}Erro ao Salvar o Planejamento {Fore.LIGHTRED_EX}{e}{Style.RESET_ALL}')
                                else:
                                    print('Projeto com erro de resposta', resp)
                    except Exception as e:
                        print('Erro ao enviar o cabeçalho para o protheus', e)

                get_planner_codes = HeaderPlanejamentoAgricola.objects.values('projeto__id_d', 'codigo_planejamento').filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
                format_data_to_send = [{**x,'area_parcela': float(x['area_parcela']) ,'codigo_planejamento': [cod['codigo_planejamento'] for cod in get_planner_codes if cod['projeto__id_d'] == x['projeto']][0] } for x in list_returned ]
                url_talhoes = 'https://api.diamanteagricola.com.br:8089/rest/planejamento/talhoes'
                payload_talhoes = {
                    'parcelas': format_data_to_send
                }

                print('Talhos a enviar: \n')
                print(format_data_to_send)

                if len(format_data_to_send) > 0:
                    try:
                        response_talhoes = requests.post(url_talhoes,data=json.dumps(payload_talhoes), headers=headers, verify=False, auth=HTTPBasicAuth('api', PROTHEUS_TOKEN))
                        print('response headers from protheus:', response_talhoes)
                        print('\n\n')
                        print('response:', response_talhoes.status_code, response_talhoes.text)

                    except Exception as e:
                        print('Erro ao enviar os talhoes para o protheus', e)

                response = {
                    "msg": "Dados consolidados para enviar ao protheus com Sucesso!!",
                    "projetos": projetos_response,
                    "parcelas": format_data_to_send,
                }
                return Response(response, status=status.HTTP_201_CREATED)
            except Exception as e:
                print('Problema em gerar os dados para enviar ao protheus', e)
                response = {
                    "message": "Arquivo desconhecido",
                    "problem": e
                    }
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['POST'])
    def open_bulcket_app_farmbox(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                file = request.FILES["insumos"]
                dict_data = file.read()
                load_data = json.loads(dict_data)
                input_id = request.data['input_id']
                input_operation = request.data['input_operation']
                list_to_work = []
                dict_app = []
                today_date = datetime.datetime.today().strftime('%Y-%m-%d')

                query_projetos = Plantio.objects.filter(safra__safra="2024/2025").values('talhao__fazenda__id_farmbox','id_farmbox','talhao__fazenda__fazenda__id_responsavel_farmbox','talhao__fazenda__fazenda__id_encarregado_farmbox')

                for i in load_data:
                    projeto = i['Projeto']
                    area = i['Area']
                    id_plantio = i['ID FarmBox']
                    dose = i["Dose Kg"]
                    get_index = [(index, x) for index, x in enumerate(dict_app) if x['projeto'] == projeto and x['dose'] == dose]
                    if area:
                        plantation_to_append = {
                                        'sought_area': float(str(area).replace(',','.')),
                                        'plantation_id': id_plantio,
                                    }
                        if get_index:
                            index_of = get_index[0][0]
                            dict_app[index_of]['plantations'].append(plantation_to_append)
                        else:
                            get_id = query_projetos.filter(id_farmbox=id_plantio).first()
                            obj_to_add = {
                                'projeto': projeto,
                                'dose': dose,
                                'date': today_date,
                                'harvest_id': 3840,
                                'farm_id':get_id['talhao__fazenda__id_farmbox'],
                                'responsible_id': get_id['talhao__fazenda__fazenda__id_responsavel_farmbox'],
                                'charge_id': get_id['talhao__fazenda__fazenda__id_encarregado_farmbox'],
                                'plantations': [plantation_to_append],
                                'inputs' : [
                                    # OPERACAO
                                    {
                                        "dosage_value": 1,
                                        "dosage_unity": "un_ha",
                                        "input_id": input_operation
                                    },
                                    # PRODUTO
                                    {
                                        "dosage_value": dose,
                                        "dosage_unity": "kg_ha",
                                        "input_id": input_id
                                    }
                                ]
                            }
                            dict_app.append(obj_to_add)

                for i in dict_app:
                    i.pop('projeto')
                    i.pop('dose')   

                # for payload in dict_app:
                #     print(payload)
                #     print('\n')

                # for payload in dict_app[2:]:
                #     print(payload)
                #     print('\n')

                # LOGICA PARA ABRIR AS APS DENTRO DO FARM
                # for payload in dict_app:
                #     url = "https://farmbox.cc/api/v1/applications"
                #     payload = payload
                #     headers = {
                #         "content-type": "application/json",
                #         "Authorization": FARMBOX_ID,
                #     }
                #     response_farm = requests.post(url, data=json.dumps(payload), headers=headers)
                #     print('responseAll from farmbox:', response_farm)
                #     print('\n\n')
                #     print('response:', response_farm.status_code)

                #     if response_farm.status_code == 201:
                #         print('Ap Aberta com sucesso!!!')

                print('Total App to Open:')
                print(len(dict_app))

                data = {
                    'dados': 'dados tratados'
                }
                response = {
                    "msg": f"Aplicação Aberta com sucesso!!!!",
                    "dados": data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET"])
    def get_plantio_planner_data(self, request, *args, **kwargs):
        safra_filter = request.query_params.get("safra")
        cicle_filter = request.query_params.get("ciclo")
        
        
        print('safra filter', safra_filter)
        print('cicle filter', cicle_filter)
        try:
            qs_planned = (
                Plantio.objects.select_related(
                    "safra",
                    "ciclo",
                    "talhao",
                    "fazenda",
                    "variedade",
                    "variedade__cultura",
                    "talhao__fazenda__fazenda",
                )
                .values(
                    "area_planejamento_plantio",
                    "id_farmbox",
                    "data_prevista_plantio",
                    "data_plantio",
                    "finalizado_plantio",
                    "data_prevista_colheita",
                    "talhao__fazenda__nome",
                    "talhao__id_talhao",
                    "variedade__variedade",
                    'variedade__dias_ciclo',
                    "variedade__cultura__cultura"
                )
                # .filter(safra__safra="2024/2025", ciclo__ciclo="3")
                .filter(safra__safra=safra_filter)
                .filter(ciclo__ciclo=cicle_filter)
                .filter(plantio_descontinuado=False)
                .filter(variedade__variedade__isnull=False)
            )
            only_proj = list(set([x["talhao__fazenda__nome"] for x in qs_planned]))

            qs_executed = (
                PlantioExtratoArea.objects.select_related(
                    "plantio__safra",
                    "plantio__ciclo",
                    "plantio__talhao",
                    "plantio__fazenda",
                    "plantio__variedade",
                    "plantio__variedade__cultura",
                    "plantio__talhao__fazenda__fazenda",
                ).values(
                    "plantio__id_farmbox",
                    "plantio__talhao__fazenda__nome",
                    "plantio__talhao__id_talhao",
                    "plantio__variedade__variedade",
                    "plantio__variedade__cultura__cultura",
                    "data_plantio",
                    "area_plantada",
                    "aguardando_chuva"
                )
                # .filter(plantio__safra__safra="2024/2025", plantio__ciclo__ciclo="3")
                .filter(plantio__safra__safra=safra_filter)
                .filter(plantio__ciclo__ciclo=cicle_filter)
                .filter(plantio__plantio_descontinuado=False)
                .filter(ativo=True)
            )

            data = {
                "qs_planned_size": len(qs_planned),
                'qs_planned_area_total': qs_planned.aggregate(Sum('area_planejamento_plantio')),
                "qs_planned": qs_planned,
                "qs_planned_projetos": only_proj,
                "qs_executed_area": qs_executed,
                }
            response = {
                "msg": f"Aplicação Aberta com sucesso!!!!",
                "dados": data,
            }
            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            print("erro ao pegar os dados do plantio", e)
            response = {
                "msg": f"Erro gerar os dados do plantio",
                "error": f"Error ao pegar os dados do plantio , Erro: {e}",  # General error message
            }
            return Response(response, status=status.HTTP_208_ALREADY_REPORTED)

    @action(detail=False, methods=["GET"])
    def get_map_plot_app_fetch_app(self, request, *args, **kwargs):
        try:
            print('Start to get Data to plot MAP on APP ')
            query_data = (
                Plantio.objects.select_related(
                    "safra",
                    "ciclo",
                    "talhao",
                    "fazenda",
                    "programa",
                    "variedade",
                    "variedade__cultura",
                    "talhao__fazenda__fazenda",
                )
                .values(
                    "talhao__id_talhao",
                    "talhao__fazenda__nome",
                    "talhao__fazenda__id_farmbox",
                    "talhao__fazenda__map_centro_id",
                    "area_colheita",
                    "map_centro_id",
                    "id_farmbox",
                    "map_geo_points",
                    "map_centro_id",
                )
                .filter(safra__safra="2025/2026", ciclo__ciclo="1")
            )
            response = {
                "msg": f"Aplicação Aberta com sucesso!!!!",
                "dados": query_data,
            }
            print('Data to plot map successfully sent')
            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            print("erro ao pegar os dados do plantio", e)
            response = {
                "msg": f"Erro gerar os dados do plantio",
                "error": f"Error ao pegar os dados do plantio , Erro: {e}",  # General error message
            }
        return Response(response, status=status.HTTP_208_ALREADY_REPORTED)

    @action(detail=False, methods=["GET"])
    def get_sent_seeds_data(self, request, *args, **kwargs):
        try:
            safra_filter = '2024/2025'
            cicle_filter = '3'
            
            safra_filter = request.query_params.get("safra")
            cicle_filter = request.query_params.get("ciclo")
            
            # # GET TOTAL PLANTED
            qs_plantio = (
                PlantioExtratoArea.objects.select_related(
                    "plantio",
                    "plantio__talhao__fazenda",
                    "plantio__talhao__fazenda__fazenda",
                    "plantio__safra",
                    "plantio__ciclo",
                    "plantio__variedade",
                    "plantio__variedade__cultura"
                ).values(
                "plantio__talhao__fazenda__fazenda__nome",  # Group by fazenda__nome
                "plantio__variedade__variedade",  # Group by variedade
                "plantio__variedade__cultura__cultura",  # Group by variedade
            ).annotate(
                total_area_plantada=Sum("area_plantada")  # Sum the area_plantada for each group
            ).filter(plantio__safra__safra=safra_filter, plantio__ciclo__ciclo=cicle_filter)
            .filter(plantio__plantio_descontinuado=False)
            ).filter(ativo=True)
            
            qs_sent_seeds = (
                SentSeeds.objects.select_related(
                    "variedade", 
                    "fazenda",
                ).values(
                "destino__nome",  # Group by fazenda__nome
                "variedade__variedade",  # Group by variedade
                "variedade__cultura__cultura", 
            ).annotate(
                total_area_plantada=Sum("peso_total")  # Sum the area_plantada for each group
            ).filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
            .filter(ativo=True)
            )
            
            qs_sent_seeds_totals = (
                SentSeeds.objects.select_related(
                    "variedade", 
                    "fazenda",
                ).values(
                "variedade__variedade",  # Group by variedade
                "variedade__cultura__cultura", 
            ).annotate(
                total_area_plantada=Sum("peso_total")  # Sum the area_plantada for each group
            ).filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
            .filter(ativo=True)
            )
            
            
            # Subquery to get the last (most recent) data_apontamento for each variedade and fazenda
            # latest_stock = SeedStock.objects.filter(
            #     fazenda=OuterRef('fazenda'),
            #     variedade=OuterRef('variedade')
            # ).order_by('-data_apontamento').values('estoque_atual')[:1]

            # Query to get the latest estoque_atual for each variedade and fazenda
            seed_stocks = SeedStock.objects.filter(
                data_apontamento=Subquery(
                    SeedStock.objects.filter(
                        fazenda=OuterRef('fazenda'),
                        variedade=OuterRef('variedade')
                    ).order_by('-data_apontamento').values('data_apontamento')[:1]
                )
            ).values(
                'fazenda__nome', 'variedade__variedade', 'estoque_atual', 'data_apontamento'
            )
            
            seed_config = SeedConfig.objects.filter(
                data_apontamento=Subquery(
                    SeedConfig.objects.filter(
                        fazenda=OuterRef('fazenda'),
                        variedade=OuterRef('variedade')
                    ).order_by('-data_apontamento').values('data_apontamento')[:1]
                )
            ).values(
                'fazenda__nome', 'variedade__variedade', 'regulagem', 'data_apontamento'
            )
            
            
            # def fetch_qs_plantio(safra_filter, cicle_filter):
            #     return (
            #         PlantioExtratoArea.objects.select_related(
            #             "plantio",
            #             "plantio__talhao__fazenda",
            #             "plantio__talhao__fazenda__fazenda",
            #             "plantio__safra",
            #             "plantio__ciclo",
            #             "plantio__variedade",
            #             "plantio__variedade__cultura"
            #         ).values(
            #             "plantio__talhao__fazenda__fazenda__nome",  # Group by fazenda__nome
            #             "plantio__variedade__variedade"  # Group by variedade
            #         ).annotate(
            #             total_area_plantada=Sum("area_plantada")  # Sum the area_plantada for each group
            #         ).filter(plantio__safra__safra=safra_filter, plantio__ciclo__ciclo=cicle_filter)
            #     )

            # def fetch_qs_sent_seeds(safra_filter, cicle_filter):
            #     return (
            #         SentSeeds.objects.select_related(
            #             "variedade", 
            #             "fazenda",
            #         ).values(
            #             "destino__nome",  # Group by fazenda__nome
            #             "variedade__variedade"  # Group by variedade
            #         ).annotate(
            #             total_area_plantada=Sum("peso_total")  # Sum the area_plantada for each group
            #         ).filter(safra__safra=safra_filter, ciclo__ciclo=cicle_filter)
            #     )

            # def fetch_seed_stocks():
            #     return SeedStock.objects.filter(
            #         data_apontamento=Subquery(
            #             SeedStock.objects.filter(
            #                 fazenda=OuterRef('fazenda'),
            #                 variedade=OuterRef('variedade')
            #             ).order_by('-data_apontamento').values('data_apontamento')[:1]
            #         )
            #     ).values(
            #         'fazenda__nome', 'variedade__variedade', 'estoque_atual', 'data_apontamento'
            #     )

            # def fetch_seed_config():
            #     return SeedConfig.objects.filter(
            #         data_apontamento=Subquery(
            #             SeedStock.objects.filter(
            #                 fazenda=OuterRef('fazenda'),
            #                 variedade=OuterRef('variedade')
            #             ).order_by('-data_apontamento').values('data_apontamento')[:1]
            #         )
            #     ).values(
            #         'fazenda__nome', 'variedade__variedade', 'regulagem', 'data_apontamento'
            #     )

            # def fetch_all_queries(safra_filter, cicle_filter):
            #     with ThreadPoolExecutor() as executor:
            #         # Submitting each query to the executor
            #         future_qs_plantio = executor.submit(fetch_qs_plantio, safra_filter, cicle_filter)
            #         future_qs_sent_seeds = executor.submit(fetch_qs_sent_seeds, safra_filter, cicle_filter)
            #         future_seed_stocks = executor.submit(fetch_seed_stocks)
            #         future_seed_config = executor.submit(fetch_seed_config)

            #         # Fetching the results
            #         qs_plantio = future_qs_plantio.result()
            #         qs_sent_seeds = future_qs_sent_seeds.result()
            #         seed_stocks = future_seed_stocks.result()
            #         seed_config = future_seed_config.result()

            #     return qs_plantio, qs_sent_seeds, seed_stocks, seed_config

            # qs_plantio, qs_sent_seeds, seed_stocks, seed_config = fetch_all_queries(safra_filter, cicle_filter)
            
            # Initialize an empty list to hold the data for the table
            table_data = []

            # Create a set of all unique (fazenda, variedade) pairs from qs_plantio and qs_sent_seeds
            all_fazenda_variedade_pairs = set()

            # Add pairs from qs_plantio
            for plantio in qs_plantio:
                fazenda_nome = plantio["plantio__talhao__fazenda__fazenda__nome"]
                variedade_nome = plantio["plantio__variedade__variedade"]
                all_fazenda_variedade_pairs.add((fazenda_nome, variedade_nome))

            # Add pairs from qs_sent_seeds
            for sent_seed in qs_sent_seeds:
                fazenda_nome = sent_seed["destino__nome"]
                variedade_nome = sent_seed["variedade__variedade"]
                all_fazenda_variedade_pairs.add((fazenda_nome, variedade_nome))

            # Loop through all unique (fazenda, variedade) pairs
            for fazenda_nome, variedade_nome in all_fazenda_variedade_pairs:
                # Fetch data from both querysets based on the current fazenda and variedade
                plantio = qs_plantio.filter(plantio__talhao__fazenda__fazenda__nome=fazenda_nome, plantio__variedade__variedade=variedade_nome).order_by('plantio__talhao__fazenda__fazenda__nome').first()
                sent_seeds = qs_sent_seeds.filter(destino__nome=fazenda_nome, variedade__variedade=variedade_nome).order_by('destino__nome').first()
                seed_stock = seed_stocks.filter(fazenda__nome=fazenda_nome, variedade__variedade=variedade_nome).order_by('fazenda__nome').first()
                seed_config_value = seed_config.filter(fazenda__nome=fazenda_nome, variedade__variedade=variedade_nome).order_by('fazenda__nome').first()

                # Fill in values for each column
                peso_total = sent_seeds["total_area_plantada"] if sent_seeds else 0
                estoque = seed_stock["estoque_atual"] if seed_stock else 0
                utilizado = peso_total - estoque
                area_plantada = plantio["total_area_plantada"] if plantio else 0
                semente_ha = round((utilizado / area_plantada), 2) if area_plantada else 0
                ultima_reg = seed_config_value["regulagem"] if seed_config_value else 0

                if ultima_reg > 0:
                    date_string = seed_config_value['data_apontamento']
                    formatted_date = date_string.strftime("%d/%m/%Y")
                    ultima_reg = f"{formatted_date} - {str(ultima_reg).replace('.', ',')} Kg"

                # Add the row to the table data
                row = {
                    "Destino": fazenda_nome,
                    "Produto": variedade_nome,
                    "Cultura": sent_seeds["variedade__cultura__cultura"] if sent_seeds else '',
                    "Peso_Total": peso_total,
                    "Estoque": estoque,
                    "Utilizado": utilizado,
                    "Area_Plantada": area_plantada,
                    "Semente_Ha": semente_ha,
                    "Ultima_Regulagem": ultima_reg
                }

                table_data.append(row)
                        
            table_data.sort(key = lambda x : x['Destino'])
            data = {
                'qs_plantio': qs_plantio,
                "qs_sent_seed": qs_sent_seeds,
                "qs_sent_seeds_totals": qs_sent_seeds_totals,
                'qs_stock': seed_stocks,
                'qs_regulagem': seed_config,
                'query_table': table_data,
                
            }
            response = {
                "msg": f"Consulta sobre as sementes realizada com sucesso!!",
                "dados": data,
            }
            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            print("erro ao pegar os dados das Sementes", e)
            response = {
                "msg": f"Erro gerar os dados das Sementes",
                "error": f"Error ao pegar os dados das Sementes , Erro: {e}",  # General error message
            }
            return Response(response, status=status.HTTP_208_ALREADY_REPORTED)

    @action(detail=False, methods=["GET", "POST"])
    def get_plot_mapa_data_fetchrn_app(self, request, pk=None):
        if request.user.is_authenticated:
            try:
                safra_filter = "2025/2026"
                cicle_filter = "3"
            
                # 1️⃣  read – query params first, then body, finally default
                safra_filter = (
                    request.query_params.get("safra")
                    or request.data.get("safra")
                    or "2025/2026"                 # <- default
                )

                ciclo_filter = (
                    request.query_params.get("ciclo")
                    or request.data.get("ciclo")
                    or "3"                         # <- default
                )

                # farm can come from ?farm=, from body, or from the URL /plots/<pk>/
                farm_filter = (
                    request.query_params.get("farm")
                    or request.data.get("farm")
                    or None                          # pk passed in the URL
                )
                qs = (
                    Plantio.objects.filter(
                        safra__safra=safra_filter,
                        ciclo__ciclo=ciclo_filter,
                        # finalizado_plantio=True,
                        # plantio_descontinuado=False,
                        talhao__fazenda__id_farmbox=farm_filter
                    )
                    .filter(~Q(variedade__cultura__cultura="Milheto"))
                    .values(
                        "talhao__id_talhao",
                        "map_geo_points",
                        "map_centro_id"
                        
                    )
                )
                response = {
                    "msg": "Consulta realizada com sucesso!!",
                    "data": qs,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)

        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['GET'])
    def check_how_fast(self, request, *args, **kwargs):
        start_time = time.time()

        # Record processing start time
        process_start = time.time()

        response_data = { 'msg': 'good' }

        # Processing done, measure the time taken
        process_end = time.time()

        # Send the response
        end_time = time.time()

        # Log the breakdown
        print(f"Total Time Taken: {end_time - start_time:.4f} seconds")
        print(f"Processing Time: {process_end - process_start:.4f} seconds")
        print(f"Other Overheads: {end_time - process_end:.4f} seconds")

        # Return the response
        return Response(response_data, status=status.HTTP_200_OK)


class DefensivoViewSet(viewsets.ModelViewSet):
    queryset = Defensivo.objects.all().order_by("produto")
    serializer_class = DefensivoSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=True, methods=["POST"])
    def save_defensivo_data(self, request, pk=None):
        if request.user.is_authenticated:
            if "defensivos" in request.data:
                try:
                    file = request.FILES["defensivos"]
                    entradas = openpyxl.load_workbook(file, data_only=True)
                    worksheet = entradas["defensivos"]
                    for col in worksheet.iter_rows(min_row=1, max_col=5, max_row=3000):
                        if col[1].value != None and col[0].value != "ID":
                            produto = col[0].value.strip()
                            unidade_medida = col[1].value
                            formulacao = col[2].value
                            tipo = col[3].value

                            try:
                                new_product = Defensivo(
                                    produto=produto,
                                    unidade_medida=unidade_medida,
                                    formulacao=formulacao,
                                    tipo=tipo,
                                )
                                new_product.save()
                                print(
                                    f"Novo Defensivo Salvo com sucesso!! - {new_product.produto}"
                                )
                            except Exception as e:
                                print(f"Erro ao Salvar o produto {produto} - Erro: {e}")
                    qs_defensivos = Defensivo.objects.all()
                    serializer_defensivos = DefensivoSerializer(
                        qs_defensivos, many=True
                    )

                    response = {
                        "msg": f"Defensivos Atualizados com sucesso!!",
                        "total_return": len(qs_defensivos),
                        "dados": serializer_defensivos.data,
                    }
                    return Response(response, status=status.HTTP_200_OK)
                except Exception as e:
                    response = {"message": f"Ocorreu um Erro: {e}"}
                    return Response(response, status=status.HTTP_400_BAD_REQUEST)
            else:
                response = {"message": "Arquivo desconhecido"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


# --------------------- ---------------------- DEFENSIVOS API END  --------------------- ----------------------#


# --------------------- ---------------------- FARMBOX APPLICATIONS UPDATE API START  --------------------- ----------------------#
    
    @action(detail=False, methods=["GET"])
    def update_farmbox_mongodb_data(self, request, pk=None):
        
        number_of_days_before = 10 if DEBUG == True else 7
        from_date = get_date(number_of_days_before)
        last_up = get_miliseconds(from_date)
        print(last_up)
        
        number_of_days_before_pluvi = 5 if DEBUG == True else 4
        from_date_pluvi = get_date(number_of_days_before_pluvi)
        last_up_pluvi = get_miliseconds(from_date_pluvi)
        print(last_up_pluvi)
        
        
        data_applications = get_applications(updated_last=last_up)
        data_applications_pluvi = get_applications_pluvi(updated_last=last_up_pluvi)
        # print(data_applications)
        
        for _ in range(2):
            print(time.ctime())
            # Prints the current time with a five second difference
            time.sleep(1)
        with Spinner("Atualizando Aplicacoes..."):
            type_up_aplicacoes = 'Aplicacoes'
            generate_file_run(type_up_aplicacoes, data_applications)
            print("\nAplicações Atualizadas.")
        
        with Spinner("Atualizando Pluviometria..."):
            type_up_pluvi = 'Pluvi'
            generate_file_run(type_up_pluvi, data_applications_pluvi)
            print("\nPluviometrias Atualizadas.")
        
        
        
        response = {
            "msg": f"Banco de Dados atualizado com sucesso!!",
            "dados": 'dados do banco',
        }
        return Response(response, status=status.HTTP_200_OK)


# --------------------- ---------------------- FARMBOX APPLICATIONS UPDATE API END  --------------------- ----------------------#


# --------------------- ---------------------- DEFENSIVOS API END  --------------------- ----------------------#


# --------------------- ---------------------- PROGRAMS API START  --------------------- ----------------------#


class ProgramasDetails(viewsets.ModelViewSet):
    queryset = Aplicacao.objects.all()
    serializer_class = AplicacaoSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=False, methods=["GET", "POST"])
    def get_operacoes(self, request):
        if request.user.is_authenticated:
            try:
                safra_filter = None
                cicle_filter = None
                try:
                    safra_filter = request.data["safra"]
                    cicle_filter = request.data["ciclo"]
                except Exception as e:
                    print(e)
                print(safra_filter)
                print(cicle_filter)
                safra_filter = "2024/2025" if safra_filter == None else safra_filter
                cicle_filter = "1" if cicle_filter == None else cicle_filter
                qs = Aplicacao.objects.values(
                    "criados",
                    # "modificado",
                    "dose",
                    "obs",
                    "operacao__estagio",
                    "operacao__prazo_dap",
                    "operacao__programa__nome",
                    # "operacao__programa__nome_fantasia",
                    "operacao__programa__cultura__cultura",
                    "operacao__programa__safra__safra",
                    "operacao__programa__ciclo__ciclo",
                    "defensivo__produto",
                    "defensivo__tipo",
                ).filter(
                    Q(ativo=True)
                    & Q(operacao__programa__ativo=True)
                    & Q(operacao__ativo=True)
                ).filter(~Q(defensivo__tipo="operacao"))
                qs_estagios = Operacao.objects.values(
                    "estagio", "programa__nome", "prazo_dap", "obs"
                ).filter(Q(programa__ativo=True) & Q(ativo=True))
                qs_programas = (
                    Programa.objects.values(
                        "nome",
                        "nome_fantasia",
                        "safra__safra",
                        "ciclo__ciclo",
                        "versao",
                    )
                    .order_by("safra", "ciclo", "nome")
                    .filter(Q(ativo=True))
                )
                qs_area_total_program = (
                    Plantio.objects.values("programa__nome")
                    .annotate(total=Sum("area_colheita"))
                    .filter(~Q(programa=None) & Q(programa__ativo=True))
                )
                # serializer = AplicacaoSerializer(qs, many=True)
                response = {
                    "msg": f"Consulta realizada com sucesso!!",
                    "total_return": len(qs),
                    "dados": qs,
                    "estagios": qs_estagios,
                    "programas": qs_programas,
                    "area_total": qs_area_total_program,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


# --------------------- ---------------------- PROGRAMS API END  --------------------- ----------------------#


def remove_leading_zeros(num):
    return num.lstrip("0")


def adjust_parcelas(parcelas):
    list_parcelas = parcelas.replace("'", "").split(";")
    if len(list_parcelas) == 2:
        return list_parcelas[:-1]
    return parcelas.replace("'", "").split(";")[0:-1]


def adjust_percent_parcelas(percent):
    if len(percent) > 0:
        list_percent = percent.split(";")
        if len(list_percent) > 2:
            new_list = [float(x) / 100 for x in list_percent[0:-1]]
            return new_list
        else:
            return list_percent[0]
    else:
        return ""


class ColheitaApiSave(viewsets.ModelViewSet):
    queryset = Colheita.objects.all()
    serializer_class = ColheitaSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=True, methods=["GET"])
    def get_colheita_detail_react_native(self, request, pk=None):
        try:
            # qs = Colheita.objects.filter(plantio__id=pk).select_related("plantio", "deposito").prefetch_related("plantio__variedade", "plantio__talhao")
            qs = Colheita.objects.filter(plantio__id=pk).values(
                "data_colheita",
                "romaneio",
                "placa",
                "motorista",
                "ticket",
                "peso_tara",
                "peso_bruto",
                "umidade",
                "desconto_umidade",
                "impureza",
                "desconto_impureza",
                "peso_liquido",
                "peso_scs_liquido",
                "id_farmtruck",
            )
            # serializer = ColheitaResumoSerializer(qs, many=True)
            response ={
                "msg": 'Consulta realizada com sucesso!!',
                "data": qs
            }
            return Response(response, status=status.HTTP_201_CREATED)
        except Exception as e:
            print("Error here:", e)
            response = {
                "msg": "Erro ao pegar os dados",
                "error": f"Erro ao pegar os dados, Erro: {str(e)}",
            }
            return Response(response, status=status.HTTP_208_ALREADY_REPORTED)

    @action(detail=False, methods=["GET", "POST"])
    def save_from_protheus(self, request):
        user_id = Token.objects.get(user=request.user)
        if request.user.is_authenticated:
            data_json = request.data
            # date_file = request.data["plantio"]
            # data_json = json.load(date_file)
            plantio_query = Plantio.objects.all().filter(
                finalizado_plantio=True,
                finalizado_colheita=False,
                plantio_descontinuado=False,
            )
            deposito_query = Deposito.objects.all()
            succes = 0
            failed = 0
            problem = []
            success_list = []
            
            data_json = sorted(data_json, key=lambda x: (x["Projeto"], int(x["Ticket"])))
            
            for i in data_json:
                try:
                    if len(i["Placa do veiculo"]) > 0:
                        data = i["Data de Pesagem Tara"]
                        if "/" in data:
                            data = datetime.datetime.strptime(
                                data, "%d/%m/%Y"
                            ).strftime("%Y-%m-%d")
                        romaneio = remove_leading_zeros(str(i["Num Romaneio"]))
                        filial = i["Filial"]
                        ticket = remove_leading_zeros(str(i["Ticket"]))
                        placa = i["Placa do veiculo"]
                        motorista = i["Motorista"]
                        origem = i["Projeto"]
                        origem_id = remove_leading_zeros(str(i["Cod Projeto"]))
                        parcelas = adjust_parcelas(i["Parcela"])
                        peso_bruto = i["Peso Bruto"]
                        peso_tara = i["Peso Tara"]
                        peso_liquido = i["Peso Liquido"]
                        umidade = str(i["Umidade Entrada %"]).replace(',', '.')
                        impureza = str(i["Impureza Entrada %"]).replace(',', '.')
                        safra = i["Safra"]
                        ciclo = i["Ciclo"]
                        destino = i["Destino"]
                        percent_parcelas = adjust_percent_parcelas(
                            i["Percentual_Parcela"]
                        )
                        id_farmtruck = i["ID_Integracao"]

                        if "UBS" in str(destino):
                            destino = 2
                        elif "BADU" in str(destino):
                            destino = 10
                        elif "FAZENDAO" in str(destino):
                            destino = 4
                        elif "DIAMANTE" in str(destino):
                            destino = 1
                        elif "BIGUA" in str(destino):
                            destino = 3
                        elif "JK" in str(destino):
                            destino = 7

                        final_ticket = f"{filial}{ticket}"
                        print(i)
                        if len(parcelas) > 1:
                            for index, parcela in enumerate(parcelas):
                                if len(percent_parcelas) > 0:
                                    peso_bruto_considerado = int(
                                        peso_bruto * percent_parcelas[index]
                                    )
                                    peso_tara_considerado = int(
                                        peso_tara * percent_parcelas[index]
                                    )

                                else:
                                    if index + 1 == len(parcelas):
                                        peso_bruto_considerado = int(
                                            peso_bruto / len(parcelas)
                                        )
                                        peso_tara_considerado = int(
                                            peso_tara / len(parcelas)
                                        )
                                        if peso_bruto % len(parcelas) != 0:
                                            peso_bruto_ajuste = peso_bruto % len(
                                                parcelas
                                            )
                                            peso_bruto_considerado += peso_bruto_ajuste
                                        if peso_tara % len(parcelas) != 0:
                                            peso_tara_ajuste = peso_tara % len(parcelas)
                                            peso_tara_considerado += peso_tara_ajuste
                                peso_liquido_considerado = int(
                                    peso_bruto_considerado - peso_tara_considerado
                                )
                                carga = {
                                    "data": data,
                                    "romaneio": romaneio,
                                    "ticket": final_ticket,
                                    "placa": placa,
                                    "motorista": motorista,
                                    "origem": origem,
                                    "origem_id": origem_id,
                                    "parcela": parcela,
                                    "peso_bruto": peso_bruto_considerado,
                                    "peso_tara": peso_tara_considerado,
                                    "peso_liquido": peso_liquido_considerado,
                                    "umidade": umidade,
                                    "impureza": impureza,
                                    "safra": safra,
                                    "ciclo": ciclo,
                                    "destino": destino,
                                    "id_farmtruck": id_farmtruck,
                                }
                                print(
                                    "Nova Carga: ",
                                    f"{Fore.CYAN}{carga}{Style.RESET_ALL}",
                                )
                                try:
                                    plantio_id = plantio_query.get(
                                        safra__safra=safra,
                                        ciclo__ciclo=ciclo,
                                        talhao__fazenda__id_d=origem_id,
                                        talhao__id_talhao=parcela,
                                    )
                                    deposito_id = deposito_query.get(pk=destino)

                                    if plantio_id and deposito_id:
                                        try:
                                            #  Check if a record with the same ticket and plantio_id already exists
                                            if Colheita.objects.filter(ticket=final_ticket, plantio=plantio_id, romaneio=romaneio).exists():
                                                print(f"Duplicate record: ticket={final_ticket}, plantio={plantio_id}")
                                                raise ValueError("Carga já cadastrada para este ticket e plantio.")
                                            new_carga = Colheita(
                                                plantio=plantio_id,
                                                deposito=deposito_id,
                                                data_colheita=data,
                                                romaneio=romaneio,
                                                placa=placa,
                                                motorista=motorista,
                                                ticket=final_ticket,
                                                peso_tara=peso_tara_considerado,
                                                peso_bruto=peso_bruto_considerado,
                                                umidade=Decimal(umidade),
                                                impureza=Decimal(impureza),
                                                id_farmtruck=id_farmtruck,
                                            )
                                            new_carga.save()
                                            succes += 1
                                            print(
                                                f"{Fore.GREEN}Nova Carga incluida com sucesso: {new_carga}{Style.RESET_ALL}"
                                            )
                                            success_load = {
                                                "parcela": parcela,
                                                "projeto": origem,
                                                "romaneio": romaneio,
                                                "ticket": ticket,
                                                "id_farmtruck": id_farmtruck,
                                            }
                                            success_list.append(success_load)
                                        except Exception as e:
                                            print(
                                                f"Proglema em salvar a carga: {Fore.LIGHTRED_EX}{e}{Style.RESET_ALL}"
                                            )
                                            failed += 1
                                            problem_load = {
                                                "parcela": parcela,
                                                "projeto": origem,
                                                "romaneio": romaneio,
                                                "ticket": ticket if ticket else 'Sem Ticket',
                                                "error": str(e),
                                                "id_farmtruck": id_farmtruck,
                                            }

                                            problem.append(problem_load)
                                    print(f"{Fore.BLUE}{deposito_id}{Style.RESET_ALL}")
                                    print(f"{Fore.BLUE}{plantio_id}{Style.RESET_ALL}")
                                except Exception as e:
                                    print(
                                        f"plantio não encontrado - {Fore.LIGHTRED_EX}Origem: {origem} | Parcela: {parcela} | Erro: {e}{Style.RESET_ALL}"
                                    )
                                    failed += 1
                                    problem_load = {
                                        "parcela": parcela,
                                        "projeto": origem,
                                        "romaneio": romaneio,
                                        "ticket": ticket if ticket else 'Sem Ticket',
                                        "error": str(e),
                                        "id_farmtruck": id_farmtruck,
                                    }
                                    problem.append(problem_load)
                        else:
                            carga = {
                                "data": data,
                                "romaneio": romaneio,
                                "ticket": final_ticket,
                                "placa": placa,
                                "motorista": motorista,
                                "origem": origem,
                                "origem_id": origem_id,
                                "parcela": parcelas[0],
                                "peso_bruto": peso_bruto,
                                "peso_tara": peso_tara,
                                "peso_liquido": peso_liquido,
                                "umidade": umidade,
                                "impureza": impureza,
                                "safra": safra,
                                "ciclo": ciclo,
                                "destino": destino,
                                "id_farmtruck": id_farmtruck,
                            }
                            print(
                                "Nova Carga 1 parcela :",
                                f"{Fore.LIGHTCYAN_EX}{carga}{Style.RESET_ALL}",
                            )
                            try:
                                plantio_id = plantio_query.get(
                                    safra__safra=safra,
                                    ciclo__ciclo=ciclo,
                                    talhao__fazenda__id_d=origem_id,
                                    talhao__id_talhao=parcelas[0],
                                )
                                deposito_id = deposito_query.get(pk=destino)
                                if plantio_id and deposito_id:
                                    try:
                                        if Colheita.objects.filter(ticket=final_ticket, plantio=plantio_id, romaneio=romaneio).exists():
                                            print(f"Duplicate record: ticket={final_ticket}, plantio={plantio_id}")
                                            raise ValueError("Carga já cadastrada para este ticket e plantio.")
                                        new_carga = Colheita(
                                            plantio=plantio_id,
                                            deposito=deposito_id,
                                            data_colheita=data,
                                            romaneio=romaneio,
                                            placa=placa,
                                            motorista=motorista,
                                            ticket=final_ticket,
                                            peso_tara=peso_tara,
                                            peso_bruto=peso_bruto,
                                            umidade=Decimal(umidade),
                                            impureza=Decimal(impureza),
                                            id_farmtruck=id_farmtruck,
                                        )
                                        new_carga.save()
                                        succes += 1
                                        print(
                                            f"{Fore.GREEN}Nova Carga incluida com sucesso: {new_carga}{Style.RESET_ALL}"
                                        )
                                        success_load = {
                                            "parcela": parcelas[0],
                                            "projeto": origem,
                                            "romaneio": romaneio,
                                            "ticket": ticket,
                                            "id_farmtruck": id_farmtruck,
                                        }
                                        success_list.append(success_load)
                                    except Exception as e:
                                        print(
                                            f"Proglema em salvar a carga: {Fore.LIGHTRED_EX}{e}{Style.RESET_ALL}"
                                        )
                                        failed += 1
                                        problem_load = {
                                            "parcela": parcelas[0],
                                            "projeto": origem,
                                            "romaneio": romaneio,
                                            "ticket": ticket if ticket else 'Sem Ticket',
                                            "error": str(e),
                                            "id_farmtruck": id_farmtruck,
                                        }
                                        problem.append(problem_load)
                                print(f"{Fore.BLUE}{deposito_id}{Style.RESET_ALL}")
                                print(f"{Fore.BLUE}{plantio_id}{Style.RESET_ALL}")
                            except Exception as e:
                                print(
                                    f"plantio não encontrado - {Fore.LIGHTRED_EX}Origem: {origem} | Parcela: {parcelas[0]} | Erro: {e}{Style.RESET_ALL}"
                                )
                                failed += 1
                                problem_load = {
                                    "parcela": parcelas[0],
                                    "projeto": origem,
                                    "romaneio": romaneio,
                                    "ticket": ticket if ticket else 'Sem Ticket',
                                    "error": str(e),
                                    "id_farmtruck": id_farmtruck,
                                }
                                problem.append(problem_load)
                        print("\n")
                    else:
                        print("Carga sem Placa e para ser computada")
                        problem_load = {
                            "parcela": "SEM PLACA",
                            "projeto": "SEM PLACA",
                            "romaneio": remove_leading_zeros(str(i["Ticket"])),
                            "error": "SEM PLACA",
                            "ticket": "SEM TICKET",
                            "id_farmtruck": id_farmtruck,
                        }
                        problem.append(problem_load)
                except Exception as e:
                    print(
                        f"{Fore.LIGHTRED_EX}Erro ao salvar a carga, verificar os campos fornecidos: {e}{Style.RESET_ALL}"
                    )
            qs = Colheita.objects.all()
            serializer = ColheitaSerializer(qs, many=True)
            # TODO
            id_farmtruck_problem_list = list(set([x["id_farmtruck"] for x in problem if len(x["id_farmtruck"]) > 5]))
            id_farmtruck_list = list(set([x["id_farmtruck"] for x in success_list if len(x["id_farmtruck"]) > 5 and x["id_farmtruck"] not in id_farmtruck_problem_list]))
            print(
                "IDs Farmtruck incluido com sucesso: \n",
            )

            print(id_farmtruck_list)
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Token {user_id}",
            }
            try:
                response = requests.post(
                    f"{main_path_upload_ids}/romaneios/update-status-protheus-uploaded/",
                    headers=headers,
                    data=json.dumps(id_farmtruck_list),
                )
                print("response from send Ids:", response)
            except Exception as e:
                print(f"erro ao enviar os Ids para o servidor: {e}")
            try:
                response = {
                    "msg": f"Cadastro das Cargas efetuado com sucesso!!!",
                    "quantidade": len(serializer.data),
                    "data": {"includes": succes, "notincludes": failed},
                    "failed_load": problem,
                    "success_load": success_list,
                    # "data": serializer.data,
                }
                return Response(response, status=status.HTTP_200_OK)
            except Exception as e:
                response = {"message": f"Ocorreu um Erro: {e}"}
                return Response(response, status=status.HTTP_400_BAD_REQUEST)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


class VisitasConsultasApi(viewsets.ModelViewSet):
    queryset = Visitas.objects.all()
    serializer_class = VisitasSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @method_decorator(cache_page(300))
    @action(detail=False, methods=["GET"])
    def get_visitas(self, request):
        qs_registros = (
            RegistroVisitas.objects.all()
            .order_by("visita__id")
            .distinct("visita")
            .select_related("visita", "visita__fazenda")
        )
        serializer_registros = RegistroVisitasSerializer(qs_registros, many=True)

        qs_visitas = (
            Visitas.objects.all()
            .order_by("-id")
            .select_related("fazenda")
            .prefetch_related("projeto")
        )
        serializer_visitas = VisitasSerializer(qs_visitas, many=True)

        response = {
            "msg": "Consulta das Visitas realizada com sucesso!!",
            "data_registros": serializer_registros.data,
            "data_visitas": serializer_visitas.data,
        }
        return Response(response, status=status.HTTP_200_OK)


class RegistroVisitasApi(viewsets.ModelViewSet):
    queryset = RegistroVisitas.objects.all()
    serializer_class = RegistroVisitasSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=False, methods=["GET", "POST"])
    def get_registro_visita(self, request, pk=None):
        if request.method == "POST":
            id_filtered = None
            if request.data:
                print(request.data)
                id_filtered = request.data["idfilter"]
                print(id_filtered)
                qs = (
                    RegistroVisitas.objects.filter(visita_id=id_filtered).values(
                        "visita",
                        "image_title",
                        "obs",
                        "visita__fazenda__nome",
                        "visita__data",
                    )
                    # .select_related("visita", "visita__fazenda")
                )
                # qs = RegistroVisitas.objects.filter(
                #     visita_id=id_filtered
                # ).select_related("visita", "visita__fazenda")
                serialize = RegistroVisitasSerializer(qs, many=True)
                response = {
                    "msg": "Consulta da Visita e as informações Realizada com sucesso!!",
                    "data": qs,
                }
                return Response(response, status.HTTP_200_OK)
            else:
                response = {
                    "msg": "Não foi informado nenhuma visita para Filtro",
                    "data": [],
                }
            return Response(response, status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["GET", "POST"])
    def get_registro_visita_url(self, request, pk=None):
        if request.method == "POST":
            id_filtered = None
            if request.data:
                print(request.data)
                id_filtered = request.data["idfilter"]
                print(id_filtered)
                qs = RegistroVisitas.objects.filter(
                    visita_id=id_filtered
                ).select_related("visita", "visita__fazenda")
                # qs = RegistroVisitas.objects.filter(
                #     visita_id=id_filtered
                # ).select_related("visita", "visita__fazenda")
                serialize = RegistroVisitasSerializer(qs, many=True)
                response = {
                    "msg": "Consulta da Visita e as informações Realizada com sucesso!!",
                    "data": serialize.data,
                }
                return Response(response, status.HTTP_200_OK)
            else:
                response = {
                    "msg": "Não foi informado nenhuma visita para Filtro",
                    "data": [],
                }
            return Response(response, status.HTTP_400_BAD_REQUEST)


class PlantioDetailResumoApi(viewsets.ModelViewSet):
    queryset = PlantioDetail.objects.all()
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    @action(detail=False, methods=["GET"])
    def get_colheita_resumo(self, request):
        try:
            safra_filter = None
            cicle_filter = None
            try:
                print(request)
                safra_filter = request.data["safra"]
                cicle_filter = request.data["ciclo"]
                print('safra e ciclo passadas pela URL:')
                print(safra_filter, cicle_filter)
            except Exception as e:
                print(e)
            if safra_filter == None or cicle_filter == None:
                print("Sem Safra e/ou ciclo informados")
                safracicle_filter = CicloAtual.objects.filter(nome="Colheita")[0]
                safra_filter = safracicle_filter.safra.safra
                cicle_filter = safracicle_filter.ciclo.ciclo
            metrics = {
                "area_total": Sum("area_colheita"),
                "area_finalizada": Case(
                    When(
                        finalizado_colheita=True, then=Coalesce(Sum("area_colheita"), 0)
                    ),
                    When(
                        finalizado_colheita=False, then=Coalesce(Sum("area_parcial"), 0)
                    ),
                    default=Value(0),
                    output_field=DecimalField(),
                    # ),
                    # "area_parcial": Case(
                    #     When(finalizado_colheita=False, then=Coalesce(Sum("area_parcial"), 0)),
                    #     default=Value(0),
                    #     output_field=DecimalField(),
                ),
            }
            query_data = (
                PlantioDetail.objects.filter(
                    safra__safra=safra_filter,
                    ciclo__ciclo=cicle_filter,
                    finalizado_plantio=True,
                    plantio_descontinuado=False,
                )
                .filter(~Q(variedade__cultura__cultura="Milheto"))
                .values(
                    "talhao__fazenda__nome",
                    "variedade__cultura__cultura",
                    "variedade__variedade",
                )
                .annotate(**metrics)
                .order_by("talhao__fazenda__nome")
            )
            cargas = get_cargas_model(safra_filter, cicle_filter, Colheita)
            new_dict = []
            for i in query_data:
                filtered_dict = filter(
                    lambda obj: obj[1]["talhao__fazenda__nome"]
                    == i["talhao__fazenda__nome"]
                    and obj[1]["variedade__cultura__cultura"]
                    == i["variedade__cultura__cultura"]
                    and obj[1]["variedade__variedade"] == i["variedade__variedade"],
                    enumerate(new_dict),
                )
                if len(list(filtered_dict)) > 0:
                    filtered_dict = list(
                        filter(
                            lambda obj: obj[1]["talhao__fazenda__nome"]
                            == i["talhao__fazenda__nome"]
                            and obj[1]["variedade__cultura__cultura"]
                            == i["variedade__cultura__cultura"]
                            and obj[1]["variedade__variedade"]
                            == i["variedade__variedade"],
                            enumerate(new_dict),
                        )
                    )
                    index_find = filtered_dict[0][0]
                    area_total = new_dict[index_find]["area_total"] + i["area_total"]
                    area_finalizada = (
                        new_dict[index_find]["area_finalizada"] + i["area_finalizada"]
                    )
                    new_dict[index_find].update(
                        {"area_total": area_total, "area_finalizada": area_finalizada}
                    )
                else:
                    new_dict.append(i)

            for carga in cargas:
                filtered_dict = list(
                    filter(
                        lambda obj: obj[1]["talhao__fazenda__nome"]
                        == carga["plantio__talhao__fazenda__nome"]
                        and obj[1]["variedade__cultura__cultura"]
                        == carga["plantio__variedade__cultura__cultura"]
                        and obj[1]["variedade__variedade"]
                        == carga["plantio__variedade__variedade"],
                        enumerate(new_dict),
                    )
                )
                index_find = filtered_dict[0][0]
                peso_total = carga["peso_scs"]
                new_dict[index_find].update({"peso_kg": peso_total})

            response = {"msg": "Consulta realizada com sucesso!!", "data": new_dict}
            return Response(response, status=status.HTTP_200_OK)
        except Exception as e:
            response = {"msg": f"Ocorreu um Erro: {e}"}
            return Response(response, status.HTTP_400_BAD_REQUEST)


def format_input_numbers(input_list):
    def format_number(value):
        # rounded_value = str(round(value, 2)).replace('.', ',')
        rounded_value = round(value, 2)
        return rounded_value
    return [{**x, "quantidade": format_number(x['quantidade'])} for x in input_list]

def format_input_numbers_for_geral(input_list):
    def format_number(value):
        # rounded_value = str(round(value, 2)).replace('.', ',')
        rounded_value = round(value, 2)
        return rounded_value
    return [{**x, "totalQuantityOpen": format_number(x['totalQuantityOpen'])} for x in input_list]

def formart_ap_list(input_list):
    formated_list = [ x.split('|')[0].strip().replace('AP', 'AP ') for x in input_list]
    return formated_list

class StViewSet(viewsets.ModelViewSet):
    queryset = StProtheusIntegration.objects.all()
    serializer_class = StProtheusIntegrationSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    emails_list_by_farm = [
        {
            "projetos": ["Fazenda Benção de Deus"],
            "emails_abertura_st": [
                "matheus.silva@diamanteagricola.com.br",
                "clessio.batista@diamanteagricola.com.br",
            ],
        },
        {
            "projetos": ["Fazenda Cacique", "Fazenda Campo Guapo", "Fazenda Safira"],
            "emails_abertura_st": [
                "Willian.junior@diamanteagricola.com.br",
                "joao.neto@diamanteagricola.com.br"
                ],
        },
        {
            "projetos": [
                "Fazenda Capivara",
                "Fazenda Cervo",
                "Fazenda Jacaré",
                "Fazenda Tucano",
                "Fazenda Tuiuiu",
            ],
            "emails_abertura_st": [
                "lara.rodrigues@diamanteagricola.com.br",
                "jardel.mota@diamanteagricola.com.br"
            ],
        },
        {
            "projetos": [
                "Fazenda Lago Verde",
                "Fazenda Fazendinha",
                "Fazenda Santa Maria",
            ],
            "emails_abertura_st": [
                "marcelo.pata@diamanteagricola.com.br",
                "juliana.silva@diamanteagricola.com.br",
                "gisely.alencar@diamanteagricola.com.br",
            ],
        },
    ]

    def find_emails_by_farm(self, farm, emails_list_by_farm):
        for entry in emails_list_by_farm:
            if farm in entry["projetos"]:
                return entry["emails_abertura_st"]
        return None  # Return None if the farm is not found
    
    
    def save_st_from_protheus(self,st_number_protheus, projetos, req_data):
        try:
            print('Start saving st from protheus print')
            logger.info('Start saving st from protheus')
            new_st_opened = StProtheusIntegration(
                st_numero=st_number_protheus,
                st_fazenda=projetos[0],
                app=req_data
            )
            new_st_opened.save()
            logger.info(f'Nova ST salva com sucesso!! {new_st_opened}')
        except Exception as e:
            print('Problema em salvar a pré st do Protheus')
            logger.error("Error in save_st_from_protheus: %s", e, exc_info=True)

    @action(detail=False, methods=["GET", "POST"])
    def open_st_by_protheus(self, request, pk=None):

        # Define if should open pre st
        generate_pre_st = True

        # Define if should send email
        if generate_pre_st:
            should_send_email = False
        else:
            should_send_email = True

        list_emails = []
        if request.user.is_authenticated:
            req_data = None
            try:
                # get data from request
                req_data = request.data['dados_st']
                # string = response.read().decode('utf-8')
                req_data = json.loads(req_data)
            except Exception as e:
                print('erro ao pegar os dados', e)
                response = {
                    "msg": f"Erro ao Pegar os dados",
                    "error": f"Error ao pegar os dados vindos do farm, Erro: {e}"  # General error message
                }
                return Response(response, status=status.HTTP_208_ALREADY_REPORTED)
            print('dados recebidos: ', req_data)

            # should change after receive from protheus'api
            st_number_protheus = 0
            if req_data:
                # handle data and format properly
                projetos = req_data.get("Projeto")
                if projetos:
                    if generate_pre_st:
                        list_emails = self.find_emails_by_farm(projetos[0], self.emails_list_by_farm)
                    else:
                        list_emails = ["mtpata@icloud.com"]
                print('\n')
                print('should sent to: ', list_emails)
                print('\n')
                datas = req_data.get("Data")

                if req_data.get("Ap"):
                    apps = formart_ap_list(req_data.get("Ap"))

                if req_data.get('produtos'):
                    produtos = format_input_numbers(req_data.get('produtos'))

                print('produtos: ', produtos)
                fazenda_destino = req_data.get('fazendaDestino')
                armazem_destino  = req_data.get('armazemDestino')
                observacao  = req_data.get('observacao')

                produtos_geral = req_data.get('produtosGeral')
                if produtos_geral:
                    produtos_geral = format_input_numbers_for_geral(produtos_geral)

                headers = {
                    "Accept": "application/json",
                    "Content-Type": "application/json",
                    "Authorization": f'Basic ${PROTHEUS_TOKEN}',
                    "Access-Control-Allow-Origin": "*"
                }

                url = "https://api.diamanteagricola.com.br:8089/rest/apisolicitacao/new"
                payload = req_data

                try:
                    if generate_pre_st:
                        response_headers = requests.post(url, data=json.dumps(payload), headers=headers, verify=False, auth=HTTPBasicAuth('api', PROTHEUS_TOKEN))
                        print('response headers from protheus:', response_headers)
                        print('\n\n')
                        print('response:', response_headers.status_code, response_headers.text)
                        if response_headers.status_code == 201:
                            print('ST Aberta com sucesso!!!')
                            parsed_json = json.loads(response_headers.text)
                            number_opened = parsed_json["codigo_pre_st"]
                            st_number_protheus = int(number_opened)
                            should_send_email = True
                            
                            # START SAVE HERE
                            Thread(target=self.save_st_from_protheus, args=(st_number_protheus,projetos,req_data)).start()
                        else:
                            parsed_json = json.loads(response_headers.text)
                            response = {
                                "msg": f"Erro ao abrir a ST no Protheus - Code: {response_headers.status_code}",
                                "error": f"Error ao Abrir ST no Protheus, erro: {parsed_json['message']}"  # General error message
                            }
                            return Response(response, status=status.HTTP_208_ALREADY_REPORTED)
                except Exception as e:
                    print('Erro ao enviar a ST', e)

            try:
                # create context to send as response to frontend
                context = {
                    "fazendas": projetos,
                    "datas": datas,
                    "aplicacoes": apps,
                    "produtos": produtos,
                    'fazenda_destino': fazenda_destino,
                    'armazem_destino': armazem_destino,
                    'st_number': st_number_protheus,
                    'observacao': observacao,
                    'produtosGeral': produtos_geral
                }

                # send email function and logic
                df = pd.DataFrame(produtos_geral)
                df_prest = pd.DataFrame(produtos)

                # Keep only the desired columns and rename them
                df = df[['inputName', 'totalQuantityOpen']].rename(columns={
                    'inputName': 'Produto',
                    'totalQuantityOpen': 'Quantidade'
                })

                # Merge df and df_prest on inputName (renamed to Produto)
                df_prest = df_prest.rename(columns={"insumo": "Produto"})
                df_combined = pd.merge(df_prest, df, on="Produto", how="left")

                # Fill NaN values in 'Quantidade' with 0 for non-matching rows
                title_column = f'Pre ST {st_number_protheus}'

                df_combined['quantidade'] = df_combined['quantidade'].fillna(0)
                df_combined['Estoque'] = ''
                df_combined = df_combined[['Produto', 'Quantidade', 'quantidade', 'Estoque']].rename(columns={
                    "quantidade": title_column,
                    'Quantidade': 'Necessidade Geral'
                })

                print("Columns in df_combined:", df_combined.columns)

                print(df)
                print('\n')
                print(df_combined)

                if should_send_email == True:

                    # Path to the pre-existing Excel file in your Django project
                    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                    file_path = os.path.join(BASE_DIR, "diamante/files", "Produtos.xlsx")

                    # Load the existing Excel file
                    workbook = load_workbook(file_path)
                    sheet_name = "Produtos"

                    # Load the target sheet or create it if it doesn't exist
                    if sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                    else:
                        sheet = workbook.create_sheet(sheet_name)
                    if len(projetos) > 1:
                        farms_names = " - ".join(name.split("Fazenda ")[-1] for name in projetos)
                    else:
                        farms_names = projetos[0].split('Fazenda ')[-1]
                    sheet.cell(row=1, column=1, value=farms_names)
                    sheet.cell(row=2, column=1, value="Produto")
                    sheet.cell(row=2, column=2, value="Necessidade Geral")
                    sheet.cell(row=2, column=3, value=title_column)
                    sheet.cell(row=2, column=4, value="Estoque")
                    # Write DataFrame to the sheet (starting from row 2 to keep headers)
                    for i, row in df_combined.iterrows():
                        sheet.cell(row=i+1 + 2, column=1, value=row["Produto"])
                        # Write "Necessidade Geral" column as a number with thousand separator
                        necessidade_cell = sheet.cell(row=i+1 + 2, column=2, value=row["Necessidade Geral"])
                        necessidade_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                        # Write dynamic title column as a number with thousand separator
                        dynamic_column_cell = sheet.cell(row=i+1 + 2, column=3, value=row[title_column])
                        dynamic_column_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                        sheet.cell(row=i + 1 + 2, column=4, value=row["Estoque"])

                    sheet_name2 = "Produtos Geral"
                    if sheet_name2 not in workbook.sheetnames:
                        sheet2 = workbook.create_sheet(sheet_name2)
                    else:
                        sheet2 = workbook[sheet_name2]

                    # Write the second DataFrame to the new tab
                    for i, row in df.iterrows():
                        sheet2.cell(row=i + 2, column=1, value=row["Produto"])
                        # Write the "Quantidade" column as a number with thousand separator
                        quantity_cell = sheet2.cell(row=i + 2, column=2, value=row["Quantidade"])
                        quantity_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

                    # Save the updated workbook to memory
                    excel_file = BytesIO()
                    workbook.save(excel_file)
                    excel_file.seek(0)

                    subject = f"Pré ST Aberta: {st_number_protheus}"
                    from_email = 'patamarcelo@gmail.com'

                    if generate_pre_st:
                        cc_list = [
                            "raylton.sousa@diamanteagricola.com.br",
                            "adriana.goncalves@diamanteagricola.com.br",
                            "marcelo.pata@diamanteagricola.com.br",
                            "marim.neto@diamanteagricola.com.br"
                        ]
                    else:
                        cc_list = ["marcelo.pata@diamanteagricola.com.br"]

                    template_name = "email/st_open.html"
                    convert_to_html_content =  render_to_string(
                        template_name=template_name, context=context
                    )

                    # plain_message = strip_tags(convert_to_html_content)

                    email = EmailMultiAlternatives(
                        subject=subject,
                        body=convert_to_html_content,  # Set plain text version
                        from_email=from_email,
                        to=list_emails,
                        cc=cc_list or [],  # Add CC list (default to empty if not provided)
                    )

                    email.attach_alternative(convert_to_html_content, "text/html")  # Add HTML alternative
                    current_date = datetime.datetime.now().strftime("%d-%m-%Y")

                    # Attach the Excel file
                    email.attach(
                        f"produtos_geral-{current_date}_{title_column}.xlsx",
                        excel_file.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Attach HTML file
                    email.attach(
                        f"resumo_solicitacao-{current_date}.html",
                        convert_to_html_content,
                        "text/html"  # MIME type for HTML
                    )

                    result = email.send()
                    check_here = 'Sim' if result > 0 else 'Não'
                    print('Email foi enviado: ', check_here)

                else:
                    print('Email não enviado, devido as configurações')
                    check_here = 'Não'
            except Exception as e:
                print('erro ao pegar os dados e enviar por email: ', e)
                response = {
                            "msg": f"Erro ao enviar o E-mail",
                            "error": f"Error ao enviar o e-mail com os dados da Pré ST, Erro: {e}"  # General error message
                        }
                return Response(response, status=status.HTTP_208_ALREADY_REPORTED)

            response = {
                'msg': 'St Aberta com Successo',
                'st_number': st_number_protheus,
                'sent_by_email': check_here,
                'dados_recebidos': req_data
            }
            return Response(response, status=status.HTTP_201_CREATED)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)


class ColheitaPlantioExtratoAreaViewSet(viewsets.ModelViewSet):
    queryset = ColheitaPlantioExtratoArea.objects.all()
    serializer_class = StProtheusIntegrationSerializer
    authentication_classes = (CachedTokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    
    
    
    
    @action(detail=False, methods=["GET", "POST"])
    def update_colheita_area_from_farmbox(self, request, pk=None):
        if request.user.is_authenticated:
            req_data = None
            try:
                # get data from request
                req_data = request.data
            except Exception as e:
                print('erro ao pegar os dados', e)
            list_of_ids = [x['plantioId'] for x in req_data]
            filtered_query = Plantio.objects.filter(id_farmbox__in=list_of_ids)
            
            for i in req_data:
                try:
                    check_this = is_older_than_7_days(i['editado'])
                    if not check_this:
                        continue

                    print('check here:', i)

                    plantio_id_to_save = i['plantioId']
                    plantio_to_save = next(
                        (x for x in filtered_query if x.id_farmbox == plantio_id_to_save), None
                    )
                    if not plantio_to_save:
                        raise ValueError(f"Plantio with id_farmbox={plantio_id_to_save} not found")

                    area_to_save = Decimal(i['Area Aplicada'].replace(',', '.'))
                    data_to_save = i['Data Aplicacao']
                    hour_to_save, minute_to_save = map(int, i['Hora Aplicacao'].split(':'))
                    total_aplicado_to_save = Decimal(i['Total Aplicado'].replace(',', '.'))

                    plantio_to_save.area_parcial = total_aplicado_to_save
                    plantio_to_save.save()

                    # Wrap critical operation in a separate atomic block
                    with transaction.atomic():
                        new_colheita = ColheitaPlantioExtratoArea(
                            plantio=plantio_to_save,
                            area_colhida=area_to_save,
                            data_colheita=data_to_save,
                            time=dateTime(hour_to_save, minute_to_save)
                        )
                        new_colheita.save()

                    print(f'{Fore.GREEN}Colheita Salva com sucesso!!{Style.RESET_ALL}')
                    print(new_colheita)
                    print('\n')

                except Exception as e:
                    print(f'{Fore.LIGHTYELLOW_EX}Problema em Salvar o Plantio: {i} \n{Fore.LIGHTRED_EX}Error: {e} {Style.RESET_ALL}')

            
            response = {
                'msg': 'Colheita Atualizada com sucesso!!'
            }
            return Response(response, status=status.HTTP_201_CREATED)
        else:
            response = {"message": "Você precisa estar logado!!!"}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)
