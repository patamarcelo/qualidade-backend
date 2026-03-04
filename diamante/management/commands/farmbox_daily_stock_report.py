import os
import json
import ssl
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo
from urllib.parse import urlencode, urljoin
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from django.core.management.base import BaseCommand

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from diamante.gmail.gmail_api import send_mail_gmail_api
from diamante.models import EmailAberturaST


BASE_URL = "https://farmbox.cc"
API_BASE = f"{BASE_URL}/api/v1"
TZ = ZoneInfo("America/Sao_Paulo")


def make_ssl_context():
    insecure = (os.environ.get("FARMBOX_INSECURE_SSL") or "").strip().lower() in ("1", "true", "yes")
    if not insecure:
        return None
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


SSL_CONTEXT = make_ssl_context()


def http_get_json(url: str, headers: dict) -> dict:
    req = Request(url, headers=headers, method="GET")
    try:
        with urlopen(req, context=SSL_CONTEXT) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw)
    except HTTPError as e:
        body = e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else ""
        raise RuntimeError(f"HTTPError {e.code} em {url}\n{body}") from e
    except URLError as e:
        raise RuntimeError(f"URLError em {url}: {e}") from e


def fetch_all_pages_list(headers: dict, first_url: str, root_key: str) -> list:
    url = first_url
    out = []
    while True:
        data = http_get_json(url, headers=headers)
        out.extend(data.get(root_key) or [])
        next_url = data.get("next_page_url")
        if not next_url:
            break
        url = urljoin(BASE_URL, next_url)
    return out


def fetch_movimentations_range(headers: dict, start_date_iso: str, end_date_iso: str) -> list:
    """
    Busca movimentações no intervalo [start_date_iso, end_date_iso].
    A paginação é resolvida via next_page_url.
    """
    q = urlencode({
        "by_start_date": start_date_iso,
        "by_end_date": end_date_iso,
    })
    url = f"{API_BASE}/movimentations?{q}"
    return fetch_all_pages_list(headers, url, "movimentations")


def fetch_users_map(headers: dict) -> dict:
    users = fetch_all_pages_list(headers, f"{API_BASE}/users", "users")
    out = {}
    for u in users:
        uid = u.get("id")
        if uid is None:
            continue
        out[int(uid)] = (u.get("name") or "").strip() or f"#{uid}"
    return out


def fetch_inputs_map(headers: dict) -> dict:
    inputs = fetch_all_pages_list(headers, f"{API_BASE}/inputs", "inputs")
    out = {}

    for i in inputs:
        iid = i.get("id")
        if iid is None:
            continue

        input_type = (i.get("input_type_name") or "").strip().lower()

        # 🔴 remove "Operação"
        if input_type == "operação":
            continue

        name = (i.get("name") or "").strip()
        if not name:
            name = f"#{iid}"

        out[int(iid)] = name

    return out


def fetch_storages_map(headers: dict) -> dict:
    storages = fetch_all_pages_list(headers, f"{API_BASE}/storages", "storages")
    out = {}
    for s in storages:
        sid = s.get("id")
        if sid is None:
            continue
        out[int(sid)] = (s.get("name") or "").strip() or f"#{sid}"
    return out


def fmt_br_2(n: float) -> str:
    try:
        x = float(n or 0)
    except Exception:
        x = 0.0
    s = f"{x:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def date_br(iso_yyyy_mm_dd: str) -> str:
    try:
        return datetime.strptime(iso_yyyy_mm_dd, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return iso_yyyy_mm_dd or ""


def parse_iso_date(d: str) -> date:
    return datetime.strptime(d, "%Y-%m-%d").date()


def enrich_rows(rows, users_map, inputs_map, storages_map):
    """
    Normaliza:
    - resolve nomes via maps
    - regra: toda saída sem AP => application_code = 'sem ap'
    """
    enriched = []
    for r in rows:
        input_id = r.get("input_id")
        storage_id = r.get("storage_id")
        user_id = r.get("user_id")

        app_info = r.get("application_info") or {}
        batch = r.get("batch_info") or {}

        mov_type = (r.get("movimentation_type") or "").strip().lower()

        application_id = app_info.get("application_id")
        application_code = (app_info.get("code") or "").strip() or None

        if mov_type == "out" and not application_code:
            application_code = "sem ap"

        enriched.append({
            "id": r.get("id"),
            "input_id": input_id,
            "storage_id": storage_id,
            "user_id": user_id,

            "product_name": inputs_map.get(int(input_id), f"#{input_id}") if input_id is not None else "",
            "storage_name": storages_map.get(int(storage_id), f"#{storage_id}") if storage_id is not None else "",
            "user_name": users_map.get(int(user_id), f"#{user_id}") if user_id is not None else "",

            "date": r.get("date"),
            "movimentation_type": mov_type,
            "quantity": float(r.get("quantity") or 0),
            "unit": r.get("unit") or "",

            "observation": r.get("observation"),
            "invoice_number": r.get("invoice_number"),
            "invoice_file_url": r.get("invoice_file_url"),

            "batch_id": batch.get("batch_id"),
            "batch_number": batch.get("batch_number"),
            "validity": batch.get("validity"),

            "application_id": application_id,
            "application_code": application_code,
        })
    return enriched


def aggregate_by_product(enriched):
    by_product = {}  # (product, unit) -> {in,out}
    sum_in = 0.0
    sum_out = 0.0

    for r in enriched:
        t = (r.get("movimentation_type") or "").lower().strip()
        qty = float(r.get("quantity") or 0)
        prod = r.get("product_name") or "(sem produto)"
        unit = r.get("unit") or ""

        by_product.setdefault((prod, unit), {"in": 0.0, "out": 0.0})

        if t == "in":
            sum_in += qty
            by_product[(prod, unit)]["in"] += qty
        elif t == "out":
            sum_out += qty
            by_product[(prod, unit)]["out"] += qty

    prod_rows = []
    for (prod, unit), v in by_product.items():
        prod_rows.append({
            "product": prod,
            "unit": unit,
            "in": v["in"],
            "out": v["out"],
            "balance": v["in"] - v["out"],
        })
    prod_rows.sort(key=lambda x: (x["out"], x["product"]), reverse=True)

    return {
        "sum_in": sum_in,
        "sum_out": sum_out,
        "balance": sum_in - sum_out,
        "products": prod_rows,
    }


def build_email_html(start_iso: str, end_iso: str, agg: dict, enriched: list):
    range_label = date_br(start_iso) if start_iso == end_iso else f"{date_br(start_iso)} → {date_br(end_iso)}"

    # =========================
    # Consolidado por fazenda -> (input_id, produto, unidade)
    # =========================
    farms = {}  # farm -> (input_id, prod, unit) -> {in,out}

    for r in enriched:
        farm = r.get("storage_name") or "(sem fazenda)"
        mov_type = (r.get("movimentation_type") or "").lower().strip()
        qty = float(r.get("quantity") or 0)

        input_id = r.get("input_id")
        prod = r.get("product_name") or "(sem produto)"
        unit = r.get("unit") or ""

        key = (input_id, prod, unit)

        farms.setdefault(farm, {})
        farms[farm].setdefault(key, {"in": 0.0, "out": 0.0})

        if mov_type == "in":
            farms[farm][key]["in"] += qty
        elif mov_type == "out":
            farms[farm][key]["out"] += qty

    farm_blocks = []
    for farm_name, items_map in farms.items():
        rows = []
        total_out_farm = 0.0
        total_in_farm = 0.0

        for (input_id, prod, unit), v in items_map.items():
            outv = float(v["out"] or 0)
            inv = float(v["in"] or 0)
            if outv <= 0:
                continue  # somente produtos com saída

            total_out_farm += outv
            total_in_farm += inv

            rows.append({
                "date": start_iso,    # diário
                "input_id": input_id,
                "product": prod,
                "unit": unit,
                "in": inv,
                "out": outv,
                "balance": inv - outv,
            })

        if not rows:
            continue

        rows.sort(key=lambda x: (x["out"], x["product"]), reverse=True)

        farm_blocks.append({
            "farm": farm_name,
            "total_in": total_in_farm,
            "total_out": total_out_farm,
            "balance": total_in_farm - total_out_farm,
            "rows": rows,
        })

    farm_blocks.sort(key=lambda x: (x["total_out"], x["farm"]), reverse=True)

    # =========================
    # Consolidado geral -> (input_id, produto, unidade)  ✅ COM ID
    # =========================
    products_map = {}  # (input_id, prod, unit) -> {in,out}

    for r in enriched:
        t = (r.get("movimentation_type") or "").lower().strip()
        qty = float(r.get("quantity") or 0)

        input_id = r.get("input_id")
        prod = r.get("product_name") or "(sem produto)"
        unit = r.get("unit") or ""

        key = (input_id, prod, unit)
        products_map.setdefault(key, {"in": 0.0, "out": 0.0})

        if t == "in":
            products_map[key]["in"] += qty
        elif t == "out":
            products_map[key]["out"] += qty

    products_out = []
    for (input_id, prod, unit), v in products_map.items():
        outv = float(v["out"] or 0)
        inv = float(v["in"] or 0)
        if outv <= 0:
            continue
        products_out.append({
            "date": start_iso,
            "input_id": input_id,
            "product": prod,
            "unit": unit,
            "in": inv,
            "out": outv,
            "balance": inv - outv,
        })

    products_out.sort(key=lambda x: (x["out"], x["product"]), reverse=True)

    def tr_prod_row(p):
        iid = p.get("input_id")
        iid_txt = str(iid) if iid not in (None, "") else ""
        date_txt = date_br(p.get("date") or "")

        return f"""
          <tr>
            <td style="padding:6px 8px;border-bottom:1px solid #e9e9ef;white-space:nowrap;width:70px;">{date_txt}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;">{p["product"]}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(p["in"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(p["out"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(p["balance"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;">{p["unit"]}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;color:#666;">{iid_txt}</td>
          </tr>
        """

    def tr_farm_row(r):
        iid = r.get("input_id")
        iid_txt = str(iid) if iid not in (None, "") else ""
        date_txt = date_br(r.get("date") or "")

        return f"""
          <tr>
            <td style="padding:6px 8px;border-bottom:1px solid #e9e9ef;white-space:nowrap;width:70px;">{date_txt}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;">{r["product"]}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(r["in"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(r["out"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;">{fmt_br_2(r["balance"])}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;">{r["unit"]}</td>
            <td style="padding:6px 10px;border-bottom:1px solid #e9e9ef;text-align:right;color:#666;">{iid_txt}</td>
          </tr>
        """

    def render_farm_block(fb):
        return f"""
          <table style="margin-top:16px; border-collapse:collapse; width:100%; max-width:980px; border:1px solid #e9e9ef;">
            <thead>
              <tr style="background:#eef1ff;">
                <th colspan="7" style="text-align:left; padding:10px 12px; font-size:15px; font-weight:900; border-bottom:1px solid #e9e9ef;">
                  Almoxarifado {fb["farm"]}
                </th>
              </tr>

              <tr style="background:#f5f6ff;">
                <th colspan="7" style="text-align:left; padding:8px 12px; font-weight:600; color:#444; border-bottom:1px solid #e9e9ef;">
                  Totais — IN: {fmt_br_2(fb["total_in"])} |
                  OUT: {fmt_br_2(fb["total_out"])} |
                  Saldo: {fmt_br_2(fb["balance"])}
                </th>
              </tr>

              <tr style="background:#f7f7fb;">
                <th style="text-align:left;padding:8px 8px;border-bottom:1px solid #e9e9ef;width:70px;">Data</th>
                <th style="text-align:left;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Produto</th>
                <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">IN</th>
                <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">OUT</th>
                <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Saldo</th>
                <th style="text-align:left;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Un</th>
                <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">ID</th>
              </tr>
            </thead>

            <tbody>
              {''.join(tr_farm_row(r) for r in fb["rows"])}
            </tbody>
          </table>
        """

    farms_html = "".join(render_farm_block(fb) for fb in farm_blocks)

    html = f"""
    <div style="font-family:Arial, Helvetica, sans-serif; color:#111;">
      <h2 style="margin:0 0 8px;">Relatório de estoque — {range_label}</h2>

      <div style="display:flex; gap:12px; flex-wrap:wrap; margin:10px 0 14px;">
        <div style="padding:10px 12px; border:1px solid #e9e9ef; border-radius:10px;">
          <div style="font-size:12px; color:#666;">Entrada (IN)</div>
          <div style="font-size:18px; font-weight:800;">{fmt_br_2(agg["sum_in"])}</div>
        </div>
        <div style="padding:10px 12px; border:1px solid #e9e9ef; border-radius:10px;">
          <div style="font-size:12px; color:#666;">Saída (OUT)</div>
          <div style="font-size:18px; font-weight:800;">{fmt_br_2(agg["sum_out"])}</div>
        </div>
        <div style="padding:10px 12px; border:1px solid #e9e9ef; border-radius:10px;">
          <div style="font-size:12px; color:#666;">Saldo (IN - OUT)</div>
          <div style="font-size:18px; font-weight:800;">{fmt_br_2(agg["balance"])}</div>
        </div>
      </div>

      <h3 style="margin:18px 0 8px;">Consolidado por produto (todos com OUT)</h3>
      <table style="border-collapse:collapse; width:100%; max-width:980px; border:1px solid #e9e9ef;">
        <thead>
          <tr style="background:#f7f7fb;">
            <th style="text-align:left;padding:8px 8px;border-bottom:1px solid #e9e9ef;width:70px;">Data</th>
            <th style="text-align:left;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Produto</th>
            <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">IN</th>
            <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">OUT</th>
            <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Saldo</th>
            <th style="text-align:left;padding:8px 10px;border-bottom:1px solid #e9e9ef;">Un</th>
            <th style="text-align:right;padding:8px 10px;border-bottom:1px solid #e9e9ef;">ID</th>
          </tr>
        </thead>
        <tbody>
          {''.join(tr_prod_row(p) for p in products_out) or '<tr><td colspan="7" style="padding:10px;">Sem saídas no período.</td></tr>'}
        </tbody>
      </table>

      <h3 style="margin:22px 0 8px;">Consolidado por fazenda (produtos com OUT)</h3>
      {farms_html or '<div style="color:#666;">Sem saídas no período.</div>'}

      <p style="color:#666; font-size:12px; margin-top:14px;">
        Anexo: Excel com extrato completo (RAW) + SAÍDAS com AP (aba SAIDAS_AP) + consolidado por produto.
      </p>
    </div>
    """
    return html

def autosize(ws, max_col=30):
    for col in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def write_table(ws, headers, rows):
    bold = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append(r)
    autosize(ws, max_col=len(headers))


def build_xlsx_bytes(start_iso: str, end_iso: str, enriched: list, agg: dict) -> bytes:
    wb = Workbook()

    # RAW
    ws_raw = wb.active
    ws_raw.title = "RAW"

    raw_headers = [
        "id", "date_iso", "date_br", "movimentation_type", "quantity", "unit",
        "product_name", "input_id",
        "storage_name", "storage_id",
        "user_name", "user_id",
        "application_code", "application_id",
        "batch_id", "batch_number", "validity",
        "invoice_number", "invoice_file_url",
        "observation",
    ]
    raw_rows = []
    for r in enriched:
        raw_rows.append([
            r.get("id"),
            r.get("date"),
            date_br(r.get("date") or ""),
            r.get("movimentation_type"),
            float(r.get("quantity") or 0),
            r.get("unit"),
            r.get("product_name"),
            r.get("input_id"),
            r.get("storage_name"),
            r.get("storage_id"),
            r.get("user_name"),
            r.get("user_id"),
            r.get("application_code"),
            r.get("application_id"),
            r.get("batch_id"),
            r.get("batch_number"),
            r.get("validity"),
            r.get("invoice_number"),
            r.get("invoice_file_url"),
            r.get("observation"),
        ])
    write_table(ws_raw, raw_headers, raw_rows)
    for row in ws_raw.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "#,##0.00"

    # SAIDAS_AP
    ws_out = wb.create_sheet("SAIDAS_AP")
    out_headers = [
        "id", "date_iso", "date_br",
        "storage_name", "product_name",
        "quantity", "unit",
        "application_code", "application_id",
        "user_name",
        "observation",
    ]
    out_rows = []
    for r in enriched:
        if (r.get("movimentation_type") or "") != "out":
            continue
        out_rows.append([
            r.get("id"),
            r.get("date"),
            date_br(r.get("date") or ""),
            r.get("storage_name"),
            r.get("product_name"),
            float(r.get("quantity") or 0),
            r.get("unit"),
            (r.get("application_code") or "sem ap"),
            r.get("application_id"),
            r.get("user_name"),
            r.get("observation"),
        ])
    write_table(ws_out, out_headers, out_rows)
    for row in ws_out.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = "#,##0.00"

    # Consolidado_Produto
    ws_p = wb.create_sheet("Consolidado_Produto")
    p_headers = ["product", "unit", "in", "out", "balance"]
    p_rows = [[x["product"], x["unit"], x["in"], x["out"], x["balance"]] for x in agg["products"]]
    write_table(ws_p, p_headers, p_rows)
    for col in (3, 4, 5):
        for row in ws_p.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = "#,##0.00"

    # META
    ws_meta = wb.create_sheet("META")
    write_table(ws_meta, ["key", "value"], [
        ["range_start_iso", start_iso],
        ["range_end_iso", end_iso],
        ["generated_at", datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S %z")],
        ["rows_total", len(enriched)],
        ["rows_out", len([r for r in enriched if (r.get("movimentation_type") or "") == "out"])],
    ])

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


class Command(BaseCommand):
    help = "Coleta movimentações na Farmbox e envia e-mail via Gmail API (HTML + XLSX)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--since",
            type=str,
            default=None,
            help="Data inicial ISO (YYYY-MM-DD). Se informado, busca do since até hoje.",
        )

    def handle(self, *args, **options):
        token = os.environ.get("FARM_API")
        if not token:
            raise RuntimeError("FARM_API não configurado (token Farmbox).")

        # to_list = (os.environ.get("STOCK_REPORT_EMAIL_TO") or "marcelo@gdourado.com.br").split(",")
        # to_list = [x.strip() for x in to_list if x.strip()]
        to_list = EmailAberturaST.objects.filter(atividade__tipo='Estoque Movimento Farmbox', ativo=True).values_list('email', flat=True)
        if not to_list:
            raise RuntimeError("STOCK_REPORT_EMAIL_TO não configurado (lista separada por vírgula).")

        cc_list = (os.environ.get("STOCK_REPORT_EMAIL_CC") or "").split(",")
        cc_list = [x.strip() for x in cc_list if x.strip()]

        from_email = os.environ.get("STOCK_REPORT_EMAIL_FROM") or "me"

        now = datetime.now(TZ)
        today_iso = now.date().strftime("%Y-%m-%d")

        since = (options.get("since") or "").strip() or None

        if since:
            # valida formato
            try:
                _ = parse_iso_date(since)
            except Exception:
                raise RuntimeError("Formato inválido em --since. Use YYYY-MM-DD.")

            start_iso = since
            end_iso = today_iso
        else:
            # padrão: ontem (um único dia)
            start_iso = (now.date() - timedelta(days=1)).strftime("%Y-%m-%d")
            end_iso = start_iso

        headers = {
            "Authorization": token,
            "Accept": "application/json",
            "User-Agent": "farmbox-daily-stock-report/1.0",
        }

        self.stdout.write(f"[1/5] Coletando movimentações {start_iso} → {end_iso}...")
        rows = fetch_movimentations_range(headers, start_date_iso=start_iso, end_date_iso=end_iso)
        self.stdout.write(f"  OK: {len(rows)} linhas (antes de filtros)")

        # garantia extra do range (caso API ignore algum limite)
        try:
            d0 = parse_iso_date(start_iso)
            d1 = parse_iso_date(end_iso)
            rows = [r for r in rows if r.get("date") and d0 <= parse_iso_date(r["date"]) <= d1]
        except Exception:
            pass

        self.stdout.write("[2/5] Carregando índices (inputs/users/storages)...")
        users_map = fetch_users_map(headers)
        inputs_map = fetch_inputs_map(headers)
        storages_map = fetch_storages_map(headers)

        # remove inputs filtrados (Operação)
        rows = [r for r in rows if r.get("input_id") in inputs_map]

        self.stdout.write("[3/5] Enriquecendo + consolidando por produto...")
        enriched = enrich_rows(rows, users_map, inputs_map, storages_map)
        enriched = [r for r in enriched if r.get("input_id") in inputs_map]

        # regra extra: saída sem AP => "sem ap"
        for r in enriched:
            if r.get("movimentation_type") == "out" and not (r.get("application_code") or "").strip():
                r["application_code"] = "sem ap"

        agg = aggregate_by_product(enriched)

        self.stdout.write("[4/5] Gerando Excel...")
        xlsx_bytes = build_xlsx_bytes(start_iso, end_iso, enriched, agg)
        xlsx_name = f"estoque_movimentacoes_{start_iso}_ate_{end_iso}.xlsx" if start_iso != end_iso else f"estoque_movimentacoes_{start_iso}.xlsx"

        self.stdout.write("[5/5] Enviando via Gmail API...")
        subject_range = date_br(start_iso) if start_iso == end_iso else f"{date_br(start_iso)} → {date_br(end_iso)}"
        subject = f"Relatório de estoque (Farmbox) — {subject_range}"
        html = build_email_html(start_iso, end_iso, agg, enriched)

        attachments = [
            (xlsx_name, xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        ]

        send_mail_gmail_api(
            subject=subject,
            body_html=html,
            from_email=from_email,
            to_emails=to_list,
            cc_emails=cc_list,
            attachments=attachments,
            fail_silently=False,
        )

        self.stdout.write(self.style.SUCCESS(f"Enviado para: {', '.join(to_list)}"))