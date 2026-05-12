from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from maquinario.models import Machine


class Command(BaseCommand):
    help = "Importa máquinas da planilha inicial da Fazenda Benção de Deus."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Caminho do arquivo .xlsx",
        )

        parser.add_argument(
            "--fazenda-id",
            required=True,
            type=int,
            help="ID da fazenda no app diamante.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options["file"]
        fazenda_id = options["fazenda_id"]

        try:
            workbook = load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Erro ao abrir planilha: {exc}")

        sheet = workbook.active

        header_row = None
        headers = {}

        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() if value is not None else "" for value in row]

            if "Identificador" in values and "Descrição" in values:
                header_row = row_index
                headers = {
                    value: index
                    for index, value in enumerate(values)
                    if value
                }
                break

        if not header_row:
            raise CommandError("Não encontrei a linha de cabeçalho com Identificador/Descrição.")

        identifier_idx = headers.get("Identificador")
        chassis_idx = headers.get("Chassi")
        description_idx = headers.get("Descrição")
        location_idx = headers.get("Restrito a localidades")
        status_idx = headers.get("Status")

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            identifier = row[identifier_idx] if identifier_idx is not None else None
            description = row[description_idx] if description_idx is not None else None

            if not identifier or not description:
                skipped_count += 1
                continue

            identifier = str(identifier).strip()
            description = str(description).strip()

            chassis = None
            if chassis_idx is not None and row[chassis_idx]:
                chassis = str(row[chassis_idx]).strip()

            location_text = None
            if location_idx is not None and row[location_idx]:
                location_text = str(row[location_idx]).strip()

            status_text = ""
            if status_idx is not None and row[status_idx]:
                status_text = str(row[status_idx]).strip().lower()

            is_active = True
            if status_text in ["inativo", "inactive", "bloqueado"]:
                is_active = False

            machine, created = Machine.objects.update_or_create(
                fazenda_id=fazenda_id,
                identifier=identifier,
                defaults={
                    "chassis": chassis,
                    "description": description,
                    "machine_type": Machine.MachineType.TRACTOR,
                    "status": Machine.Status.OPERATION,
                    "location_text": location_text,
                    "is_active": is_active,
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Importação finalizada. Criadas: {created_count}. "
                f"Atualizadas: {updated_count}. Ignoradas: {skipped_count}."
            )
        )